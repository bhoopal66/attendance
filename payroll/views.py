"""
Payroll calculation views.
"""

import io
import json
import datetime
import calendar
import logging
from collections import defaultdict
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as _get_col_letter

from django.core.management import call_command
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required, user_passes_test

from attendance.models import (
    Employee, RemoteEmployee, Holiday,
    LeaveRequest, MonthlySummary, RemoteMonthlySummary, AnnualLeave,
    AttendanceRecord, RemoteCallRecord,
)
from collections import defaultdict as _defaultdict
from attendance.views.utils import (
    MONTH_CHOICES, MONTH_NAMES, YEAR_RANGE,
    get_selected_month_year, section_required,
    get_active_special_periods_for_month, get_remote_thresholds_from_period,
    get_bridge_sunday_days,
    SALES_PERFORMANCE_V2_START, compute_sales_performance_v2_days,
)
from .models import PayrollAdjustment, Bank, BankSubmission, DeductionEntry, DeductionCarryover, GeneratedDocument, ExchangeRate, FrozenPayrollMonth, PaidSalaryRecord, CommissionTierSettings, DEDUCTION_CATEGORY_CHOICES, PAYMENT_METHOD_CHOICES, OTHER_DEDUCTION_CATEGORIES, PayrollNote
from .services import get_effective_salary_structure
from .services import convert_employee_deduction_currency

FOREIGN_CURRENCIES = ('INR', 'NPR')  # non-AED currencies with dashboard-level differentiation

logger = logging.getLogger('payroll')


# ============================================
# Payroll Calculation Helpers
# ============================================

def _count_holidays(year, month, days_in_month):
    """Return total non-working days (Sundays + custom holidays) for a month.
    Uses a set to avoid double-counting Sundays that are also custom holidays.
    """
    non_working = set(
        day for day in range(1, days_in_month + 1)
        if datetime.date(year, month, day).weekday() == 6
    )
    for holiday in Holiday.objects.filter(date__year=year, date__month=month):
        non_working.add(holiday.date.day)
    return len(non_working)


def _count_holidays_in_period(period_start, period_end):
    """Return total non-working days (Sundays + custom holidays) in a date range."""
    non_working = set()
    curr = period_start
    while curr <= period_end:
        if curr.weekday() == 6:
            non_working.add(curr)
        curr += datetime.timedelta(days=1)
    for holiday in Holiday.objects.filter(date__gte=period_start, date__lte=period_end):
        non_working.add(holiday.date)
    return len(non_working)


def _get_employee_pay_period(cycle_start_day, selected_year, selected_month):
    """Return (period_start, period_end, days_in_period, total_holidays) for a given pay cycle."""
    if cycle_start_day == 1:
        period_start = datetime.date(selected_year, selected_month, 1)
        last_day = calendar.monthrange(selected_year, selected_month)[1]
        period_end = datetime.date(selected_year, selected_month, last_day)
    else:
        prev_year = selected_year - 1 if selected_month == 1 else selected_year
        prev_month = 12 if selected_month == 1 else selected_month - 1
        prev_month_last = calendar.monthrange(prev_year, prev_month)[1]
        start_day = min(cycle_start_day, prev_month_last)
        period_start = datetime.date(prev_year, prev_month, start_day)
        period_end = datetime.date(selected_year, selected_month, cycle_start_day - 1)
    days_in_period = (period_end - period_start).days + 1
    total_holidays = _count_holidays_in_period(period_start, period_end)
    return period_start, period_end, days_in_period, total_holidays


def _leave_days_in_month(leave, month_start, month_end):
    """Return the effective paid leave days that fall within the given month.
    Pro-rates approved_days based on how much of the leave span falls in the month.
    """
    effective = leave.get_effective_days()
    if not effective:
        return 0
    # If entirely within the month, no pro-rating needed
    if leave.start_date >= month_start and leave.end_date <= month_end:
        return effective
    # Clip to month boundaries and pro-rate
    overlap_start = max(leave.start_date, month_start)
    overlap_end = min(leave.end_date, month_end)
    full_span = (leave.end_date - leave.start_date).days + 1
    overlap_span = (overlap_end - overlap_start).days + 1
    return round(effective * overlap_span / full_span)


def _annual_leave_day_counts(al, month_start, month_end):
    """Return (total_days, 0) for the annual leave overlap with the month.

    total_days — all calendar days in the overlap (Mon–Sun, including holidays).
                 MonthlySummary.leave_days now includes Sundays/holidays within
                 AnnualLeave periods, so the base deduction already covers them.
                 The second element is always 0; it exists only for call-site
                 compatibility and has no financial effect.
    """
    overlap_start = max(al.start_date, month_start)
    overlap_end = min(al.end_date, month_end)
    if overlap_end < overlap_start:
        return 0, 0

    total_days = (overlap_end - overlap_start).days + 1
    return total_days, 0


def _sales_annual_leave_compensation(emp, emp_type, year, month, month_start, month_end, daily_rate):
    """Annual leave compensation for sales fixed-salary / attendance-based employees.

    The sales absent_days calculation uses total_working_days (excluding Sundays/holidays),
    so compensation must also count only working days within each leave period.
    """
    holiday_dates = set(
        Holiday.objects.filter(
            date__year=year, date__month=month
        ).values_list('date', flat=True)
    )
    if emp_type == 'inhouse':
        al_qs = AnnualLeave.objects.filter(
            employee=emp, start_date__lte=month_end, end_date__gte=month_start
        )
    else:
        al_qs = AnnualLeave.objects.filter(
            remote_employee=emp, start_date__lte=month_end, end_date__gte=month_start
        )

    compensation = 0.0
    leave_working_days = 0
    for al in al_qs:
        overlap_start = max(al.start_date, month_start)
        overlap_end = min(al.end_date, month_end)
        curr = overlap_start
        working_days = 0
        while curr <= overlap_end:
            if curr.weekday() != 6 and curr not in holiday_dates:
                working_days += 1
            curr += datetime.timedelta(days=1)
        leave_working_days += working_days
        salary_pct = float(al.salary_percentage) if al.is_paid else 0.0
        compensation += daily_rate * working_days * salary_pct / 100.0

    return round(compensation, 2), leave_working_days


# Legacy fallback if a foreign currency has no CommissionTierSettings row yet.
DEFAULT_TIER_THRESHOLD = 4
DEFAULT_TIER_OVERFLOW_RATE = Decimal('3000')


def _get_tier_settings(currency):
    """Return (threshold, overflow_rate) for a foreign currency's tiered commission.
    Falls back to the legacy hardcoded INR default if nothing is configured yet."""
    setting = CommissionTierSettings.objects.filter(currency=currency).first()
    if setting:
        return setting.threshold, setting.overflow_rate
    return DEFAULT_TIER_THRESHOLD, DEFAULT_TIER_OVERFLOW_RATE


def _calc_tiered_commission(pairs, threshold, overflow_rate):
    """Tiered commission for a foreign currency (INR, NPR, ...).

    pairs: list of (count, bank_rate) in processing order (alphabetical by bank).
    - Accounts 1-threshold: bank's per-account charge in that currency.
    - Each account beyond threshold: overflow_rate flat.

    Returns (total_commission_float, per_pair_commission_list).
    """
    total = Decimal('0')
    used = 0
    per_pair = []
    for count, rate in pairs:
        rate = Decimal(str(rate))
        if used >= threshold:
            commission = count * overflow_rate
        elif used + count <= threshold:
            commission = count * rate
            used += count
        else:
            within = threshold - used
            overflow = count - within
            commission = within * rate + overflow * overflow_rate
            used = threshold
        per_pair.append(float(commission))
        total += commission
    return float(total), per_pair


def _get_commission(year, month, employee=None, remote_employee=None, currency='AED'):
    """Calculate total commission from bank submissions for the period.
    Foreign-currency employees (INR, NPR, ...) use tiered pricing per CommissionTierSettings:
    first `threshold` accounts at the bank's rate for that currency, then a flat overflow rate."""
    if employee:
        submissions = list(BankSubmission.objects.filter(
            employee=employee, year=year, month=month
        ).select_related('bank').order_by('bank__name'))
    else:
        submissions = list(BankSubmission.objects.filter(
            remote_employee=remote_employee, year=year, month=month
        ).select_related('bank').order_by('bank__name'))

    if currency != 'AED':
        threshold, overflow_rate = _get_tier_settings(currency)
        pairs = [(s.submission_count, s.bank.charge_for_currency(currency)) for s in submissions]
        total, _ = _calc_tiered_commission(pairs, threshold, overflow_rate)
        return total

    return float(sum(
        s.submission_count * s.bank.per_account_charge
        for s in submissions
    ))


def _get_inhouse_payroll_row(emp, year, month, month_start, month_end, total_holidays, days_in_period=None):
    """Build a payroll data row for one in-house employee."""
    _cross_month = (month_start.month != month_end.month or month_start.year != month_end.year)
    _holiday_dates = set()

    if _cross_month and days_in_period is not None:
        # Cross-month period (e.g. 21st–20th): MonthlySummary is calendar-month only,
        # so compute attendance directly from AttendanceRecord.
        _holiday_dates = set(Holiday.objects.filter(
            date__gte=month_start, date__lte=month_end
        ).values_list('date', flat=True))
        _approved_leave_dates = set()
        for lr in LeaveRequest.objects.filter(
            employee=emp, status='approved',
            start_date__lte=month_end, end_date__gte=month_start,
        ):
            _c = max(lr.start_date, month_start)
            _e = min(lr.end_date, month_end)
            while _c <= _e:
                if _c.weekday() != 6 and _c not in _holiday_dates:
                    _approved_leave_dates.add(_c)
                _c += datetime.timedelta(days=1)
        _recs = {r.date: r for r in AttendanceRecord.objects.filter(
            employee=emp, date__gte=month_start, date__lte=month_end
        )}
        full_days = half_days = late_days = absent_days = 0
        _c = month_start
        while _c <= month_end:
            if _c.weekday() == 6 or _c in _holiday_dates:
                _c += datetime.timedelta(days=1)
                continue
            _r = _recs.get(_c)
            if _r and (_r.first_in or _r.is_work_from_home):
                if _r.is_work_from_home or emp.is_fixed_salary:
                    full_days += 1
                elif _r.first_in > datetime.time(12, 0):
                    half_days += 1
                else:
                    full_days += 1
            elif _r and _r.is_paid_leave:
                pass  # admin-marked paid leave — not deducted
            elif _c not in _approved_leave_dates:
                absent_days += 1
            _c += datetime.timedelta(days=1)
        # Sundays/holidays inside AnnualLeave spans are handled exclusively by
        # annual_leave_extra_deduction below (charges (100 − salary_pct)% per day).
        # They are intentionally NOT added to absent_days here to avoid double-counting.
        effective_work_days = full_days + half_days * 0.5
        days_in_month = days_in_period
    else:
        summary = MonthlySummary.objects.filter(
            employee=emp, year=year, month=month
        ).first()
        if summary:
            half_days = summary.half_days or 0
            full_days = summary.working_days - half_days
            effective_work_days = full_days + (half_days * 0.5)
            absent_days = summary.leave_days or 0
            late_days = summary.late_days or 0
        else:
            full_days = 0
            half_days = 0
            effective_work_days = 0.0
            absent_days = 0
            late_days = 0
        days_in_month = calendar.monthrange(year, month)[1]

    salary = float(emp.salary) if emp.salary else 0.0

    # Phase 5 fix: use the employee's real approved SalaryStructure breakdown
    # (Basic/Housing/Transport/Phone/Other) instead of a synthetic percentage
    # split of the flat gross salary. Employee.salary stays the source of
    # truth for the net-pay math below (it's kept in sync with the structure's
    # gross by the Salary tab save handler) — only this display breakdown
    # changes source.
    _salary_structure = get_effective_salary_structure(emp, month_end)
    has_salary_structure = _salary_structure is not None
    if _salary_structure:
        basic_salary = float(_salary_structure.basic)
        housing_allowance = float(_salary_structure.housing)
        transport_allowance = float(_salary_structure.transport)
        # Phone is no longer a salary component (Phase E9 — the business does
        # not provide a phone allowance). The column is gone from every
        # breakdown, but any amount still sitting in the field is folded into
        # Other rather than dropped: silently discarding it would make the
        # components stop summing to the gross salary. The model field is
        # deliberately left in place so nothing is destroyed and the component
        # can be reinstated if that ever changes.
        phone_allowance = 0.0
        other_allowance_amt = float(_salary_structure.other_allowance) + float(_salary_structure.phone)
    else:
        # No approved SalaryStructure on file for this employee — flag it
        # (surfaces on the dashboard row and the exception centre) rather
        # than silently guessing a split.
        basic_salary = 0.0
        housing_allowance = 0.0
        transport_allowance = 0.0
        phone_allowance = 0.0
        other_allowance_amt = 0.0

    approved_leaves = LeaveRequest.objects.filter(
        employee=emp,
        status='approved',
        start_date__lte=month_end,
        end_date__gte=month_start,
    )
    paid_leave_days = sum(_leave_days_in_month(leave, month_start, month_end) for leave in approved_leaves)

    daily_rate = salary / days_in_month if salary > 0 else 0.0
    # Every 3 late days = 1 half-day deduction
    late_half_days = late_days // 3

    # Sandwich-rule unpaid Sundays: a Sunday between an approved-leave Sat
    # and an approved-leave Mon is treated as unpaid leave (one daily_rate off).
    bridge_sunday_count = len(get_bridge_sunday_days(emp, month_start, month_end))

    # Performance-based payroll: no late/leave deductions
    if getattr(emp, 'payroll_type', 'attendance') == 'performance':
        total_deduction_days = 0
        deduction = 0.0
    else:
        # Deduct absent days, half-day shortfalls, late penalties, and bridge Sundays
        total_deduction_days = (
            absent_days + (half_days * 0.5) + (late_half_days * 0.5) + bridge_sunday_count
        )
        deduction = daily_rate * total_deduction_days
    base_payroll = salary - deduction

    adjustments = PayrollAdjustment.objects.filter(employee=emp, year=year, month=month)
    incentives = float(
        adjustments.filter(adjustment_type='incentive').aggregate(total=Sum('amount'))['total'] or 0
    )
    reductions = float(
        adjustments.filter(adjustment_type='reduction').aggregate(total=Sum('amount'))['total'] or 0
    )
    commission = _get_commission(year, month, employee=emp, currency=emp.currency)

    # Annual leave adjustment:
    # - Paid leave at X%: compensate X% of daily_rate × working_leave_days
    #   (offsets the absent-day deduction already applied via base_payroll)
    # - Unpaid leave: no compensation for working days (deduction stands),
    #   but also deduct Sundays + holidays within the leave period because
    #   those days are normally paid — on unpaid leave they should not be.
    annual_leaves = AnnualLeave.objects.filter(
        employee=emp,
        start_date__lte=month_end,
        end_date__gte=month_start,
    )
    annual_leave_compensation = 0.0
    annual_leave_extra_deduction = 0.0
    annual_leave_days = 0
    for al in annual_leaves:
        if _cross_month:
            # Cross-month path: MonthlySummary is not used, so manually count
            # working vs non-working days within the leave overlap.
            # _holiday_dates is already populated from the attendance loop above.
            al_overlap_start = max(al.start_date, month_start)
            al_overlap_end = min(al.end_date, month_end)
            al_working = 0
            al_non_working = 0
            if al_overlap_end >= al_overlap_start:
                _d = al_overlap_start
                while _d <= al_overlap_end:
                    if _d.weekday() == 6 or _d in _holiday_dates:
                        al_non_working += 1
                    else:
                        al_working += 1
                    _d += datetime.timedelta(days=1)
            working_days, non_working_days = al_working, al_non_working
        else:
            working_days, non_working_days = _annual_leave_day_counts(al, month_start, month_end)
        annual_leave_days += working_days
        salary_pct = float(al.salary_percentage) if al.is_paid else 0.0
        # Working days: compensate salary_pct% of daily rate (offsets absent-day deduction)
        annual_leave_compensation += daily_rate * working_days * salary_pct / 100.0
        # Non-working days (Sundays/holidays): normally paid at 100%, during leave
        # only paid at salary_pct% — deduct the remaining (100 - salary_pct)%
        annual_leave_extra_deduction += daily_rate * non_working_days * (100.0 - salary_pct) / 100.0

    net_payroll = base_payroll + annual_leave_compensation - annual_leave_extra_deduction + incentives + commission - reductions

    return {
        'employee': emp,
        'employee_type': 'inhouse',
        'currency': emp.currency,
        'salary': salary,
        'basic_salary': basic_salary,
        'housing_allowance': housing_allowance,
        'transport_allowance': transport_allowance,
        'phone_allowance': phone_allowance,
        'other_allowance_amt': other_allowance_amt,
        'has_salary_structure': has_salary_structure,
        'full_days': full_days,
        'half_days': half_days,
        'effective_work_days': round(effective_work_days, 1),
        'absent_days': absent_days,
        'late_days': late_days,
        'late_half_days': late_half_days,
        'paid_leave_days': paid_leave_days,
        'bridge_sunday_count': bridge_sunday_count,
        'annual_leave_days': annual_leave_days,
        'annual_leave_compensation': round(annual_leave_compensation, 2),
        'annual_leave_extra_deduction': round(annual_leave_extra_deduction, 2),
        'total_deduction_days': round(total_deduction_days, 1),
        'daily_rate': round(daily_rate, 2),
        'deduction': round(deduction, 2),
        'base_payroll': round(base_payroll, 2),
        'incentives': round(incentives, 2),
        'reductions': round(reductions, 2),
        'commission': round(commission, 2),
        'net_payroll': round(net_payroll, 2),
    }


def _attach_gross_breakdown(row, period_end):
    """Populate Basic/Housing/Transport/Phone/Other on a payroll row dict
    (Phase E5). In-house employees use their real approved SalaryStructure —
    the same source the individual Payslip page already draws from. Remote
    employees have no SalaryStructure model at all, so fall back to the same
    40/60 Basic/Other estimate already used for remote payslips elsewhere,
    and only when the row actually has a salary figure to split — flagged
    is_estimated so the template can mark it clearly as not itemized.
    Rows with neither (no approved in-house structure, or no salary concept
    at all — e.g. pure-commission Sales Performance rows) get
    has_salary_structure=False so the template shows a dash instead of a
    fabricated number.
    """
    emp = row.get('employee')
    salary = row.get('salary')
    if row.get('employee_type') == 'inhouse' and emp is not None:
        structure = get_effective_salary_structure(emp, period_end)
        if structure:
            row['basic_salary'] = float(structure.basic)
            row['housing_allowance'] = float(structure.housing)
            row['transport_allowance'] = float(structure.transport)
            # Phase E9 — phone folded into Other; see _get_inhouse_payroll_row.
            row['phone_allowance'] = 0.0
            row['other_allowance_amt'] = float(structure.other_allowance) + float(structure.phone)
            row['has_salary_structure'] = True
            row['is_estimated'] = False
        else:
            row['basic_salary'] = 0.0
            row['housing_allowance'] = 0.0
            row['transport_allowance'] = 0.0
            row['phone_allowance'] = 0.0
            row['other_allowance_amt'] = 0.0
            row['has_salary_structure'] = False
            row['is_estimated'] = False
        return
    if salary:
        row['basic_salary'] = round(salary * 0.40, 2)
        row['housing_allowance'] = 0.0
        row['transport_allowance'] = 0.0
        row['phone_allowance'] = 0.0
        row['other_allowance_amt'] = round(salary * 0.60, 2)
        row['has_salary_structure'] = True
        row['is_estimated'] = True
    else:
        row['basic_salary'] = 0.0
        row['housing_allowance'] = 0.0
        row['transport_allowance'] = 0.0
        row['phone_allowance'] = 0.0
        row['other_allowance_amt'] = 0.0
        row['has_salary_structure'] = False
        row['is_estimated'] = False


def _get_sales_payroll_row(emp, year, month, emp_type, banks, days_in_month=None, total_holidays=0, period_start=None, period_end=None):
    """Build a payroll data row for a sales employee (commission-only, no attendance deductions).
    Fixed salary employees get attendance-based salary instead of commission."""
    if emp_type == 'inhouse':
        submissions_qs = BankSubmission.objects.filter(employee=emp, year=year, month=month)
        adjustments = PayrollAdjustment.objects.filter(employee=emp, year=year, month=month)
    else:
        submissions_qs = BankSubmission.objects.filter(remote_employee=emp, year=year, month=month)
        adjustments = PayrollAdjustment.objects.filter(remote_employee=emp, year=year, month=month)

    emp_currency = emp.currency if hasattr(emp, 'currency') else 'AED'
    bank_counts = {s.bank_id: s.submission_count for s in submissions_qs}
    bank_counts_list = [bank_counts.get(b.id, 0) for b in banks]

    if emp_currency != 'AED':
        threshold, overflow_rate = _get_tier_settings(emp_currency)
        pairs = [(bank_counts.get(b.id, 0), b.charge_for_currency(emp_currency)) for b in banks]
        commission, _ = _calc_tiered_commission(pairs, threshold, overflow_rate)
    else:
        commission = sum(
            bank_counts.get(b.id, 0) * float(b.per_account_charge)
            for b in banks
        )

    incentives = float(
        adjustments.filter(adjustment_type='incentive').aggregate(total=Sum('amount'))['total'] or 0
    )
    reductions = float(
        adjustments.filter(adjustment_type='reduction').aggregate(total=Sum('amount'))['total'] or 0
    )

    # display_type is based on location field (same logic as the Salary Setup page)
    if emp_type == 'inhouse':
        display_type = 'inhouse'
    else:
        display_type = 'inhouse' if (emp.location and emp.location.lower() == 'inhouse') else 'remote'

    # Fixed salary: attendance-based salary, bank counts kept for record only.
    # Present days are computed on-the-fly from raw records (a single punch / call
    # on a workday counts as Present) so the calculation reflects the *current*
    # fixed-salary rule even if stored summaries / attendance_status fields are
    # stale from before the toggle was flipped.
    if emp.is_fixed_salary:
        if days_in_month is None:
            days_in_month = calendar.monthrange(year, month)[1]
        salary = float(emp.salary) if emp.salary else 0.0
        daily_rate = salary / days_in_month if salary > 0 else 0.0
        total_working_days = days_in_month - total_holidays

        _ms = period_start or datetime.date(year, month, 1)
        _me = period_end or datetime.date(year, month, days_in_month)
        if emp_type == 'inhouse':
            present_days = AttendanceRecord.objects.filter(
                employee=emp, date__gte=_ms, date__lte=_me,
            ).filter(
                Q(first_in__isnull=False) | Q(is_work_from_home=True) | Q(is_paid_leave=True)
            ).count()
        else:
            call_records = RemoteCallRecord.objects.filter(
                employee=emp, date__gte=_ms, date__lte=_me,
            ).only('answered_calls', 'no_answered', 'busy', 'failed')
            present_days = sum(
                1 for r in call_records
                if (r.answered_calls or 0) + (r.no_answered or 0)
                   + (r.busy or 0) + (r.failed or 0) > 0
            )
        bridge_sunday_count = len(get_bridge_sunday_days(emp, _ms, _me))

        absent_days = max(0, total_working_days - present_days) + bridge_sunday_count
        deduction = daily_rate * absent_days
        base_salary = round(salary - deduction, 2)
        al_compensation, al_days = _sales_annual_leave_compensation(
            emp, emp_type, year, month, _ms, _me, daily_rate
        )
        net_payroll = round(base_salary + al_compensation + incentives - reductions, 2)

        return {
            'employee': emp,
            'employee_type': emp_type,
            'display_type': display_type,
            'currency': emp.currency,
            'is_fixed_salary': True,
            'salary': round(salary, 2),
            'daily_rate': round(daily_rate, 2),
            'absent_days': absent_days,
            'deduction': round(deduction, 2),
            'annual_leave_compensation': al_compensation,
            'annual_leave_days': al_days,
            'base_salary': base_salary,
            'bank_counts_list': bank_counts_list,
            'commission': round(commission, 2),
            'incentives': round(incentives, 2),
            'reductions': round(reductions, 2),
            'net_payroll': net_payroll,
        }

    # Attendance-based: salary scales with actual attendance, plus any commission.
    # In-house: any punch-in counts as present (half-day detection lives in the
    # report layer). Remote: present/half/absent are computed on-the-fly from
    # raw talk-time, mirroring the calendar view logic: special period thresholds
    # are applied per-day when an active SpecialShiftPeriod covers that date,
    # otherwise the defaults (weekday 45/90, friday 30/60, saturday 21/45) are
    # used. The stored RemoteCallRecord.attendance_status is NOT consulted because
    # it can be stale: if the employee was ever toggled to is_fixed_salary, every
    # record was saved with status 'present'/'absent' only (no 'half_day'), and
    # toggling back doesn't re-trigger save(). Computing inline is also independent
    # of the is_fixed_salary flag, which is the right behavior for an
    # attendance-based payroll. A 'half_day' counts as 0.5 of a worked day.
    # Sundays + custom holidays are excluded from both the count and the
    # working-day denominator.
    if getattr(emp, 'payroll_type', 'attendance') == 'attendance' and emp.salary:
        if emp_type == 'remote' and (year, month) >= SALES_PERFORMANCE_V2_START:
            _ms = period_start or datetime.date(year, month, 1)
            _me = period_end or datetime.date(year, month, calendar.monthrange(year, month)[1])
            _days = days_in_month if days_in_month is not None else (_me - _ms).days + 1
            v2 = _get_sales_performance_test_row(emp, _ms, _me, _days, total_holidays, year=year, month=month)
            base_salary = v2['net_payroll_test']
            deduction = round(v2['salary'] - base_salary, 2)
            net_payroll = round(base_salary + incentives - reductions, 2)
            return {
                'employee': emp,
                'employee_type': emp_type,
                'display_type': display_type,
                'currency': emp.currency,
                'is_fixed_salary': False,
                'is_attendance_based': True,
                'payroll_method': 'v2',
                'salary': v2['salary'],
                'daily_rate': v2['daily_rate'],
                'present_days': v2['full_days'],
                'half_days': v2['half_days'],
                'proportional_days': v2['proportional_days'],
                'non_working_days': v2['non_working_days'],
                'new_joiner_days': v2['new_joiner_days'],
                'grace_denials': v2['grace_denials'],
                'target_achieved': v2['target_achieved'],
                'absent_days': v2['leave_days'],
                'deduction': deduction,
                'annual_leave_compensation': 0.0,
                'annual_leave_days': 0,
                'base_salary': base_salary,
                'bank_counts_list': bank_counts_list,
                'commission': round(commission, 2),
                'incentives': round(incentives, 2),
                'reductions': round(reductions, 2),
                'net_payroll': net_payroll,
            }
        if days_in_month is None:
            days_in_month = calendar.monthrange(year, month)[1]
        salary = float(emp.salary)
        daily_rate = salary / days_in_month if salary > 0 else 0.0
        total_working_days = days_in_month - total_holidays

        _ms = period_start or datetime.date(year, month, 1)
        _me = period_end or datetime.date(year, month, days_in_month)
        bridge_sunday_count = len(get_bridge_sunday_days(emp, _ms, _me))
        if emp_type == 'inhouse':
            present_days = AttendanceRecord.objects.filter(
                employee=emp, date__gte=_ms, date__lte=_me,
            ).filter(
                Q(first_in__isnull=False) | Q(is_work_from_home=True)
            ).count()
            half_days = 0
        else:
            holiday_dates = set(
                Holiday.objects.filter(
                    date__gte=_ms, date__lte=_me
                ).values_list('date', flat=True)
            )
            call_records = RemoteCallRecord.objects.filter(
                employee=emp, date__gte=_ms, date__lte=_me,
            ).only('date', 'total_talk_duration')
            special_periods = get_active_special_periods_for_month(_ms, _me)
            present_days = 0
            half_days = 0
            for r in call_records:
                wd = r.date.weekday()
                if wd == 6 or r.date in holiday_dates:
                    continue
                if not r.total_talk_duration:
                    continue
                minutes = r.total_talk_duration.total_seconds() / 60
                # Apply special period thresholds when one covers this day,
                # falling back to defaults — mirrors the calendar view logic.
                active_period = next(
                    (p for p in special_periods if p.start_date <= r.date <= p.end_date),
                    None,
                ) if special_periods else None
                period_thresholds = get_remote_thresholds_from_period(active_period) if active_period else None
                t = period_thresholds or {}
                if wd == 5:
                    half_min, present_min = t.get('saturday', RemoteCallRecord.DEFAULT_THRESHOLDS['saturday'])
                elif wd == 4:
                    half_min, present_min = t.get('friday', RemoteCallRecord.DEFAULT_THRESHOLDS['friday'])
                else:
                    half_min, present_min = t.get('weekday', RemoteCallRecord.DEFAULT_THRESHOLDS['weekday'])
                if minutes >= present_min:
                    present_days += 1
                elif minutes >= half_min:
                    half_days += 1

        al_compensation, al_days = _sales_annual_leave_compensation(
            emp, emp_type, year, month, _ms, _me, daily_rate
        )
        effective_present = present_days + (half_days * 0.5)
        if effective_present == 0 and al_compensation == 0:
            # No attendance and no paid leave at all: deduct the entire calendar
            # month (including Sundays/holidays) so net salary is zero rather
            # than paying the non-working-day portion of the daily rate.
            absent_days = float(days_in_month)
        else:
            absent_days = max(0, total_working_days - effective_present) + bridge_sunday_count
        deduction = daily_rate * absent_days
        base_salary = round(salary - deduction, 2)
        net_payroll = round(base_salary + al_compensation + incentives - reductions, 2)

        return {
            'employee': emp,
            'employee_type': emp_type,
            'display_type': display_type,
            'currency': emp.currency,
            'is_fixed_salary': False,
            'is_attendance_based': True,
            'salary': round(salary, 2),
            'daily_rate': round(daily_rate, 2),
            'present_days': present_days,
            'half_days': half_days,
            'absent_days': round(absent_days, 1),
            'deduction': round(deduction, 2),
            'annual_leave_compensation': al_compensation,
            'annual_leave_days': al_days,
            'base_salary': base_salary,
            'bank_counts_list': bank_counts_list,
            'commission': round(commission, 2),
            'incentives': round(incentives, 2),
            'reductions': round(reductions, 2),
            'net_payroll': net_payroll,
        }

    net_payroll = commission + incentives - reductions
    return {
        'employee': emp,
        'employee_type': emp_type,
        'display_type': display_type,
        'is_fixed_salary': False,
        'currency': emp.currency,
        'bank_counts_list': bank_counts_list,
        'commission': round(commission, 2),
        'incentives': round(incentives, 2),
        'reductions': round(reductions, 2),
        'net_payroll': round(net_payroll, 2),
    }


def _get_sales_performance_test_row(emp, period_start, period_end, days_in_period, total_holidays, year=None, month=None):
    """EXPERIMENTAL — "Method 2" daily pay for remote Sales:Performance employees.
    Not used anywhere outside the /payroll/test/ 'Sales: Performance Test' section.

    Implements a per-day regime + grace-day gate model:

      DAILY_WAGE = ROUND(salary / days_in_period)   # days_in_period = total
                   calendar days in the pay period (Sundays/holidays included)
      HALF_DAY_PAY = ROUND(DAILY_WAGE / 2)

    Regime routing (Step 0 of the spec):
      - Friday/Saturday -> FRIDAY_SATURDAY regime (always, overrides everything else)
      - Monday-Thursday, the day falls within a rolling 30-day window starting
        on the employee's `joining_date` (day 0 through day 29 inclusive) ->
        NEW_JOINER_MONTH_1 regime. A calendar-month window would shortchange
        someone who joins near month-end (e.g. the 29th) to just a couple of
        lenient days.
      - Monday-Thursday otherwise -> STANDARD regime
    Sundays and custom holidays are outside all regimes and paid at the full
    daily wage, same as the live payroll sections, so the two figures stay
    comparable — except when the day falls inside an AnnualLeave span, in
    which case it's paid at the leave's own salary_percentage instead (0%
    for unpaid leave), matching the in-house annual_leave_extra_deduction
    behavior.

    FRIDAY_SATURDAY regime — binary, no half/proportional zone:
      <30 min talktime  -> absent (0 pay)
      >=30 min talktime -> full day (100% of daily wage)

    NEW_JOINER_MONTH_1 regime (Mon-Thu only, no grace band):
      <45 min           -> absent (0 pay)
      45-59 min (incl.)  -> half day (50% of daily wage)
      >=60 min           -> full day (100% of daily wage)

    STANDARD regime (Mon-Thu, Month 2+):
      <45 min             -> absent (0 pay)
      45-54 min (incl.)   -> half day (50% of daily wage)
      55-89 min (incl.)   -> grace-band: subject to the three gates below
      >=90 min             -> full day (100% of daily wage)

    Grace-band (55-89 min) gates, checked in order — any failure downgrades
    the day to a half day instead of prorata pay, and does not consume a
    grace slot:
      1. Monthly cap: employee has already used 7 grace days in this
         calendar month
      2. Weekly cap: employee has already used 2 grace days in this
         Mon-Sun work-week
      3. Consecutive cap: the employee's last 2 STANDARD-regime workdays
         were both grace/prorata days
    If all three gates pass: PRORATA pay = ROUND((minutes/90) * daily_wage),
    and the monthly/weekly/consecutive counters all increment. The
    consecutive-day counter resets to 0 on any STANDARD full/half/absent day,
    and is left untouched by Friday/Saturday and NEW_JOINER_MONTH_1 days.

    DATA LIMITATION: the source spec expects per-call records (call_id,
    called_number, duration_seconds, is_internal, crm_disposition_entered)
    to derive a filtered "TVM" (total valid minutes) before classification —
    excluding short calls, internal calls, calls beyond 1 re-dial per number,
    and calls missing CRM disposition. This system only stores daily
    aggregates (`RemoteCallRecord.total_talk_duration`), not individual call
    legs, so that filtering step cannot be reproduced here: TVM is taken
    directly from the stored total talk duration, same as it was in the
    original Method 2. Likewise, there is no stored per-day incentive field,
    so the spec's incentive step is not modeled (always 0).

    TARGET OVERRIDE (foreign-currency employees only): if the employee's total
    BankSubmission count across all banks for `year`/`month` reaches their
    currency's CommissionTierSettings threshold (4 by default),
    they've hit the account target and are paid full salary for the period
    regardless of attendance — the day-by-day breakdown is still computed and
    returned for informational purposes, but net_payroll_test is overridden
    to the full salary amount.
    """
    STANDARD_FULL_DAY_THRESHOLD = 90

    salary = float(emp.salary) if emp.salary else 0.0
    # Kept at full decimal precision (not rounded to a whole rupee) so a day's
    # pay never drifts from salary / days_in_period — rounding only happens
    # once, on the final total below.
    daily_wage = (salary / days_in_period) if (salary > 0 and days_in_period) else 0
    half_day_pay = daily_wage / 2

    holiday_dates = set(
        Holiday.objects.filter(date__gte=period_start, date__lte=period_end)
        .values_list('date', flat=True)
    )
    # Day classification (regime + full/half/proportional/leave) lives in
    # attendance.views.utils.compute_sales_performance_v2_days, shared with
    # the calendar/report/portal views so pay and displayed attendance status
    # never drift apart.
    result = compute_sales_performance_v2_days(emp, period_start, period_end, holiday_dates)

    total_pay = 0.0
    for day, info in result['days'].items():
        regime = info['regime']
        classification = info['classification']
        tvm = info['tvm']
        if regime == 'non_working':
            # Sundays/holidays are only paid if they fall on or after the
            # employee's joining date — otherwise the employee wasn't on
            # payroll yet and the day shouldn't be compensated, mirroring how
            # pre-joining weekdays already fall through to unpaid 'leave'.
            if not emp.joining_date or day >= emp.joining_date:
                al_pct = info.get('annual_leave_salary_pct')
                if al_pct is not None:
                    # Sunday/holiday inside an AnnualLeave span: follow the
                    # leave's own pay rate instead of the normal full wage —
                    # otherwise an unpaid leave still pays its Sundays in full.
                    total_pay += daily_wage * al_pct / 100.0
                else:
                    total_pay += daily_wage
        elif regime == 'friday_saturday':
            if classification == 'full':
                total_pay += daily_wage
        elif regime == 'new_joiner':
            if classification == 'full':
                total_pay += daily_wage
            elif classification == 'half':
                total_pay += half_day_pay
        else:  # standard
            if classification == 'full':
                total_pay += daily_wage
            elif classification == 'proportional':
                total_pay += (tvm / STANDARD_FULL_DAY_THRESHOLD) * daily_wage
            elif classification == 'half':
                total_pay += half_day_pay

    # Target override: foreign-currency employees (INR, NPR, ...) who hit their
    # currency's account target for the month get full salary regardless of attendance.
    target_achieved = False
    if emp.currency != 'AED' and year is not None and month is not None:
        threshold, _overflow_rate = _get_tier_settings(emp.currency)
        total_submissions = BankSubmission.objects.filter(
            remote_employee=emp, year=year, month=month,
        ).aggregate(total=Sum('submission_count'))['total'] or 0
        if total_submissions >= threshold:
            target_achieved = True
            total_pay = salary

    return {
        'employee': emp,
        'employee_type': 'remote',
        'currency': emp.currency,
        'salary': round(salary, 2),
        'daily_rate': round(daily_wage, 2),
        'leave_days': result['leave_days'],
        'half_days': result['half_days'],
        'proportional_days': result['proportional_days'],
        'full_days': result['full_days'],
        'non_working_days': result['non_working_days'],
        'new_joiner_days': result['new_joiner_days'],
        'grace_denials': result['grace_denials'],
        'target_achieved': target_achieved,
        'net_payroll_test': round(total_pay, 2),
    }


def _build_deduction_recovery_map():
    """Per-employee open-carryover-chain map, used to decide whether a past
    DeductionEntry has actually been recovered or is still tangled up in an
    unresolved carryover chain.

    A month M going into carryover means that month's payroll came out
    negative. That debt is proven "resolved" once month M+1 has actually
    been marked paid *and* did not itself spawn a new carryover row (if it
    had gone negative too, a "from M+1" row would exist). So for each
    employee we look at their most recent carryover month, and check whether
    the following month has been marked paid — if so, the debt was cleanly
    absorbed there (it can't be in the open set, since we already took the
    max). If that following month hasn't been processed yet, the balance is
    still open and pending.

    Returns {(emp_type, emp_id): (is_open, chain_start_idx)} — chain_start_idx
    is the earliest month (as year*12+(month-1)) in the unbroken run of
    negative months ending at the open balance, only meaningful when is_open.
    """
    from_idxs_by_emp = _defaultdict(set)
    for co in DeductionCarryover.objects.filter(is_skipped=False):
        key = ('inhouse', co.employee_id) if co.employee_id else ('remote', co.remote_employee_id)
        from_idxs_by_emp[key].add(co.from_year * 12 + (co.from_month - 1))

    paid_idx_by_emp = _defaultdict(set)
    for pr in PaidSalaryRecord.objects.all():
        key = ('inhouse', pr.employee_id) if pr.employee_id else ('remote', pr.remote_employee_id)
        paid_idx_by_emp[key].add(pr.year * 12 + (pr.month - 1))

    recovery_map = {}
    for key, idx_set in from_idxs_by_emp.items():
        max_from = max(idx_set)
        next_idx = max_from + 1
        if next_idx in paid_idx_by_emp.get(key, set()):
            recovery_map[key] = (False, None)
            continue
        chain_start = max_from
        while (chain_start - 1) in idx_set:
            chain_start -= 1
        recovery_map[key] = (True, chain_start)
    return recovery_map


def _entry_is_recovered(entry_end_idx, key, recovery_map, today_idx):
    """True if a DeductionEntry (identified by its last installment month
    index and (emp_type, emp_id) key) has been fully recovered."""
    if entry_end_idx >= today_idx:
        return False
    is_open, chain_start = recovery_map.get(key, (False, None))
    if is_open and entry_end_idx >= chain_start:
        return False
    return True


def _build_section_totals(rows):
    """Sum net, incentives, reductions, commission across a list of payroll rows."""
    return (
        round(sum(r['net_payroll'] for r in rows), 2),
        round(sum(r['incentives'] for r in rows), 2),
        round(sum(r['reductions'] for r in rows), 2),
        round(sum(r.get('commission', 0) for r in rows), 2),
    )


# ============================================
# Freeze / Snapshot Helpers
# ============================================

def _serialize_payroll_context(ctx):
    """Convert a computed payroll context dict into a JSON-serialisable snapshot.
    Replaces Employee/RemoteEmployee model objects with flat primitive fields so
    the snapshot can be stored in FrozenPayrollMonth.snapshot (JSONField).
    """
    def _emp_fields(emp):
        return {
            'employee_id': emp.id,
            'employee_name': emp.name,
            'employee_designation': getattr(emp, 'designation', None),
            'employee_location': getattr(emp, 'location', '') or '',
            'employee_department': getattr(emp, 'department', '') or '',
            'employee_currency': getattr(emp, 'currency', 'AED'),
            'employee_is_fixed_salary': getattr(emp, 'is_fixed_salary', False),
        }

    def _ser_rows(rows):
        result = []
        for row in rows:
            d = {}
            for k, v in row.items():
                if k == 'employee':
                    d.update(_emp_fields(v))
                else:
                    d[k] = v
            result.append(d)
        return result

    banks = ctx.get('banks', [])
    return {
        'meta': {
            'total_holidays': ctx.get('total_holidays', 0),
            'inr_exchange_rate': ctx.get('inr_exchange_rate'),
        },
        'admin_data': _ser_rows(ctx.get('admin_data', [])),
        'total_admin': ctx.get('total_admin', 0),
        'admin_incentives_total': ctx.get('admin_incentives_total', 0),
        'admin_reductions_total': ctx.get('admin_reductions_total', 0),
        'admin_commission_data': _ser_rows(ctx.get('admin_commission_data', [])),
        'total_admin_commission': ctx.get('total_admin_commission', 0),
        'banks': [
            {
                'id': b.id, 'name': b.name,
                'per_account_charge': float(b.per_account_charge),
                'inr_per_account_charge': float(b.inr_per_account_charge) if b.inr_per_account_charge else None,
            }
            for b in banks
        ],
        'all_sales_data': _ser_rows(ctx.get('all_sales_data', [])),
        'total_sales_inhouse': ctx.get('total_sales_inhouse', 0),
        'total_remote': ctx.get('total_remote', 0),
        'total_sales': ctx.get('total_sales', 0),
        'total_sales_aed': ctx.get('total_sales_aed', 0),
        'total_sales_inr': ctx.get('total_sales_inr', 0),
        'total_sales_commission_aed': ctx.get('total_sales_commission_aed', 0),
        'total_sales_commission_inr': ctx.get('total_sales_commission_inr', 0),
        'sales_incentives_total': ctx.get('sales_incentives_total', 0),
        'sales_reductions_total': ctx.get('sales_reductions_total', 0),
        'total_sales_all_aed': ctx.get('total_sales_all_aed', 0),
        'section3_rows': _ser_rows(ctx.get('section3_rows', [])),
        's3_total_ded': ctx.get('s3_total_ded', 0),
        's3_total_add': ctx.get('s3_total_add', 0),
        's3_net': ctx.get('s3_net', 0),
        's3_total_ded_aed': ctx.get('s3_total_ded_aed', 0),
        's3_total_ded_inr': ctx.get('s3_total_ded_inr', 0),
        's3_total_add_aed': ctx.get('s3_total_add_aed', 0),
        's3_total_add_inr': ctx.get('s3_total_add_inr', 0),
        's3_net_aed': ctx.get('s3_net_aed', 0),
        's3_net_inr': ctx.get('s3_net_inr', 0),
        'final_rows': _ser_rows(ctx.get('final_rows', [])),
        'final_total': ctx.get('final_total', 0),
        'final_total_aed': ctx.get('final_total_aed', 0),
        'final_total_inr': ctx.get('final_total_inr', 0),
        'final_total_combined_aed': ctx.get('final_total_combined_aed', 0),
        'grand_total': ctx.get('grand_total', 0),
        'grand_total_aed': ctx.get('grand_total_aed', 0),
        'grand_total_inr': ctx.get('grand_total_inr', 0),
    }


def _row_to_ns(d):
    """Wrap a serialised row dict in SimpleNamespace so Django template dot-access works."""
    emp = SimpleNamespace(
        id=d.get('employee_id'),
        name=d.get('employee_name', ''),
        designation=d.get('employee_designation'),
        location=d.get('employee_location', ''),
        department=d.get('employee_department', ''),
        currency=d.get('employee_currency', 'AED'),
        is_fixed_salary=d.get('employee_is_fixed_salary', False),
    )
    ns = SimpleNamespace(**{k: v for k, v in d.items()})
    ns.employee = emp
    return ns


def _deserialize_payroll_context(snapshot, selected_month, selected_year):
    """Rebuild a template-compatible context from a frozen JSON snapshot."""
    banks = [SimpleNamespace(**b) for b in snapshot.get('banks', [])]
    banks_json = json.dumps([
        {
            'id': b.id, 'name': b.name,
            'rate': b.per_account_charge,
            'inr_rate': b.inr_per_account_charge,
        }
        for b in banks
    ])
    meta = snapshot.get('meta', {})
    return {
        'total_holidays': meta.get('total_holidays', 0),
        'inr_exchange_rate': meta.get('inr_exchange_rate'),
        'banks': banks,
        'banks_json': banks_json,
        'admin_data': [_row_to_ns(r) for r in snapshot.get('admin_data', [])],
        'total_admin': snapshot.get('total_admin', 0),
        'admin_incentives_total': snapshot.get('admin_incentives_total', 0),
        'admin_reductions_total': snapshot.get('admin_reductions_total', 0),
        'admin_commission_data': [_row_to_ns(r) for r in snapshot.get('admin_commission_data', [])],
        'total_admin_commission': snapshot.get('total_admin_commission', 0),
        'all_sales_data': [_row_to_ns(r) for r in snapshot.get('all_sales_data', [])],
        'total_sales_inhouse': snapshot.get('total_sales_inhouse', 0),
        'total_remote': snapshot.get('total_remote', 0),
        'total_sales': snapshot.get('total_sales', 0),
        'total_sales_aed': snapshot.get('total_sales_aed', 0),
        'total_sales_inr': snapshot.get('total_sales_inr', 0),
        'total_sales_commission_aed': snapshot.get('total_sales_commission_aed', 0),
        'total_sales_commission_inr': snapshot.get('total_sales_commission_inr', 0),
        'sales_incentives_total': snapshot.get('sales_incentives_total', 0),
        'sales_reductions_total': snapshot.get('sales_reductions_total', 0),
        'total_sales_all_aed': snapshot.get('total_sales_all_aed', 0),
        'section3_rows': [_row_to_ns(r) for r in snapshot.get('section3_rows', [])],
        'all_deductions_list': [],  # edit modal disabled when frozen
        'all_employees_json': '[]',
        's3_total_ded': snapshot.get('s3_total_ded', 0),
        's3_total_add': snapshot.get('s3_total_add', 0),
        's3_net': snapshot.get('s3_net', 0),
        's3_total_ded_aed': snapshot.get('s3_total_ded_aed', 0),
        's3_total_ded_inr': snapshot.get('s3_total_ded_inr', 0),
        's3_total_add_aed': snapshot.get('s3_total_add_aed', 0),
        's3_total_add_inr': snapshot.get('s3_total_add_inr', 0),
        's3_net_aed': snapshot.get('s3_net_aed', 0),
        's3_net_inr': snapshot.get('s3_net_inr', 0),
        'final_rows': [_row_to_ns(r) for r in snapshot.get('final_rows', [])],
        'final_total': snapshot.get('final_total', 0),
        'final_total_aed': snapshot.get('final_total_aed', 0),
        'final_total_inr': snapshot.get('final_total_inr', 0),
        'final_total_combined_aed': snapshot.get('final_total_combined_aed', 0),
        'grand_total': snapshot.get('grand_total', 0),
        'grand_total_aed': snapshot.get('grand_total_aed', 0),
        'grand_total_inr': snapshot.get('grand_total_inr', 0),
        # Carryover timeline is always live (spans all months, not specific to frozen month)
        'all_carryovers': DeductionCarryover.objects.none(),
        'carryover_timeline': [],
        # Template navigation
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_name': MONTH_NAMES[selected_month],
        'months': MONTH_CHOICES,
        'years': YEAR_RANGE,
        'INR_COMMISSION_THRESHOLD': _get_tier_settings('INR')[0],
        'INR_OVERFLOW_RATE': int(_get_tier_settings('INR')[1]),
        'DEDUCTION_CATEGORY_CHOICES': DEDUCTION_CATEGORY_CHOICES,
        'is_frozen': True,
    }


# ============================================
# Main Dashboard
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def payroll_dashboard(request):
    """Payroll dashboard showing Admin and Sales sections."""
    selected_month, selected_year = get_selected_month_year(request)

    # --- Frozen month: serve snapshot instead of recalculating ---
    frozen_obj = FrozenPayrollMonth.objects.filter(
        year=selected_year, month=selected_month
    ).first()
    if frozen_obj:
        ctx = _deserialize_payroll_context(frozen_obj.snapshot, selected_month, selected_year)
        ctx['frozen_at'] = frozen_obj.frozen_at
        ctx['frozen_by'] = frozen_obj.frozen_by
        return render(request, 'payroll/dashboard.html', ctx)

    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    month_start = datetime.date(selected_year, selected_month, 1)
    month_end = datetime.date(selected_year, selected_month, days_in_month)

    total_holidays = _count_holidays(selected_year, selected_month, days_in_month)

    # Fetch active banks for sales spreadsheet
    banks = list(Bank.objects.filter(is_active=True).order_by('name'))
    banks_json = json.dumps([
        {
            'id': b.id,
            'name': b.name,
            'rate': float(b.per_account_charge),
            'inr_rate': float(b.inr_per_account_charge) if b.inr_per_account_charge else None,
        }
        for b in banks
    ])

    # --- Admin section (in-house Admin dept) ---
    admin_employees = list(Employee.objects.filter(
        department='Admin', is_active=True
    ).order_by('name'))
    admin_data = [
        _get_inhouse_payroll_row(emp, selected_year, selected_month, month_start, month_end, total_holidays)
        for emp in admin_employees
    ]
    total_admin, admin_incentives_total, admin_reductions_total, _ = _build_section_totals(admin_data)

    # Admin commission spreadsheet data (for Sales tab sub-section)
    admin_commission_data = []
    for emp in admin_employees:
        submissions_qs = BankSubmission.objects.filter(employee=emp, year=selected_year, month=selected_month)
        bank_counts = {s.bank_id: s.submission_count for s in submissions_qs}
        bank_counts_list = [bank_counts.get(b.id, 0) for b in banks]
        commission = _get_commission(selected_year, selected_month, employee=emp, currency=emp.currency)
        admin_commission_data.append({
            'employee': emp,
            'employee_type': 'inhouse',
            'currency': emp.currency,
            'bank_counts_list': bank_counts_list,
            'commission': round(commission, 2),
        })
    total_admin_commission = round(sum(r['commission'] for r in admin_commission_data), 2)

    # --- Sales section: in-house Sales employees (commission-only) ---
    sales_inhouse_employees = Employee.objects.filter(
        department='Sales', is_active=True
    ).order_by('name')
    sales_inhouse_data = [
        _get_sales_payroll_row(emp, selected_year, selected_month, 'inhouse', banks, days_in_month, total_holidays)
        for emp in sales_inhouse_employees
    ]
    total_sales_inhouse, _, _, _ = _build_section_totals(sales_inhouse_data)

    # --- Sales section: remote employees (commission-only) ---
    # Exclude remote employees whose tcr_id already appears as an in-house employee
    # (same person tracked in both tables) to prevent duplicates
    all_inhouse_tcr_ids = set(
        Employee.objects.filter(is_active=True)
        .exclude(tcr_id='')
        .values_list('tcr_id', flat=True)
    )
    remote_employees = RemoteEmployee.objects.filter(is_active=True).order_by('name')
    if all_inhouse_tcr_ids:
        remote_employees = remote_employees.exclude(tcr_id__in=all_inhouse_tcr_ids)
    remote_data = [
        _get_sales_payroll_row(emp, selected_year, selected_month, 'remote', banks, days_in_month, total_holidays)
        for emp in remote_employees
    ]
    total_remote, _, _, _ = _build_section_totals(remote_data)

    # Combined sales data: salary-bearing rows (fixed / attendance-based) first,
    # then pure commission rows
    def _has_base_salary(r):
        return r.get('is_fixed_salary') or r.get('is_attendance_based')

    all_sales_data = sorted(
        sales_inhouse_data + remote_data,
        key=lambda r: (0 if _has_base_salary(r) else 1, r['employee'].name.lower())
    )

    # Combined Sales totals
    all_sales_rows = sales_inhouse_data + remote_data
    total_sales, sales_incentives_total, sales_reductions_total, _ = _build_section_totals(all_sales_rows)
    total_sales_aed = round(sum(r['net_payroll'] for r in all_sales_rows if r.get('currency', 'AED') == 'AED'), 2)
    total_sales_inr = round(sum(r['net_payroll'] for r in all_sales_rows if r.get('currency', 'AED') == 'INR'), 2)
    # Tfoot totals split by currency. Salary-bearing rows surface net_payroll
    # (they may have salary + commission), pure commission rows surface commission only.
    def _sales_tfoot(rows, currency):
        return round(
            sum(r.get('commission', 0) for r in rows if not _has_base_salary(r) and r.get('currency', 'AED') == currency) +
            sum(r['net_payroll'] for r in rows if _has_base_salary(r) and r.get('currency', 'AED') == currency),
            2
        )
    total_sales_commission_aed = _sales_tfoot(all_sales_rows, 'AED')
    total_sales_commission_inr = _sales_tfoot(all_sales_rows, 'INR')

    # --- Exchange rate for INR → AED conversion (rate on 10th of the month) ---
    inr_exchange_rate = None
    inr_rate_obj = ExchangeRate.objects.filter(
        currency='INR', year=selected_year, month=selected_month
    ).first()
    if inr_rate_obj:
        inr_exchange_rate = float(inr_rate_obj.rate)

    # Add net_payroll_aed to each sales row
    for row in all_sales_data:
        amount = row.get('net_payroll') if (row.get('is_fixed_salary') or row.get('is_attendance_based')) else row.get('commission', 0)
        if row.get('currency', 'AED') == 'INR' and inr_exchange_rate and inr_exchange_rate > 0:
            row['net_payroll_aed'] = round(amount / inr_exchange_rate, 2)
        else:
            row['net_payroll_aed'] = round(amount, 2)

    total_sales_all_aed = round(sum(r['net_payroll_aed'] for r in all_sales_data), 2)

    grand_total = round(total_admin + total_sales, 2)

    # --- Section 3: Employee Deduction Spreadsheet ---
    _month_names = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec',
    }
    _DED_COLS = ['advance', 'visa_status_change', 'clawback', 'leave_deduction', 'late_deduction', 'other_deduction']
    _ADD_COLS = ['last_month_balance', 'paid_leave', 'other_addition']
    _ALL_CATS = _DED_COLS + _ADD_COLS
    target_idx = selected_year * 12 + (selected_month - 1)

    # Load all DeductionEntry records; build active-this-month map and all-entries list
    active_by_emp = _defaultdict(list)
    all_deductions_list = []
    for entry in DeductionEntry.objects.select_related('employee', 'remote_employee').order_by('-created_at'):
        emp = entry.employee or entry.remote_employee
        emp_type = 'inhouse' if entry.employee else 'remote'
        start_idx = entry.start_year * 12 + (entry.start_month - 1)
        if start_idx <= target_idx < start_idx + entry.split_months:
            active_by_emp[(emp_type, emp.id)].append(entry)
        end_y, end_m = entry.end_month_year()
        all_deductions_list.append({
            'id': entry.id,
            'employee': emp,
            'employee_type': emp_type,
            'currency': emp.currency,
            'category_display': entry.get_category_display(),
            'entry_type': entry.entry_type,
            'total_amount': float(entry.total_amount),
            'split_months': entry.split_months,
            'installment_amount': float(entry.installment_amount),
            'start_month_name': _month_names[entry.start_month],
            'start_year': entry.start_year,
            'end_month_name': _month_names[end_m],
            'end_year': end_y,
            'note': entry.note,
            'created_at': entry.created_at.strftime('%d %b %Y'),
        })

    # Pre-load monthly summaries for auto leave/late deductions (avoid N+1)
    inhouse_summaries = {
        s.employee_id: s
        for s in MonthlySummary.objects.filter(year=selected_year, month=selected_month)
    }

    # Load incoming carryovers for the selected month (overflow from previous month)
    incoming_carryovers = DeductionCarryover.objects.filter(
        to_year=selected_year, to_month=selected_month
    ).exclude(is_skipped=True)
    carryover_by_emp = {}
    for co in incoming_carryovers:
        key = ('inhouse', co.employee_id) if co.employee_id else ('remote', co.remote_employee_id)
        carryover_by_emp[key] = co.overflow_amount

    # Use the same employee sets as the payroll sections (Admin + Sales dept for in-house,
    # tcr_id-deduped for remote). Merge deductions from any same-name duplicates.
    _payroll_inhouse = list(Employee.objects.filter(
        is_active=True, department__in=['Admin', 'Sales']
    ).order_by('department', 'name'))
    _payroll_inhouse_ids = {e.id for e in _payroll_inhouse}
    _payroll_inhouse_names = {e.name: e.id for e in _payroll_inhouse}

    # Map: payroll employee id -> [duplicate employee ids with same name]
    _inhouse_dupe_ids = _defaultdict(list)
    for _emp in Employee.objects.filter(is_active=True).exclude(id__in=_payroll_inhouse_ids):
        main_id = _payroll_inhouse_names.get(_emp.name)
        if main_id:
            _inhouse_dupe_ids[main_id].append(_emp.id)

    # remote_employees is already deduped by tcr_id (defined above in sales section)
    _remote_ids = {e.id for e in remote_employees}
    _remote_names = {e.name: e.id for e in remote_employees}
    _remote_dupe_ids = _defaultdict(list)
    for _emp in RemoteEmployee.objects.filter(is_active=True).exclude(id__in=_remote_ids):
        main_id = _remote_names.get(_emp.name)
        if main_id:
            _remote_dupe_ids[main_id].append(_emp.id)

    section3_rows = []
    for emp in _payroll_inhouse:
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get(('inhouse', emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        for dupe_id in _inhouse_dupe_ids.get(emp.id, []):
            for ded_entry in active_by_emp.get(('inhouse', dupe_id), []):
                cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        # Auto-compute leave/late from attendance (skip for performance-based payroll)
        summary = inhouse_summaries.get(emp.id)
        if emp.salary and emp.payroll_type != 'performance':
            daily = float(emp.salary) / days_in_month
            bridge_count = len(get_bridge_sunday_days(emp, month_start, month_end))
            base_leave_days = (summary.leave_days if summary else 0) or 0
            base_late_days = (summary.late_days if summary else 0) or 0
            cat['leave_deduction'] = round(daily * (base_leave_days + bridge_count), 2)
            cat['late_deduction'] = round(daily * (base_late_days // 3) * 0.5, 2)
        # Exclude leave/late from total_deductions when payroll_net already bakes in the
        # attendance deduction (Admin always; Sales when salary + non-performance payroll).
        # Pure-commission Sales rows have payroll_net = commission with no attendance hit.
        _has_attendance_salary = (
            emp.salary and
            getattr(emp, 'payroll_type', 'attendance') != 'performance'
        )
        if emp.department == 'Admin' or _has_attendance_salary:
            ded_cols_for_total = [c for c in _DED_COLS if c not in ('leave_deduction', 'late_deduction')]
        else:
            ded_cols_for_total = _DED_COLS
        total_ded = round(sum(cat[c] for c in ded_cols_for_total), 2)
        carryover_in = float(carryover_by_emp.get(('inhouse', emp.id), 0))
        for dupe_id in _inhouse_dupe_ids.get(emp.id, []):
            carryover_in += float(carryover_by_emp.get(('inhouse', dupe_id), 0))
        total_ded = round(total_ded + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        row = {'employee': emp, 'employee_type': 'inhouse', 'is_inhouse': True, 'currency': emp.currency}
        row.update(cat)
        row.update({'total_deductions': total_ded, 'total_additions': total_add, 'net': round(total_add - total_ded, 2), 'carryover_in': carryover_in})
        section3_rows.append(row)

    for emp in remote_employees:
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get(('remote', emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        for dupe_id in _remote_dupe_ids.get(emp.id, []):
            for ded_entry in active_by_emp.get(('remote', dupe_id), []):
                cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        total_ded = round(sum(cat[c] for c in _DED_COLS), 2)
        carryover_in = float(carryover_by_emp.get(('remote', emp.id), 0))
        for dupe_id in _remote_dupe_ids.get(emp.id, []):
            carryover_in += float(carryover_by_emp.get(('remote', dupe_id), 0))
        total_ded = round(total_ded + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        row = {'employee': emp, 'employee_type': 'remote', 'is_inhouse': False, 'currency': emp.currency}
        row.update(cat)
        row.update({'total_deductions': total_ded, 'total_additions': total_add, 'net': round(total_add - total_ded, 2), 'carryover_in': carryover_in})
        section3_rows.append(row)

    s3_total_ded = round(sum(r['total_deductions'] for r in section3_rows), 2)
    s3_total_add = round(sum(r['total_additions'] for r in section3_rows), 2)
    s3_net = round(s3_total_add - s3_total_ded, 2)
    s3_total_ded_aed = round(sum(r['total_deductions'] for r in section3_rows if r.get('currency', 'AED') == 'AED'), 2)
    s3_total_ded_inr = round(sum(r['total_deductions'] for r in section3_rows if r.get('currency', 'AED') == 'INR'), 2)
    s3_total_add_aed = round(sum(r['total_additions'] for r in section3_rows if r.get('currency', 'AED') == 'AED'), 2)
    s3_total_add_inr = round(sum(r['total_additions'] for r in section3_rows if r.get('currency', 'AED') == 'INR'), 2)
    s3_net_aed = round(s3_total_add_aed - s3_total_ded_aed, 2)
    s3_net_inr = round(s3_total_add_inr - s3_total_ded_inr, 2)

    # All employees for Add modal dropdown (same sets as payroll sections)
    all_employees_json = json.dumps([
        {'id': emp.id, 'name': emp.name, 'type': 'inhouse', 'dept': emp.department or '', 'currency': emp.currency}
        for emp in _payroll_inhouse
    ] + [
        {'id': emp.id, 'name': emp.name, 'type': 'remote', 'dept': 'Remote', 'currency': emp.currency}
        for emp in remote_employees
    ])

    # --- Section 5: Final Summary ---
    # Build lookup dicts keyed by (emp_type, emp_id) for payroll and deductions
    payroll_by_emp = {}
    for row in admin_data:
        payroll_by_emp[('inhouse', row['employee'].id)] = row
    for row in all_sales_data:
        payroll_by_emp[(row['employee_type'], row['employee'].id)] = row

    ded_by_emp = {}
    for row in section3_rows:
        ded_by_emp[(row['employee_type'], row['employee'].id)] = row

    # Compute next calendar month for carryover targeting
    if selected_month == 12:
        _co_to_month, _co_to_year = 1, selected_year + 1
    else:
        _co_to_month, _co_to_year = selected_month + 1, selected_year

    # Employees already marked paid for this month have a locked PaidSalaryRecord
    # snapshot (including its own carryover_out at the time of payment). Don't
    # recompute/delete their carryover rows using whatever the employee's *current*
    # settings (salary, currency, cycle) are — that silently destroys real debt
    # when those settings change after payment.
    _paid_keys_this_month = set(
        PaidSalaryRecord.objects.filter(year=selected_year, month=selected_month).values_list(
            'employee_id', 'remote_employee_id'
        )
    )
    _paid_emp_keys = {('inhouse', e) for e, r in _paid_keys_this_month if e} | {
        ('remote', r) for e, r in _paid_keys_this_month if r
    }

    final_rows = []
    for emp in _payroll_inhouse:
        key = ('inhouse', emp.id)
        p = payroll_by_emp.get(key)
        d = ded_by_emp.get(key)
        payroll_net = p['net_payroll'] if p else 0.0
        total_ded = d['total_deductions'] if d else 0.0
        total_add = d['total_additions'] if d else 0.0
        carryover_in = d['carryover_in'] if d else 0.0
        # Carryover tracking: use full total_ded (including carryover) to decide
        # whether to create/delete a carryover record for next month.
        accounting_net = round(payroll_net - total_ded + total_add, 2)
        if key not in _paid_emp_keys:
            if accounting_net < 0:
                overflow = Decimal(str(abs(accounting_net)))
                DeductionCarryover.objects.update_or_create(
                    employee=emp, from_year=selected_year, from_month=selected_month,
                    defaults={'overflow_amount': overflow, 'to_year': _co_to_year, 'to_month': _co_to_month,
                              'remote_employee': None, 'currency': emp.currency},
                )
            else:
                DeductionCarryover.objects.filter(
                    employee=emp, from_year=selected_year, from_month=selected_month
                ).delete()
            if carryover_in > 0:
                incoming_co = incoming_carryovers.filter(employee=emp).first()
                if incoming_co:
                    incoming_co.applied_amount = Decimal(str(min(carryover_in, float(incoming_co.overflow_amount))))
                    incoming_co.save(update_fields=['applied_amount'])
        # Displayed final salary matches the payslip:
        # - Admin employees: payslip reconstructs attendance deductions without bridge Sunday,
        #   so add it back here.
        # - All employees: payslip does not deduct carryover, so exclude it.
        if p and emp.department == 'Admin':
            bridge_adj = (p.get('bridge_sunday_count') or 0) * (p.get('daily_rate') or 0.0)
        else:
            bridge_adj = 0.0
        final_salary = round(payroll_net + bridge_adj - (total_ded - carryover_in) + total_add, 2)
        if final_salary < 0:
            final_salary = 0.0
        final_rows.append({
            'employee': emp,
            'employee_type': 'inhouse',
            'department': emp.department or 'In-House',
            'currency': emp.currency,
            'payroll_net': payroll_net,
            'total_deductions': total_ded,
            'total_additions': total_add,
            'final_salary': final_salary,
        })

    for emp in remote_employees:
        key = ('remote', emp.id)
        p = payroll_by_emp.get(key)
        d = ded_by_emp.get(key)
        payroll_net = p['net_payroll'] if p else 0.0
        total_ded = d['total_deductions'] if d else 0.0
        total_add = d['total_additions'] if d else 0.0
        carryover_in = d['carryover_in'] if d else 0.0
        # Carryover tracking uses full total_ded.
        accounting_net = round(payroll_net - total_ded + total_add, 2)
        if key not in _paid_emp_keys:
            if accounting_net < 0:
                overflow = Decimal(str(abs(accounting_net)))
                DeductionCarryover.objects.update_or_create(
                    remote_employee=emp, from_year=selected_year, from_month=selected_month,
                    defaults={'overflow_amount': overflow, 'to_year': _co_to_year, 'to_month': _co_to_month,
                              'employee': None, 'currency': emp.currency},
                )
            else:
                DeductionCarryover.objects.filter(
                    remote_employee=emp, from_year=selected_year, from_month=selected_month
                ).delete()
            if carryover_in > 0:
                incoming_co = incoming_carryovers.filter(remote_employee=emp).first()
                if incoming_co:
                    incoming_co.applied_amount = Decimal(str(min(carryover_in, float(incoming_co.overflow_amount))))
                    incoming_co.save(update_fields=['applied_amount'])
        # Displayed final salary: exclude carryover (matches payslip).
        final_salary = round(payroll_net - (total_ded - carryover_in) + total_add, 2)
        if final_salary < 0:
            final_salary = 0.0
        final_rows.append({
            'employee': emp,
            'employee_type': 'remote',
            'department': 'Remote',
            'currency': emp.currency,
            'payroll_net': payroll_net,
            'total_deductions': total_ded,
            'total_additions': total_add,
            'final_salary': final_salary,
        })

    final_total_aed = round(sum(r['final_salary'] for r in final_rows if r.get('currency', 'AED') == 'AED'), 2)
    final_total_inr = round(sum(r['final_salary'] for r in final_rows if r.get('currency', 'AED') == 'INR'), 2)
    final_total = round(sum(r['final_salary'] for r in final_rows), 2)
    # Combined total: convert INR to AED and add
    if inr_exchange_rate and inr_exchange_rate > 0 and final_total_inr > 0:
        final_total_combined_aed = round(final_total_aed + (final_total_inr / inr_exchange_rate), 2)
    else:
        final_total_combined_aed = final_total_aed

    # Carryover history for all active employees
    from django.db.models import Q as _Q
    all_carryovers = DeductionCarryover.objects.filter(
        _Q(employee__in=Employee.objects.filter(is_active=True)) |
        _Q(remote_employee__in=RemoteEmployee.objects.filter(is_active=True))
    ).select_related('employee', 'remote_employee').order_by('-from_year', '-from_month')

    # Pre-build deduction entry lookup for timeline breakdown
    _all_ded_entries_raw = DeductionEntry.objects.select_related('employee', 'remote_employee').filter(
        _Q(employee__in=Employee.objects.filter(is_active=True)) |
        _Q(remote_employee__in=RemoteEmployee.objects.filter(is_active=True))
    )
    _entries_by_emp = defaultdict(list)
    for _e in _all_ded_entries_raw:
        _ek2 = ('inhouse', _e.employee_id) if _e.employee_id else ('remote', _e.remote_employee_id)
        _entries_by_emp[_ek2].append(_e)

    # Build per-employee 12-month deduction/carryover timeline for selected year
    _co_timeline_dict = {}

    # Helper: collect all entries for an employee (including duplicates)
    def _get_emp_entries(_ek):
        _entries = list(_entries_by_emp.get(_ek, []))
        if _ek[0] == 'inhouse':
            for _did in _inhouse_dupe_ids.get(_ek[1], []):
                _entries.extend(_entries_by_emp.get(('inhouse', _did), []))
        else:
            for _did in _remote_dupe_ids.get(_ek[1], []):
                _entries.extend(_entries_by_emp.get(('remote', _did), []))
        return _entries

    # Helper: build breakdown for a given month index from a list of entries
    def _build_breakdown(_entries, _target_idx):
        _brkd_ded, _brkd_add, _adv = {}, {}, []
        for _e in _entries:
            _e_start = _e.start_year * 12 + (_e.start_month - 1)
            if _e_start <= _target_idx < _e_start + _e.split_months:
                _amt = round(float(_e.installment_amount), 2)
                if _e.category == 'advance':
                    _adv.append({'entry_id': _e.id, 'amount': _amt, 'note': _e.note or 'Cash Advance'})
                elif _e.entry_type == 'deduction':
                    _cat = _e.get_category_display()
                    _brkd_ded[_cat] = round(_brkd_ded.get(_cat, 0) + _amt, 2)
                else:
                    _cat = _e.get_category_display()
                    _brkd_add[_cat] = round(_brkd_add.get(_cat, 0) + _amt, 2)
        return _brkd_ded, _brkd_add, _adv

    # Step 1: Build timeline from ALL employees with deduction entries
    # Collect entries per employee: payroll employees (with duplicate merging) + any others
    _timeline_emp_data = {}  # _ek -> (entries_list, name, dept, currency)
    _covered_keys = set()
    for emp in _payroll_inhouse:
        _ek = ('inhouse', emp.id)
        _all_entries = _get_emp_entries(_ek)
        if _all_entries:
            _timeline_emp_data[_ek] = (_all_entries, emp.name, emp.department or 'In-House', emp.currency)
        _covered_keys.add(_ek)
        for _did in _inhouse_dupe_ids.get(emp.id, []):
            _covered_keys.add(('inhouse', _did))
    for emp in remote_employees:
        _ek = ('remote', emp.id)
        _all_entries = _get_emp_entries(_ek)
        if _all_entries:
            _timeline_emp_data[_ek] = (_all_entries, emp.name, 'Remote', emp.currency)
        _covered_keys.add(_ek)
        for _did in _remote_dupe_ids.get(emp.id, []):
            _covered_keys.add(('remote', _did))
    # Include employees with entries that aren't in the payroll sets
    for _ek, _entries in _entries_by_emp.items():
        if _ek in _covered_keys:
            continue
        _eobj = _entries[0].employee if _ek[0] == 'inhouse' else _entries[0].remote_employee
        _timeline_emp_data[_ek] = (
            list(_entries), _eobj.name,
            _eobj.department or 'Other' if _ek[0] == 'inhouse' else 'Remote',
            _eobj.currency,
        )

    for _ek, (_all_entries, _ename, _edept, _ecurr) in _timeline_emp_data.items():
        cells = [None] * 12
        has_any = False
        for _mi in range(12):
            _tidx = selected_year * 12 + _mi
            _bd, _ba, _adv = _build_breakdown(_all_entries, _tidx)
            if _adv or _bd or _ba:
                has_any = True
                cells[_mi] = {
                    'type': 'entries',
                    'advance_items': _adv,
                    'breakdown_ded': _bd,
                    'breakdown_add': _ba,
                }
        if has_any:
            _co_timeline_dict[_ek] = {
                'name': _ename, 'dept': _edept, 'currency': _ecurr,
                'emp_type': _ek[0], 'emp_id': _ek[1], 'cells': cells,
            }

    # Step 2: Overlay carryover data on top of entry cells
    _co_year_qs = DeductionCarryover.objects.filter(
        (_Q(employee__in=Employee.objects.filter(is_active=True)) |
         _Q(remote_employee__in=RemoteEmployee.objects.filter(is_active=True))) &
        (_Q(from_year=selected_year) | _Q(to_year=selected_year))
    ).select_related('employee', 'remote_employee').order_by('from_year', 'from_month')
    for _co in _co_year_qs:
        if _co.employee_id:
            _ek = ('inhouse', _co.employee_id)
            _ename = _co.employee.name
            _edept = _co.employee.department or 'In-House'
            _ecurr = _co.employee.currency
        else:
            _ek = ('remote', _co.remote_employee_id)
            _ename = _co.remote_employee.name
            _edept = 'Remote'
            _ecurr = _co.remote_employee.currency
        if _ek not in _co_timeline_dict:
            _co_timeline_dict[_ek] = {
                'name': _ename, 'dept': _edept, 'currency': _ecurr,
                'emp_type': _ek[0], 'emp_id': _ek[1],
                'cells': [None] * 12,
            }
        _cells = _co_timeline_dict[_ek]['cells']
        _status = 'applied' if _co.applied_amount >= _co.overflow_amount else ('partial' if _co.applied_amount > 0 else 'pending')
        if _co.from_year == selected_year:
            _existing = _cells[_co.from_month - 1]
            # Grab entry breakdown from existing cell (entries or carryover-with-entries)
            _brkd_ded = _existing.get('breakdown_ded', {}) if _existing else {}
            _brkd_add = _existing.get('breakdown_add', {}) if _existing else {}
            _advance_items = _existing.get('advance_items', []) if _existing else []
            # Preserve incoming carryover info if overwriting a carryover cell
            _incoming_co_info = None
            if _existing and _existing.get('type') == 'carryover':
                _incoming_co_info = {
                    'amount': _existing['overflow'],
                    'applied': _existing['applied'],
                    'from_month_name': _existing['from_month_name'],
                    'from_year': _existing['from_year'],
                }
            elif _existing and _existing.get('incoming_carryover'):
                _incoming_co_info = _existing['incoming_carryover']
            _cells[_co.from_month - 1] = {
                'type': 'overflow',
                'amount': float(_co.overflow_amount),
                'to_month': _co.to_month,
                'to_month_name': _month_names[_co.to_month],
                'to_year': _co.to_year,
                'cross_year': _co.to_year != selected_year,
                'status': _status,
                'breakdown_ded': _brkd_ded,
                'breakdown_add': _brkd_add,
                'advance_items': _advance_items,
                'incoming_carryover': _incoming_co_info,
            }
        if _co.to_year == selected_year:
            _existing = _cells[_co.to_month - 1]
            # Don't overwrite an overflow cell (it already merged everything)
            if not _existing or _existing.get('type') != 'overflow':
                _carryover_cell = {
                    'type': 'carryover',
                    'applied': float(_co.applied_amount),
                    'overflow': float(_co.overflow_amount),
                    'remaining': round(float(_co.overflow_amount) - float(_co.applied_amount), 2),
                    'from_month': _co.from_month,
                    'from_month_name': _month_names[_co.from_month],
                    'from_year': _co.from_year,
                    'cross_year': _co.from_year != selected_year,
                    'status': _status,
                }
                # Preserve entry breakdown if overwriting an entries cell
                if _existing and _existing.get('type') == 'entries':
                    _carryover_cell['advance_items'] = _existing['advance_items']
                    _carryover_cell['breakdown_ded'] = _existing['breakdown_ded']
                    _carryover_cell['breakdown_add'] = _existing['breakdown_add']
                _cells[_co.to_month - 1] = _carryover_cell
    carryover_timeline = list(_co_timeline_dict.values())

    # Grand total split by currency
    all_rows = admin_data + all_sales_data
    grand_total_aed = round(sum(r['net_payroll'] for r in all_rows if r.get('currency', 'AED') == 'AED'), 2)
    grand_total_inr = round(sum(r['net_payroll'] for r in all_rows if r.get('currency', 'AED') == 'INR'), 2)

    context = {
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_name': MONTH_NAMES[selected_month],
        'months': MONTH_CHOICES,
        'years': YEAR_RANGE,
        'total_holidays': total_holidays,
        # Admin
        'admin_data': admin_data,
        'total_admin': total_admin,
        'admin_incentives_total': admin_incentives_total,
        'admin_reductions_total': admin_reductions_total,
        'admin_commission_data': admin_commission_data,
        'total_admin_commission': total_admin_commission,
        # Sales (spreadsheet)
        'banks': banks,
        'banks_json': banks_json,
        'INR_COMMISSION_THRESHOLD': _get_tier_settings('INR')[0],
        'INR_OVERFLOW_RATE': int(_get_tier_settings('INR')[1]),
        'all_sales_data': all_sales_data,
        'total_sales_inhouse': total_sales_inhouse,
        'total_remote': total_remote,
        'total_sales': total_sales,
        'total_sales_aed': total_sales_aed,
        'total_sales_inr': total_sales_inr,
        'total_sales_commission_aed': total_sales_commission_aed,
        'total_sales_commission_inr': total_sales_commission_inr,
        'sales_incentives_total': sales_incentives_total,
        'sales_reductions_total': sales_reductions_total,
        'inr_exchange_rate': inr_exchange_rate,
        'total_sales_all_aed': total_sales_all_aed,
        # Deductions & Additions (Section 3)
        'section3_rows': section3_rows,
        'all_deductions_list': all_deductions_list,
        's3_total_ded': s3_total_ded,
        's3_total_add': s3_total_add,
        's3_net': s3_net,
        's3_total_ded_aed': s3_total_ded_aed,
        's3_total_ded_inr': s3_total_ded_inr,
        's3_total_add_aed': s3_total_add_aed,
        's3_total_add_inr': s3_total_add_inr,
        's3_net_aed': s3_net_aed,
        's3_net_inr': s3_net_inr,
        'all_employees_json': all_employees_json,
        # Carryover schedule
        'all_carryovers': all_carryovers,
        'carryover_timeline': carryover_timeline,
        # Final Summary
        'final_rows': final_rows,
        'final_total': final_total,
        'final_total_aed': final_total_aed,
        'final_total_inr': final_total_inr,
        'final_total_combined_aed': final_total_combined_aed,
        # Grand total
        'grand_total': grand_total,
        'grand_total_aed': grand_total_aed,
        'grand_total_inr': grand_total_inr,
        # Freeze state
        'is_frozen': False,
        'frozen_at': None,
        'frozen_by': None,
        'DEDUCTION_CATEGORY_CHOICES': DEDUCTION_CATEGORY_CHOICES,
    }

    return render(request, 'payroll/dashboard.html', context)


# ============================================
# Freeze / Unfreeze Payroll Month
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def freeze_payroll(request):
    """Compute and store an immutable snapshot of payroll for the given month."""
    try:
        data = json.loads(request.body)
        year = int(data['year'])
        month = int(data['month'])
    except (json.JSONDecodeError, KeyError, ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

    if FrozenPayrollMonth.objects.filter(year=year, month=month).exists():
        return JsonResponse({'success': False, 'error': 'This month is already frozen.'}, status=400)

    # Run the full payroll calculation inline (mirrors payroll_dashboard logic).
    _, days_in_month = calendar.monthrange(year, month)
    month_start = datetime.date(year, month, 1)
    month_end = datetime.date(year, month, days_in_month)
    total_holidays = _count_holidays(year, month, days_in_month)
    banks = list(Bank.objects.filter(is_active=True).order_by('name'))

    admin_employees = list(Employee.objects.filter(department='Admin', is_active=True).order_by('name'))
    admin_data = [
        _get_inhouse_payroll_row(emp, year, month, month_start, month_end, total_holidays)
        for emp in admin_employees
    ]
    total_admin, admin_incentives_total, admin_reductions_total, _ = _build_section_totals(admin_data)

    admin_commission_data = []
    for emp in admin_employees:
        submissions_qs = BankSubmission.objects.filter(employee=emp, year=year, month=month)
        bank_counts = {s.bank_id: s.submission_count for s in submissions_qs}
        bank_counts_list = [bank_counts.get(b.id, 0) for b in banks]
        commission = _get_commission(year, month, employee=emp, currency=emp.currency)
        admin_commission_data.append({
            'employee': emp,
            'employee_type': 'inhouse',
            'currency': emp.currency,
            'bank_counts_list': bank_counts_list,
            'commission': round(commission, 2),
        })
    total_admin_commission = round(sum(r['commission'] for r in admin_commission_data), 2)

    sales_inhouse_employees = Employee.objects.filter(department='Sales', is_active=True).order_by('name')
    sales_inhouse_data = [
        _get_sales_payroll_row(emp, year, month, 'inhouse', banks, days_in_month, total_holidays)
        for emp in sales_inhouse_employees
    ]
    total_sales_inhouse, _, _, _ = _build_section_totals(sales_inhouse_data)

    all_inhouse_tcr_ids = set(
        Employee.objects.filter(is_active=True).exclude(tcr_id='').values_list('tcr_id', flat=True)
    )
    remote_employees = RemoteEmployee.objects.filter(is_active=True).order_by('name')
    if all_inhouse_tcr_ids:
        remote_employees = remote_employees.exclude(tcr_id__in=all_inhouse_tcr_ids)
    remote_data = [
        _get_sales_payroll_row(emp, year, month, 'remote', banks, days_in_month, total_holidays)
        for emp in remote_employees
    ]
    total_remote, _, _, _ = _build_section_totals(remote_data)

    def _has_base_salary(r):
        return r.get('is_fixed_salary') or r.get('is_attendance_based')

    all_sales_data = sorted(
        sales_inhouse_data + remote_data,
        key=lambda r: (0 if _has_base_salary(r) else 1, r['employee'].name.lower())
    )
    all_sales_rows = sales_inhouse_data + remote_data
    total_sales, sales_incentives_total, sales_reductions_total, _ = _build_section_totals(all_sales_rows)
    total_sales_aed = round(sum(r['net_payroll'] for r in all_sales_rows if r.get('currency', 'AED') == 'AED'), 2)
    total_sales_inr = round(sum(r['net_payroll'] for r in all_sales_rows if r.get('currency', 'AED') == 'INR'), 2)

    def _sales_tfoot(rows, currency):
        return round(
            sum(r.get('commission', 0) for r in rows if not _has_base_salary(r) and r.get('currency', 'AED') == currency) +
            sum(r['net_payroll'] for r in rows if _has_base_salary(r) and r.get('currency', 'AED') == currency),
            2
        )
    total_sales_commission_aed = _sales_tfoot(all_sales_rows, 'AED')
    total_sales_commission_inr = _sales_tfoot(all_sales_rows, 'INR')

    inr_exchange_rate = None
    inr_rate_obj = ExchangeRate.objects.filter(currency='INR', year=year, month=month).first()
    if inr_rate_obj:
        inr_exchange_rate = float(inr_rate_obj.rate)

    for row in all_sales_data:
        amount = row.get('net_payroll') if (_has_base_salary(row)) else row.get('commission', 0)
        if row.get('currency', 'AED') == 'INR' and inr_exchange_rate and inr_exchange_rate > 0:
            row['net_payroll_aed'] = round(amount / inr_exchange_rate, 2)
        else:
            row['net_payroll_aed'] = round(amount, 2)
    total_sales_all_aed = round(sum(r['net_payroll_aed'] for r in all_sales_data), 2)

    _month_names_short = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec',
    }
    _DED_COLS = ['advance', 'visa_status_change', 'clawback', 'leave_deduction', 'late_deduction', 'other_deduction']
    _ADD_COLS = ['last_month_balance', 'paid_leave', 'other_addition']
    _ALL_CATS = _DED_COLS + _ADD_COLS
    target_idx = year * 12 + (month - 1)

    active_by_emp = _defaultdict(list)
    for entry in DeductionEntry.objects.select_related('employee', 'remote_employee').order_by('-created_at'):
        emp = entry.employee or entry.remote_employee
        emp_type = 'inhouse' if entry.employee else 'remote'
        start_idx = entry.start_year * 12 + (entry.start_month - 1)
        if start_idx <= target_idx < start_idx + entry.split_months:
            active_by_emp[(emp_type, emp.id)].append(entry)

    inhouse_summaries = {
        s.employee_id: s
        for s in MonthlySummary.objects.filter(year=year, month=month)
    }
    incoming_carryovers = DeductionCarryover.objects.filter(to_year=year, to_month=month).exclude(is_skipped=True)
    carryover_by_emp = {}
    for co in incoming_carryovers:
        key = ('inhouse', co.employee_id) if co.employee_id else ('remote', co.remote_employee_id)
        carryover_by_emp[key] = co.overflow_amount

    _payroll_inhouse = list(Employee.objects.filter(
        is_active=True, department__in=['Admin', 'Sales']
    ).order_by('department', 'name'))
    _payroll_inhouse_ids = {e.id for e in _payroll_inhouse}
    _payroll_inhouse_names = {e.name: e.id for e in _payroll_inhouse}
    _inhouse_dupe_ids = _defaultdict(list)
    for _emp in Employee.objects.filter(is_active=True).exclude(id__in=_payroll_inhouse_ids):
        main_id = _payroll_inhouse_names.get(_emp.name)
        if main_id:
            _inhouse_dupe_ids[main_id].append(_emp.id)

    _remote_ids = {e.id for e in remote_employees}
    _remote_names = {e.name: e.id for e in remote_employees}
    _remote_dupe_ids = _defaultdict(list)
    for _emp in RemoteEmployee.objects.filter(is_active=True).exclude(id__in=_remote_ids):
        main_id = _remote_names.get(_emp.name)
        if main_id:
            _remote_dupe_ids[main_id].append(_emp.id)

    section3_rows = []
    for emp in _payroll_inhouse:
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get(('inhouse', emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        for dupe_id in _inhouse_dupe_ids.get(emp.id, []):
            for ded_entry in active_by_emp.get(('inhouse', dupe_id), []):
                cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        summary = inhouse_summaries.get(emp.id)
        if emp.salary and emp.payroll_type != 'performance':
            daily = float(emp.salary) / days_in_month
            bridge_count = len(get_bridge_sunday_days(emp, month_start, month_end))
            base_leave_days = (summary.leave_days if summary else 0) or 0
            base_late_days = (summary.late_days if summary else 0) or 0
            cat['leave_deduction'] = round(daily * (base_leave_days + bridge_count), 2)
            cat['late_deduction'] = round(daily * (base_late_days // 3) * 0.5, 2)
        _has_attendance_salary = (
            emp.salary and
            getattr(emp, 'payroll_type', 'attendance') != 'performance'
        )
        if emp.department == 'Admin' or _has_attendance_salary:
            ded_cols_for_total = [c for c in _DED_COLS if c not in ('leave_deduction', 'late_deduction')]
        else:
            ded_cols_for_total = _DED_COLS
        total_ded = round(sum(cat[c] for c in ded_cols_for_total), 2)
        carryover_in = float(carryover_by_emp.get(('inhouse', emp.id), 0))
        for dupe_id in _inhouse_dupe_ids.get(emp.id, []):
            carryover_in += float(carryover_by_emp.get(('inhouse', dupe_id), 0))
        total_ded = round(total_ded + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        row = {'employee': emp, 'employee_type': 'inhouse', 'is_inhouse': True, 'currency': emp.currency}
        row.update(cat)
        row.update({'total_deductions': total_ded, 'total_additions': total_add, 'net': round(total_add - total_ded, 2), 'carryover_in': carryover_in})
        section3_rows.append(row)

    for emp in remote_employees:
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get(('remote', emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        for dupe_id in _remote_dupe_ids.get(emp.id, []):
            for ded_entry in active_by_emp.get(('remote', dupe_id), []):
                cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        total_ded = round(sum(cat[c] for c in _DED_COLS), 2)
        carryover_in = float(carryover_by_emp.get(('remote', emp.id), 0))
        for dupe_id in _remote_dupe_ids.get(emp.id, []):
            carryover_in += float(carryover_by_emp.get(('remote', dupe_id), 0))
        total_ded = round(total_ded + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        row = {'employee': emp, 'employee_type': 'remote', 'is_inhouse': False, 'currency': emp.currency}
        row.update(cat)
        row.update({'total_deductions': total_ded, 'total_additions': total_add, 'net': round(total_add - total_ded, 2), 'carryover_in': carryover_in})
        section3_rows.append(row)

    s3_total_ded = round(sum(r['total_deductions'] for r in section3_rows), 2)
    s3_total_add = round(sum(r['total_additions'] for r in section3_rows), 2)
    s3_net = round(s3_total_add - s3_total_ded, 2)
    s3_total_ded_aed = round(sum(r['total_deductions'] for r in section3_rows if r.get('currency', 'AED') == 'AED'), 2)
    s3_total_ded_inr = round(sum(r['total_deductions'] for r in section3_rows if r.get('currency', 'AED') == 'INR'), 2)
    s3_total_add_aed = round(sum(r['total_additions'] for r in section3_rows if r.get('currency', 'AED') == 'AED'), 2)
    s3_total_add_inr = round(sum(r['total_additions'] for r in section3_rows if r.get('currency', 'AED') == 'INR'), 2)
    s3_net_aed = round(s3_total_add_aed - s3_total_ded_aed, 2)
    s3_net_inr = round(s3_total_add_inr - s3_total_ded_inr, 2)

    payroll_by_emp = {}
    for row in admin_data:
        payroll_by_emp[('inhouse', row['employee'].id)] = row
    for row in all_sales_data:
        payroll_by_emp[(row['employee_type'], row['employee'].id)] = row
    ded_by_emp = {}
    for row in section3_rows:
        ded_by_emp[(row['employee_type'], row['employee'].id)] = row

    if month == 12:
        _co_to_month, _co_to_year = 1, year + 1
    else:
        _co_to_month, _co_to_year = month + 1, year

    final_rows = []
    for emp in _payroll_inhouse:
        key = ('inhouse', emp.id)
        p = payroll_by_emp.get(key)
        d = ded_by_emp.get(key)
        payroll_net = p['net_payroll'] if p else 0.0
        total_ded = d['total_deductions'] if d else 0.0
        total_add = d['total_additions'] if d else 0.0
        carryover_in = d['carryover_in'] if d else 0.0
        final_salary = round(payroll_net - total_ded + total_add, 2)
        if final_salary < 0:
            overflow = Decimal(str(abs(final_salary)))
            final_salary = 0.0
            DeductionCarryover.objects.update_or_create(
                employee=emp, from_year=year, from_month=month,
                defaults={'overflow_amount': overflow, 'to_year': _co_to_year, 'to_month': _co_to_month,
                          'remote_employee': None, 'currency': emp.currency},
            )
        else:
            DeductionCarryover.objects.filter(employee=emp, from_year=year, from_month=month).delete()
        if carryover_in > 0:
            inc_co = incoming_carryovers.filter(employee=emp).first()
            if inc_co:
                inc_co.applied_amount = Decimal(str(min(carryover_in, float(inc_co.overflow_amount))))
                inc_co.save(update_fields=['applied_amount'])
        final_rows.append({
            'employee': emp,
            'employee_type': 'inhouse',
            'department': emp.department or 'In-House',
            'currency': emp.currency,
            'payroll_net': payroll_net,
            'total_deductions': total_ded,
            'total_additions': total_add,
            'final_salary': final_salary,
        })

    for emp in remote_employees:
        key = ('remote', emp.id)
        p = payroll_by_emp.get(key)
        d = ded_by_emp.get(key)
        payroll_net = p['net_payroll'] if p else 0.0
        total_ded = d['total_deductions'] if d else 0.0
        total_add = d['total_additions'] if d else 0.0
        carryover_in = d['carryover_in'] if d else 0.0
        final_salary = round(payroll_net - total_ded + total_add, 2)
        if final_salary < 0:
            overflow = Decimal(str(abs(final_salary)))
            final_salary = 0.0
            DeductionCarryover.objects.update_or_create(
                remote_employee=emp, from_year=year, from_month=month,
                defaults={'overflow_amount': overflow, 'to_year': _co_to_year, 'to_month': _co_to_month,
                          'employee': None, 'currency': emp.currency},
            )
        else:
            DeductionCarryover.objects.filter(remote_employee=emp, from_year=year, from_month=month).delete()
        if carryover_in > 0:
            inc_co = incoming_carryovers.filter(remote_employee=emp).first()
            if inc_co:
                inc_co.applied_amount = Decimal(str(min(carryover_in, float(inc_co.overflow_amount))))
                inc_co.save(update_fields=['applied_amount'])
        final_rows.append({
            'employee': emp,
            'employee_type': 'remote',
            'department': 'Remote',
            'currency': emp.currency,
            'payroll_net': payroll_net,
            'total_deductions': total_ded,
            'total_additions': total_add,
            'final_salary': final_salary,
        })

    final_total_aed = round(sum(r['final_salary'] for r in final_rows if r.get('currency', 'AED') == 'AED'), 2)
    final_total_inr = round(sum(r['final_salary'] for r in final_rows if r.get('currency', 'AED') == 'INR'), 2)
    final_total = round(sum(r['final_salary'] for r in final_rows), 2)
    if inr_exchange_rate and inr_exchange_rate > 0 and final_total_inr > 0:
        final_total_combined_aed = round(final_total_aed + (final_total_inr / inr_exchange_rate), 2)
    else:
        final_total_combined_aed = final_total_aed

    all_rows = admin_data + all_sales_data
    grand_total_aed = round(sum(r['net_payroll'] for r in all_rows if r.get('currency', 'AED') == 'AED'), 2)
    grand_total_inr = round(sum(r['net_payroll'] for r in all_rows if r.get('currency', 'AED') == 'INR'), 2)
    grand_total = round(total_admin + total_sales, 2)

    snapshot_ctx = {
        'total_holidays': total_holidays,
        'inr_exchange_rate': inr_exchange_rate,
        'banks': banks,
        'admin_data': admin_data,
        'total_admin': total_admin,
        'admin_incentives_total': admin_incentives_total,
        'admin_reductions_total': admin_reductions_total,
        'admin_commission_data': admin_commission_data,
        'total_admin_commission': total_admin_commission,
        'all_sales_data': all_sales_data,
        'total_sales_inhouse': total_sales_inhouse,
        'total_remote': total_remote,
        'total_sales': total_sales,
        'total_sales_aed': total_sales_aed,
        'total_sales_inr': total_sales_inr,
        'total_sales_commission_aed': total_sales_commission_aed,
        'total_sales_commission_inr': total_sales_commission_inr,
        'sales_incentives_total': sales_incentives_total,
        'sales_reductions_total': sales_reductions_total,
        'total_sales_all_aed': total_sales_all_aed,
        'section3_rows': section3_rows,
        's3_total_ded': s3_total_ded,
        's3_total_add': s3_total_add,
        's3_net': s3_net,
        's3_total_ded_aed': s3_total_ded_aed,
        's3_total_ded_inr': s3_total_ded_inr,
        's3_total_add_aed': s3_total_add_aed,
        's3_total_add_inr': s3_total_add_inr,
        's3_net_aed': s3_net_aed,
        's3_net_inr': s3_net_inr,
        'final_rows': final_rows,
        'final_total': final_total,
        'final_total_aed': final_total_aed,
        'final_total_inr': final_total_inr,
        'final_total_combined_aed': final_total_combined_aed,
        'grand_total': grand_total,
        'grand_total_aed': grand_total_aed,
        'grand_total_inr': grand_total_inr,
    }
    snapshot = _serialize_payroll_context(snapshot_ctx)
    now = datetime.datetime.now(tz=datetime.timezone.utc)
    FrozenPayrollMonth.objects.create(
        year=year, month=month,
        frozen_at=now,
        frozen_by=request.user.username,
        snapshot=snapshot,
    )
    logger.info("Payroll frozen for %s/%s by %s", month, year, request.user.username)
    return JsonResponse({'success': True, 'frozen_at': now.isoformat()})


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def unfreeze_payroll(request):
    """Delete the frozen snapshot for a month, reverting to live calculation."""
    try:
        data = json.loads(request.body)
        year = int(data['year'])
        month = int(data['month'])
        password = data.get('password', '')
    except (json.JSONDecodeError, KeyError, ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

    if not request.user.check_password(password):
        return JsonResponse({'success': False, 'error': 'Incorrect password.'}, status=403)

    deleted, _ = FrozenPayrollMonth.objects.filter(year=year, month=month).delete()
    if not deleted:
        return JsonResponse({'success': False, 'error': 'This month is not frozen.'}, status=400)

    logger.info("Payroll unfrozen for %s/%s by %s", month, year, request.user.username)
    return JsonResponse({'success': True})


# ============================================
# Bank Management
# ============================================

@login_required
@user_passes_test(section_required('banks'), login_url='/report/')
def manage_banks(request):
    """Bank management page — list/add/edit/deactivate banks."""
    banks = Bank.objects.all().order_by('name')
    existing_tiers = {t.currency: t for t in CommissionTierSettings.objects.filter(currency__in=FOREIGN_CURRENCIES)}
    tier_rows = []
    for currency in FOREIGN_CURRENCIES:
        t = existing_tiers.get(currency)
        tier_rows.append({
            'currency': currency,
            'threshold': t.threshold if t else DEFAULT_TIER_THRESHOLD,
            'overflow_rate': t.overflow_rate if t else None,
        })
    return render(request, 'payroll/banks.html', {'banks': banks, 'tier_rows': tier_rows})


@login_required
@user_passes_test(section_required('banks'), login_url='/report/')
@require_http_methods(["GET", "POST"])
def banks_api(request):
    """GET: list active banks. POST: add a new bank."""
    if request.method == 'GET':
        banks = Bank.objects.filter(is_active=True).order_by('name')
        data = [{'id': b.id, 'name': b.name, 'per_account_charge': float(b.per_account_charge)} for b in banks]
        return JsonResponse({'success': True, 'banks': data})

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    name = data.get('name', '').strip()
    per_account_charge = data.get('per_account_charge')
    inr_per_account_charge = data.get('inr_per_account_charge')
    npr_per_account_charge = data.get('npr_per_account_charge')

    if not name or per_account_charge is None:
        return JsonResponse({'success': False, 'error': 'Name and per_account_charge are required'}, status=400)

    if Bank.objects.filter(name__iexact=name).exists():
        return JsonResponse({'success': False, 'error': 'A bank with this name already exists'}, status=400)

    try:
        inr_charge = Decimal(str(inr_per_account_charge)) if inr_per_account_charge not in (None, '', 0, '0') else None
        npr_charge = Decimal(str(npr_per_account_charge)) if npr_per_account_charge not in (None, '', 0, '0') else None
        bank = Bank.objects.create(
            name=name,
            per_account_charge=Decimal(str(per_account_charge)),
            inr_per_account_charge=inr_charge,
            npr_per_account_charge=npr_charge,
        )
    except (ValueError, TypeError) as e:
        return JsonResponse({'success': False, 'error': f'Invalid data: {e}'}, status=400)

    logger.info("Bank added: %s (AED %s/account) by %s", bank.name, bank.per_account_charge, request.user.username)
    return JsonResponse({'success': True, 'bank': {
        'id': bank.id, 'name': bank.name,
        'per_account_charge': float(bank.per_account_charge),
        'inr_per_account_charge': float(bank.inr_per_account_charge) if bank.inr_per_account_charge else None,
        'npr_per_account_charge': float(bank.npr_per_account_charge) if bank.npr_per_account_charge else None,
        'is_active': bank.is_active,
    }})


@login_required
@user_passes_test(section_required('banks'), login_url='/report/')
@require_http_methods(["POST"])
def bank_detail_api(request, bank_id):
    """Update or toggle a bank."""
    try:
        bank = Bank.objects.get(id=bank_id)
    except Bank.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Bank not found'}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    action = data.get('action')

    if action == 'update':
        name = data.get('name', '').strip()
        per_account_charge = data.get('per_account_charge')
        inr_per_account_charge = data.get('inr_per_account_charge')
        npr_per_account_charge = data.get('npr_per_account_charge')
        if not name or per_account_charge is None:
            return JsonResponse({'success': False, 'error': 'Name and charge are required'}, status=400)
        # Check uniqueness excluding self
        if Bank.objects.filter(name__iexact=name).exclude(id=bank_id).exists():
            return JsonResponse({'success': False, 'error': 'Another bank with this name already exists'}, status=400)
        try:
            bank.name = name
            bank.per_account_charge = Decimal(str(per_account_charge))
            bank.inr_per_account_charge = (
                Decimal(str(inr_per_account_charge))
                if inr_per_account_charge not in (None, '', 0, '0')
                else None
            )
            bank.npr_per_account_charge = (
                Decimal(str(npr_per_account_charge))
                if npr_per_account_charge not in (None, '', 0, '0')
                else None
            )
            bank.save()
        except (ValueError, TypeError) as e:
            return JsonResponse({'success': False, 'error': f'Invalid data: {e}'}, status=400)
        logger.info("Bank updated: %s by %s", bank.name, request.user.username)
        return JsonResponse({'success': True, 'bank': {
            'id': bank.id, 'name': bank.name,
            'per_account_charge': float(bank.per_account_charge),
            'inr_per_account_charge': float(bank.inr_per_account_charge) if bank.inr_per_account_charge else None,
            'npr_per_account_charge': float(bank.npr_per_account_charge) if bank.npr_per_account_charge else None,
            'is_active': bank.is_active,
        }})

    elif action == 'toggle':
        bank.is_active = not bank.is_active
        bank.save()
        logger.info("Bank %s: %s by %s", 'activated' if bank.is_active else 'deactivated', bank.name, request.user.username)
        return JsonResponse({'success': True, 'is_active': bank.is_active})

    return JsonResponse({'success': False, 'error': 'Unknown action'}, status=400)


@login_required
@user_passes_test(section_required('banks'), login_url='/report/')
@require_http_methods(["POST"])
def save_commission_tier(request):
    """Save/update the tiered-commission threshold + flat overflow rate for a foreign currency (INR, NPR, ...)."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    currency = (data.get('currency') or '').strip().upper()
    threshold = data.get('threshold')
    overflow_rate = data.get('overflow_rate')

    if currency not in FOREIGN_CURRENCIES:
        return JsonResponse({'success': False, 'error': 'Currency must be one of: ' + ', '.join(FOREIGN_CURRENCIES)}, status=400)
    if threshold is None or overflow_rate is None:
        return JsonResponse({'success': False, 'error': 'threshold and overflow_rate are required'}, status=400)

    try:
        threshold_val = int(threshold)
        overflow_val = Decimal(str(overflow_rate))
        if threshold_val < 0 or overflow_val < 0:
            raise ValueError
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'threshold must be a non-negative integer and overflow_rate a non-negative number'}, status=400)

    obj, created = CommissionTierSettings.objects.update_or_create(
        currency=currency,
        defaults={'threshold': threshold_val, 'overflow_rate': overflow_val},
    )
    logger.info(
        "Commission tier settings saved for %s: threshold=%s overflow_rate=%s by %s",
        currency, threshold_val, overflow_val, request.user.username,
    )
    return JsonResponse({
        'success': True, 'currency': currency,
        'threshold': obj.threshold, 'overflow_rate': float(obj.overflow_rate), 'created': created,
    })


# ============================================
# API: Bank Submissions
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def get_submissions(request, emp_type, employee_id):
    """Get bank submissions for an employee for a specific month."""
    try:
        year = int(request.GET.get('year', datetime.datetime.now().year))
        month = int(request.GET.get('month', datetime.datetime.now().month))
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'})

    if emp_type == 'inhouse':
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        submissions_qs = BankSubmission.objects.filter(
            employee=employee, year=year, month=month
        ).select_related('bank')
    else:
        try:
            employee = RemoteEmployee.objects.get(id=employee_id)
        except RemoteEmployee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        submissions_qs = BankSubmission.objects.filter(
            remote_employee=employee, year=year, month=month
        ).select_related('bank')

    submission_map = {s.bank_id: s.submission_count for s in submissions_qs}
    banks = list(Bank.objects.filter(is_active=True).order_by('name'))
    emp_currency = employee.currency if hasattr(employee, 'currency') else 'AED'

    if emp_currency != 'AED':
        threshold, overflow_rate = _get_tier_settings(emp_currency)
        pairs = [(submission_map.get(b.id, 0), b.charge_for_currency(emp_currency)) for b in banks]
        total_commission, per_bank_commissions = _calc_tiered_commission(pairs, threshold, overflow_rate)
        data = [
            {
                'bank_id': b.id,
                'bank_name': b.name,
                'per_account_charge': float(b.charge_for_currency(emp_currency)),
                'currency': emp_currency,
                'submission_count': submission_map.get(b.id, 0),
                'commission': round(per_bank_commissions[i], 2),
            }
            for i, b in enumerate(banks)
        ]
    else:
        data = []
        total_commission = 0.0
        for bank in banks:
            count = submission_map.get(bank.id, 0)
            charge = float(bank.per_account_charge)
            commission = round(count * charge, 2)
            total_commission += commission
            data.append({
                'bank_id': bank.id,
                'bank_name': bank.name,
                'per_account_charge': charge,
                'currency': emp_currency,
                'submission_count': count,
                'commission': commission,
            })

    return JsonResponse({
        'success': True,
        'employee_name': employee.name,
        'currency': emp_currency,
        'banks': data,
        'total_commission': round(total_commission, 2),
    })


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def save_submissions(request):
    """Save bank submission counts for an employee for a month."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    emp_type = data.get('emp_type') or data.get('employee_type')
    employee_id = data.get('employee_id')
    year = data.get('year')
    month = data.get('month')
    submissions_raw = data.get('submissions', [])
    # Accept both list [{bank_id, count}] and dict {bank_id: count} formats
    if isinstance(submissions_raw, list):
        submissions = {str(s['bank_id']): s['count'] for s in submissions_raw if 'bank_id' in s}
    else:
        submissions = submissions_raw

    if not all([emp_type, employee_id, year, month]):
        return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

    if emp_type == 'inhouse':
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        fk_kwargs = {'employee': employee}
    else:
        try:
            employee = RemoteEmployee.objects.get(id=employee_id)
        except RemoteEmployee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        fk_kwargs = {'remote_employee': employee}

    emp_currency = employee.currency if hasattr(employee, 'currency') else 'AED'

    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'}, status=400)

    for bank_id_str, count_val in submissions.items():
        try:
            bank_id = int(bank_id_str)
            count = int(count_val)
            bank = Bank.objects.get(id=bank_id, is_active=True)
        except (ValueError, TypeError, Bank.DoesNotExist):
            continue

        if count <= 0:
            BankSubmission.objects.filter(bank=bank, year=year, month=month, **fk_kwargs).delete()
        else:
            BankSubmission.objects.update_or_create(
                bank=bank, year=year, month=month, **fk_kwargs,
                defaults={'submission_count': count},
            )

    # Re-calculate total commission after saving using the same tiered logic as the dashboard
    total_commission = _get_commission(year, month, currency=emp_currency, **{
        'employee' if emp_type == 'inhouse' else 'remote_employee': employee
    })

    logger.info("Bank submissions saved for %s by %s", employee.name, request.user.username)
    return JsonResponse({'success': True, 'total_commission': round(total_commission, 2)})


# ============================================
# API: In-house Employee Adjustments
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def get_adjustments(request, employee_id):
    """Get adjustments for an in-house employee for a specific month."""
    try:
        year = int(request.GET.get('year', datetime.datetime.now().year))
        month = int(request.GET.get('month', datetime.datetime.now().month))
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'})

    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

    adjustments = PayrollAdjustment.objects.filter(
        employee=employee, year=year, month=month
    )
    data = [{
        'id': adj.id,
        'type': adj.adjustment_type,
        'amount': float(adj.amount),
        'reason': adj.reason,
        'created_at': adj.created_at.strftime('%Y-%m-%d %H:%M'),
    } for adj in adjustments]

    return JsonResponse({
        'success': True,
        'employee_name': employee.name,
        'adjustments': data,
    })


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def add_adjustment(request):
    """Add a new adjustment for an in-house employee."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    employee_id = data.get('employee_id')
    year = data.get('year')
    month = data.get('month')
    adjustment_type = data.get('type')
    amount = data.get('amount')
    reason = data.get('reason', '')

    if not all([employee_id, year, month, adjustment_type, amount]):
        return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

    if adjustment_type not in ('incentive', 'reduction'):
        return JsonResponse({'success': False, 'error': 'Invalid adjustment type'}, status=400)

    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

    try:
        adjustment = PayrollAdjustment.objects.create(
            employee=employee,
            year=int(year),
            month=int(month),
            adjustment_type=adjustment_type,
            amount=Decimal(str(amount)),
            reason=reason,
        )
    except (ValueError, TypeError) as e:
        return JsonResponse({'success': False, 'error': f'Invalid data: {e}'}, status=400)

    logger.info(
        "Payroll adjustment added: %s %s %s by %s",
        employee.name, adjustment_type, amount, request.user.username
    )
    return JsonResponse({
        'success': True,
        'message': 'Adjustment added successfully',
        'adjustment': {
            'id': adjustment.id,
            'type': adjustment.adjustment_type,
            'amount': float(adjustment.amount),
            'reason': adjustment.reason,
        },
    })


# ============================================
# API: Remote Employee Adjustments
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def get_remote_adjustments(request, employee_id):
    """Get adjustments for a remote employee for a specific month."""
    try:
        year = int(request.GET.get('year', datetime.datetime.now().year))
        month = int(request.GET.get('month', datetime.datetime.now().month))
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'})

    try:
        employee = RemoteEmployee.objects.get(id=employee_id)
    except RemoteEmployee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

    adjustments = PayrollAdjustment.objects.filter(
        remote_employee=employee, year=year, month=month
    )
    data = [{
        'id': adj.id,
        'type': adj.adjustment_type,
        'amount': float(adj.amount),
        'reason': adj.reason,
        'created_at': adj.created_at.strftime('%Y-%m-%d %H:%M'),
    } for adj in adjustments]

    return JsonResponse({
        'success': True,
        'employee_name': employee.name,
        'adjustments': data,
    })


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def add_remote_adjustment(request):
    """Add a new adjustment for a remote employee."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    employee_id = data.get('employee_id')
    year = data.get('year')
    month = data.get('month')
    adjustment_type = data.get('type')
    amount = data.get('amount')
    reason = data.get('reason', '')

    if not all([employee_id, year, month, adjustment_type, amount]):
        return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

    if adjustment_type not in ('incentive', 'reduction'):
        return JsonResponse({'success': False, 'error': 'Invalid adjustment type'}, status=400)

    try:
        employee = RemoteEmployee.objects.get(id=employee_id)
    except RemoteEmployee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

    try:
        adjustment = PayrollAdjustment.objects.create(
            remote_employee=employee,
            year=int(year),
            month=int(month),
            adjustment_type=adjustment_type,
            amount=Decimal(str(amount)),
            reason=reason,
        )
    except (ValueError, TypeError) as e:
        return JsonResponse({'success': False, 'error': f'Invalid data: {e}'}, status=400)

    logger.info(
        "Remote payroll adjustment added: %s %s %s by %s",
        employee.name, adjustment_type, amount, request.user.username
    )
    return JsonResponse({
        'success': True,
        'message': 'Adjustment added successfully',
        'adjustment': {
            'id': adjustment.id,
            'type': adjustment.adjustment_type,
            'amount': float(adjustment.amount),
            'reason': adjustment.reason,
        },
    })


# ============================================
# API: Recalculate Summaries
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def recalculate_summaries(request):
    """Trigger recalculation of monthly summaries for the selected month/year."""
    try:
        data = json.loads(request.body)
        year = int(data.get('year'))
        month = int(data.get('month'))
    except (json.JSONDecodeError, TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'}, status=400)

    if not (1 <= month <= 12) or not (2000 <= year <= 2099):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'}, status=400)

    call_command('recalculate_summaries', year, month, verbosity=0)
    logger.info("Payroll summaries recalculated for %s/%s by %s", year, month, request.user.username)
    return JsonResponse({'success': True})


# ============================================
# API: Delete Adjustment (both types)
# ============================================

# ============================================
# Upload: Bank Submissions XLSX
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def upload_submissions(request):
    """
    Parse a Target vs Achieved XLSX file and upsert BankSubmission records.

    Expected columns (any order, header detection by keyword):
      - Id      → employee TCR ID (optional; used when present)
      - Agent   → employee name (fallback lookup when Id is absent/empty)
      - {BankName} Ach → achieved submission count for each bank

    Each row = one employee. The achieved count for each bank is read directly
    from the "{Bank} Ach" column (Total Ach is ignored).
    Employees are looked up by tcr_id (via Id column) or name (via Agent column)
    across both Employee and RemoteEmployee.
    """
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

    try:
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid year/month'}, status=400)

    if not (1 <= month <= 12) or not (2000 <= year <= 2099):
        return JsonResponse({'success': False, 'error': 'Invalid year/month range'}, status=400)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Could not read file: {e}'}, status=400)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return JsonResponse({'success': False, 'error': 'File is empty'}, status=400)

    # Detect columns from header row
    raw_headers = [str(c).strip() if c else '' for c in rows[0]]
    headers_lower = [h.lower() for h in raw_headers]

    agent_col = id_col = None
    # {bank_name_lower: col_index} for "{Bank} Ach" columns (excluding "total ach")
    bank_ach_cols = {}

    for i, h in enumerate(headers_lower):
        if h == 'id':
            id_col = i
        elif h == 'agent':
            agent_col = i
        elif h.endswith(' ach') and not h.startswith('total'):
            bank_name = raw_headers[i][:-4].strip()  # strip " Ach" suffix
            bank_ach_cols[bank_name.lower()] = (i, bank_name)

    if agent_col is None:
        return JsonResponse(
            {'success': False, 'error': 'Could not find "Agent" column in header'},
            status=400
        )
    if not bank_ach_cols:
        return JsonResponse(
            {'success': False, 'error': 'Could not find any "{Bank} Ach" columns in header'},
            status=400
        )

    # Preload active banks for fast lookup (case-insensitive)
    bank_map = {b.name.lower(): b for b in Bank.objects.filter(is_active=True)}

    matched = 0
    unmatched_agents = set()
    unmatched_banks = set()

    for row in rows[1:]:
        if not any(row):
            continue

        agent_name = str(row[agent_col]).strip() if row[agent_col] else ''
        tcr_id = str(row[id_col]).strip() if (id_col is not None and row[id_col]) else ''

        if not agent_name and not tcr_id:
            continue

        # Employee lookup: prefer tcr_id, fall back to name
        employee = remote_employee = None
        if tcr_id:
            employee = Employee.objects.filter(tcr_id=tcr_id, is_active=True).first()
            if not employee:
                remote_employee = RemoteEmployee.objects.filter(tcr_id=tcr_id, is_active=True).first()
        if not employee and not remote_employee and agent_name:
            employee = Employee.objects.filter(name__iexact=agent_name, is_active=True).first()
            if not employee:
                remote_employee = RemoteEmployee.objects.filter(name__iexact=agent_name, is_active=True).first()

        if not employee and not remote_employee:
            unmatched_agents.add(tcr_id or agent_name)
            continue

        fk_kwargs = {'employee': employee} if employee else {'remote_employee': remote_employee}

        for bank_lower, (col_idx, bank_display_name) in bank_ach_cols.items():
            try:
                count = int(row[col_idx]) if row[col_idx] is not None else 0
            except (ValueError, TypeError):
                count = 0

            bank = bank_map.get(bank_lower)
            if not bank:
                unmatched_banks.add(bank_display_name)
                continue

            if count <= 0:
                BankSubmission.objects.filter(bank=bank, year=year, month=month, **fk_kwargs).delete()
            else:
                BankSubmission.objects.update_or_create(
                    bank=bank, year=year, month=month, **fk_kwargs,
                    defaults={'submission_count': count},
                )
            matched += 1

    logger.info(
        "Submission upload for %d/%d: %d saved, %d unmatched agents, %d unmatched banks by %s",
        year, month, matched, len(unmatched_agents), len(unmatched_banks), request.user.username,
    )

    return JsonResponse({
        'success': True,
        'stats': {
            'matched': matched,
            'unmatched_agents': sorted(unmatched_agents),
            'unmatched_banks': sorted(unmatched_banks),
        },
    })


# ============================================
# Payroll Employee Database
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def payroll_employees(request):
    """Redirects to the unified employee management page."""
    return redirect('employee_management')


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def payroll_employee_update(request, emp_type, employee_id):
    """Update payroll-relevant fields (salary, currency, designation, department) for an employee."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    if emp_type == 'inhouse':
        try:
            emp = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
    elif emp_type == 'remote':
        try:
            emp = RemoteEmployee.objects.get(id=employee_id)
        except RemoteEmployee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid employee type'}, status=400)

    if 'salary' in data:
        val = data['salary']
        try:
            emp.salary = Decimal(str(val)) if val not in (None, '', 0) else None
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid salary value'}, status=400)

    old_currency = emp.currency
    if 'currency' in data:
        if data['currency'] not in ('AED', 'INR', 'NPR'):
            return JsonResponse({'success': False, 'error': 'Currency must be AED, INR, or NPR'}, status=400)
        emp.currency = data['currency']

    if 'designation' in data:
        emp.designation = str(data['designation']).strip() or None

    if 'department' in data:
        dept = str(data['department']).strip()
        if dept not in ('Sales', 'Admin', ''):
            return JsonResponse({'success': False, 'error': 'Invalid department'}, status=400)
        emp.department = dept or None

    if 'payroll_type' in data:
        pt = str(data['payroll_type']).strip()
        if pt not in ('attendance', 'performance'):
            return JsonResponse({'success': False, 'error': 'Invalid payroll type'}, status=400)
        emp.payroll_type = pt

    if 'is_fixed_salary' in data:
        emp.is_fixed_salary = bool(data['is_fixed_salary'])

    if 'salary_cycle_start_day' in data:
        try:
            cycle_day = int(data['salary_cycle_start_day'])
            if not (1 <= cycle_day <= 28):
                return JsonResponse({'success': False, 'error': 'salary_cycle_start_day must be 1–28'}, status=400)
            emp.salary_cycle_start_day = cycle_day
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid salary_cycle_start_day'}, status=400)

    if 'visa_provider' in data:
        vp = str(data['visa_provider']).strip()
        if vp and vp not in ('Jumbo', 'OnTime', 'Taamul'):
            return JsonResponse({'success': False, 'error': 'Invalid visa provider'}, status=400)
        emp.visa_provider = vp or None

    emp.save()
    if old_currency != emp.currency:
        convert_employee_deduction_currency(emp_type, emp.id, old_currency, emp.currency)
        logger.info(
            "Currency changed for %s (%s): %s -> %s, outstanding deductions converted by %s",
            emp.name, emp_type, old_currency, emp.currency, request.user.username,
        )
    logger.info("Payroll employee updated: %s (%s) by %s", emp.name, emp_type, request.user.username)
    return JsonResponse({
        'success': True,
        'salary': float(emp.salary) if emp.salary else None,
        'currency': emp.currency,
        'designation': emp.designation or '',
        'department': emp.department or '',
        'payroll_type': emp.payroll_type,
        'is_fixed_salary': emp.is_fixed_salary,
        'visa_provider': emp.visa_provider or '',
        'salary_cycle_start_day': emp.salary_cycle_start_day,
    })


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def delete_adjustment(request, adjustment_id):
    """Delete an adjustment (works for both in-house and remote)."""
    try:
        adjustment = PayrollAdjustment.objects.get(id=adjustment_id)
    except PayrollAdjustment.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Adjustment not found'}, status=404)

    logger.info(
        "Payroll adjustment deleted: id=%s by %s",
        adjustment_id, request.user.username
    )
    adjustment.delete()
    return JsonResponse({'success': True, 'message': 'Adjustment deleted'})


# ============================================
# API: Deductions & Additions
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def add_deduction(request):
    """Create a new DeductionEntry (deduction or addition) for an employee."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    emp_type = data.get('emp_type')
    employee_id = data.get('employee_id')
    category = data.get('category')
    total_amount = data.get('total_amount')
    split_months = data.get('split_months', 1)
    start_year = data.get('start_year')
    start_month = data.get('start_month')
    note = data.get('note', '')

    if not all([emp_type, employee_id, category, total_amount, start_year, start_month]):
        return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

    valid_categories = {c[0] for c in DEDUCTION_CATEGORY_CHOICES}
    if category not in valid_categories:
        return JsonResponse({'success': False, 'error': 'Invalid category'}, status=400)

    if emp_type == 'inhouse':
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        fk_kwargs = {'employee': employee}
    elif emp_type == 'remote':
        try:
            employee = RemoteEmployee.objects.get(id=employee_id)
        except RemoteEmployee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        fk_kwargs = {'remote_employee': employee}
    else:
        return JsonResponse({'success': False, 'error': 'Invalid employee type'}, status=400)

    try:
        entry = DeductionEntry.objects.create(
            **fk_kwargs,
            category=category,
            total_amount=Decimal(str(total_amount)),
            currency=employee.currency,
            split_months=max(1, int(split_months)),
            start_year=int(start_year),
            start_month=int(start_month),
            note=note,
        )
    except (ValueError, TypeError) as e:
        return JsonResponse({'success': False, 'error': f'Invalid data: {e}'}, status=400)

    logger.info(
        "DeductionEntry added: %s %s %s by %s",
        employee.name, category, total_amount, request.user.username,
    )
    return JsonResponse({'success': True})


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def delete_deduction_entry(request, deduction_id):
    """Delete a DeductionEntry."""
    try:
        entry = DeductionEntry.objects.get(id=deduction_id)
    except DeductionEntry.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Deduction not found'}, status=404)

    logger.info("DeductionEntry deleted: id=%s by %s", deduction_id, request.user.username)
    entry.delete()
    return JsonResponse({'success': True})


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def toggle_carryover_skip(request, carryover_id):
    """Waive (or restore) a single carried-over deduction month.

    Skipping excludes the record from live deduction totals (it is treated as
    forgiven, not recovered) without deleting its history. Toggling back
    un-skips it, restoring normal recovery.
    """
    try:
        co = DeductionCarryover.objects.get(id=carryover_id)
    except DeductionCarryover.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Carryover record not found'}, status=404)

    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        data = {}
    skip = data.get('skip', not co.is_skipped)

    co.is_skipped = bool(skip)
    if co.is_skipped:
        co.skipped_at = timezone.now()
        co.skipped_by = request.user.username
        co.skip_reason = (data.get('reason') or '')[:255]
    else:
        co.skipped_at = None
        co.skipped_by = ''
        co.skip_reason = ''
    co.save(update_fields=['is_skipped', 'skipped_at', 'skipped_by', 'skip_reason'])

    emp = co.employee or co.remote_employee
    logger.info(
        "DeductionCarryover %s: id=%s employee=%s %s/%s by %s",
        'skipped' if co.is_skipped else 'unskipped',
        carryover_id, emp.name if emp else '?', co.from_month, co.from_year, request.user.username,
    )
    return JsonResponse({'success': True, 'is_skipped': co.is_skipped})


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["GET"])
def autofill_deduction(request):
    """Return auto-computed leave/late deduction amounts for an in-house employee."""
    try:
        employee_id = int(request.GET.get('employee_id'))
        year = int(request.GET.get('year'))
        month = int(request.GET.get('month'))
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid parameters'}, status=400)

    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)

    _, days_in_month = calendar.monthrange(year, month)
    summary = MonthlySummary.objects.filter(employee=employee, year=year, month=month).first()

    if not summary or not employee.salary or employee.payroll_type == 'performance':
        return JsonResponse({'success': True, 'leave_amount': 0.0, 'late_amount': 0.0})

    salary = float(employee.salary)
    daily_rate = salary / days_in_month
    absent_days = summary.leave_days or 0
    late_days = summary.late_days or 0
    late_half_days = late_days // 3

    return JsonResponse({
        'success': True,
        'leave_amount': round(daily_rate * absent_days, 2),
        'late_amount': round(daily_rate * late_half_days * 0.5, 2),
        'daily_rate': round(daily_rate, 2),
        'absent_days': absent_days,
        'late_days': late_days,
        'late_half_days': late_half_days,
    })


# ============================================
# Payslip Download
# ============================================

def _amount_in_words(amount):
    """Convert a numeric amount (integer AED) to English words for payslips."""
    ones = [
        '', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
        'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
        'Seventeen', 'Eighteen', 'Nineteen',
    ]
    tens_words = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def below_thousand(n):
        if n == 0:
            return ''
        if n < 20:
            return ones[n]
        if n < 100:
            rem = ones[n % 10]
            return tens_words[n // 10] + (' ' + rem if rem else '')
        rem = below_thousand(n % 100)
        return ones[n // 100] + ' Hundred' + (' ' + rem if rem else '')

    n = int(round(abs(amount)))
    if n == 0:
        return 'Zero Only'

    parts = []
    if n >= 1_000_000:
        parts.append(below_thousand(n // 1_000_000) + ' Million')
        n %= 1_000_000
    if n >= 1000:
        parts.append(below_thousand(n // 1000) + ' Thousand')
        n %= 1000
    if n > 0:
        parts.append(below_thousand(n))
    return ' '.join(parts) + ' Only'


@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def download_payslip(request, emp_type, emp_id):
    """Render a printable HTML payslip for one employee for the selected month."""
    selected_month, selected_year = get_selected_month_year(request)
    month_name = MONTH_NAMES[selected_month]

    if emp_type == 'inhouse':
        emp = get_object_or_404(Employee, id=emp_id)
        paid_record = PaidSalaryRecord.objects.filter(
            employee=emp, year=selected_year, month=selected_month
        ).first()
    else:
        emp = get_object_or_404(RemoteEmployee, id=emp_id)
        paid_record = PaidSalaryRecord.objects.filter(
            remote_employee=emp, year=selected_year, month=selected_month
        ).first()

    if paid_record and paid_record.snapshot:
        # Employee's salary for this month is locked via "Mark as Paid" — render
        # from that immutable snapshot instead of recomputing live, so the payslip
        # always matches the locked Payroll Summary figure even if attendance,
        # deductions, salary, or the payroll formula change afterward.
        snap = paid_record.snapshot
        salary = float(snap.get('salary') or 0)
        days_in_month = snap.get('days_in_period') or 0
        absent_days_display = snap.get('absent_days') or 0
        # Locked/paid snapshots never captured a per-component salary
        # breakdown (only the flat gross) — this branch is intentionally left
        # exactly as it always computed, per the rule that PaidSalaryRecord
        # snapshots are never retroactively altered. housing/transport/phone
        # are always 0 here; other_allowance_full carries the same 60% figure
        # this locked payslip always showed.
        if salary:
            basic_full = round(salary * 0.40, 2)
            other_allowance_full = round(salary * 0.60, 2)
        else:
            basic_full = 0.0
            other_allowance_full = 0.0
        housing_full = 0.0
        transport_full = 0.0
        phone_full = 0.0
        commission = 0.0 if snap.get('is_fixed_salary') else snap.get('commission', 0.0)
        incentives = snap.get('incentives', 0.0)
        cat = snap.get('deductions_breakdown', {})
        advance_ded = round(cat.get('advance', 0.0), 2)
        leave_deduction = round(cat.get('leave_deduction', 0.0), 2)
        late_deduction = round(cat.get('late_deduction', 0.0), 2)
        other_ded = round(
            cat.get('other_deduction', 0.0) + cat.get('visa_status_change', 0.0)
            + cat.get('clawback', 0.0) + snap.get('reductions', 0.0), 2
        )
        additions = round(
            cat.get('paid_leave', 0.0) + cat.get('last_month_balance', 0.0)
            + cat.get('other_addition', 0.0), 2
        )
        carryover_ded = round(snap.get('carryover_in', 0.0), 2)
        incentives_commission = round(incentives + commission, 2)
        total_earnings = round(basic_full + housing_full + transport_full + phone_full + other_allowance_full + incentives_commission + additions, 2)
        net_salary = round(snap.get('final_salary', 0.0), 2)
        total_deductions = round(total_earnings - net_salary, 2)
        salary_words = _amount_in_words(net_salary) if net_salary > 0 else 'Zero Only'
        return _render_payslip_response(
            request, emp, emp_type, selected_year, selected_month, month_name,
            days_in_month, absent_days_display, salary,
            basic_full, housing_full, transport_full, phone_full, other_allowance_full,
            incentives_commission, additions, leave_deduction, late_deduction,
            advance_ded, other_ded, carryover_ded, total_earnings, total_deductions,
            net_salary, salary_words,
        )

    salary = float(emp.salary) if emp.salary else 0.0

    # Use the employee's salary cycle — same logic as the payroll dashboard.
    cycle_day = getattr(emp, 'salary_cycle_start_day', None) or 21
    period_start, period_end, days_in_period, total_holidays = _get_employee_pay_period(
        cycle_day, selected_year, selected_month
    )
    days_in_month = days_in_period  # "Total No. of days" on the payslip

    # --- Payroll figures (identical path to the dashboard) ---
    if emp_type == 'inhouse' and emp.department == 'Admin':
        payroll = _get_inhouse_payroll_row(
            emp, selected_year, selected_month,
            period_start, period_end, total_holidays,
            days_in_period=days_in_period,
        )
        daily_rate = payroll['daily_rate']
        absent_days = payroll['absent_days']
        half_days = payroll['half_days']
        late_half_days = payroll['late_half_days']
        incentives = payroll['incentives']
        commission = payroll['commission']
        reductions = payroll['reductions']
        annual_leave_compensation = payroll['annual_leave_compensation']
        annual_leave_extra_deduction = payroll['annual_leave_extra_deduction']
        total_deduction_days = payroll['total_deduction_days']
        absent_days_display = round(total_deduction_days, 2)
        # Real Basic/Housing/Transport/Phone/Other breakdown from the
        # employee's approved SalaryStructure (Phase 5) — no approved
        # structure on file means we can't print an accurate payslip.
        if not payroll['has_salary_structure']:
            return _render_payslip_missing_structure(request, emp, selected_year, selected_month, month_name)
        basic_full = payroll['basic_salary']
        housing_full = payroll['housing_allowance']
        transport_full = payroll['transport_allowance']
        phone_full = payroll['phone_allowance']
        other_allowance_full = payroll['other_allowance_amt']
        # Attendance-based deduction components
        att_leave_ded = round(daily_rate * (absent_days + half_days * 0.5)
                              - annual_leave_compensation + annual_leave_extra_deduction, 2)
        att_late_ded = round(daily_rate * (late_half_days * 0.5), 2)
    else:
        banks = list(Bank.objects.filter(is_active=True).order_by('name'))
        payroll = _get_sales_payroll_row(
            emp, selected_year, selected_month, emp_type, banks,
            days_in_period, total_holidays,
            period_start=period_start, period_end=period_end,
        )
        incentives = payroll['incentives']
        commission = 0.0 if payroll.get('is_fixed_salary') else payroll['commission']
        reductions = payroll['reductions']
        absent_days_display = payroll.get('absent_days', 0)
        raw_salary = payroll.get('salary', 0.0)
        if emp_type == 'inhouse':
            # In-house Sales employee — SalaryStructure applies to them too
            # (it's not Admin-department-specific), so use the real breakdown
            # the same way as the Admin branch above.
            _salary_structure = get_effective_salary_structure(emp, period_end)
            if not _salary_structure:
                return _render_payslip_missing_structure(request, emp, selected_year, selected_month, month_name)
            basic_full = float(_salary_structure.basic)
            housing_full = float(_salary_structure.housing)
            transport_full = float(_salary_structure.transport)
            # Phase E9 — phone is not a salary component; fold any residual
            # amount into Other so the payslip's earnings lines still add up
            # to the gross. See _get_inhouse_payroll_row for the full note.
            phone_full = 0.0
            other_allowance_full = float(_salary_structure.other_allowance) + float(_salary_structure.phone)
        else:
            # Remote employees have no SalaryStructure model to draw from —
            # keep the existing flat-salary Basic 40% / Other Allowance 60%
            # split, unchanged from before this fix.
            if raw_salary:
                basic_full = round(raw_salary * 0.40, 2)
                other_allowance_full = round(raw_salary * 0.60, 2)
            else:
                basic_full = 0.0
                other_allowance_full = 0.0
            housing_full = 0.0
            transport_full = 0.0
            phone_full = 0.0
        al_comp = payroll.get('annual_leave_compensation', 0.0)
        att_leave_ded = max(0.0, payroll.get('deduction', 0.0) - al_comp)
        att_late_ded = 0.0

    # --- Active DeductionEntry records for this month ---
    target_idx = selected_year * 12 + (selected_month - 1)
    if emp_type == 'inhouse':
        deduction_entries = list(DeductionEntry.objects.filter(employee=emp))
    else:
        deduction_entries = list(DeductionEntry.objects.filter(remote_employee=emp))

    advance_ded = 0.0
    leave_ded_manual = 0.0
    late_ded_manual = 0.0
    other_ded_manual = 0.0
    additions = 0.0
    for entry in deduction_entries:
        start_idx = entry.start_year * 12 + (entry.start_month - 1)
        if start_idx <= target_idx < start_idx + entry.split_months:
            amt = float(entry.installment_amount)
            if entry.category == 'advance':
                advance_ded += amt
            elif entry.category in ('paid_leave', 'last_month_balance', 'other_addition'):
                additions += amt
            elif entry.category == 'leave_deduction':
                leave_ded_manual += amt
            elif entry.category == 'late_deduction':
                late_ded_manual += amt
            elif entry.category in ('visa_status_change', 'clawback', 'other_deduction'):
                other_ded_manual += amt

    advance_ded = round(advance_ded, 2)
    additions = round(additions, 2)
    # Combine attendance-based with manual entries; include PA reductions in other
    leave_deduction = round(att_leave_ded + leave_ded_manual, 2)
    late_deduction = round(att_late_ded + late_ded_manual, 2)
    other_ded = round(other_ded_manual + reductions, 2)

    incentives_commission = round(incentives + commission, 2)
    total_earnings = round(basic_full + housing_full + transport_full + phone_full + other_allowance_full + incentives_commission + additions, 2)

    # Carryover deduction (overflow from prior month that went negative)
    if emp_type == 'inhouse':
        _co = DeductionCarryover.objects.filter(employee=emp, to_year=selected_year, to_month=selected_month).first()
    else:
        _co = DeductionCarryover.objects.filter(remote_employee=emp, to_year=selected_year, to_month=selected_month).first()
    carryover_ded = round(float(_co.overflow_amount), 2) if _co else 0.0

    # Derive net from payroll['net_payroll'] — avoids rounding drift that accumulates
    # when daily_rate (already rounded to 2dp in the dict) is multiplied per-component.
    # Formula mirrors the dashboard's final_salary: net_payroll + bridge_adj - ded_entries - carryover + add_entries.
    # bridge_adj is added back for Admin employees because the payslip excludes bridge Sundays
    # (the dashboard compensates the same way in its final_salary calculation).
    # payroll['net_payroll'] already deducted reductions, so only DeductionEntry items and carryover are subtracted here.
    _ded_entries = advance_ded + leave_ded_manual + late_ded_manual + other_ded_manual
    if emp_type == 'inhouse' and emp.department == 'Admin':
        _bridge_adj = payroll.get('bridge_sunday_count', 0) * payroll['daily_rate']
        net_salary = round(payroll['net_payroll'] + _bridge_adj + additions - _ded_entries - carryover_ded, 2)
    else:
        net_salary = round(payroll['net_payroll'] + additions - _ded_entries - carryover_ded, 2)
    total_deductions = round(total_earnings - net_salary, 2)

    salary_words = _amount_in_words(net_salary) if net_salary > 0 else 'Zero Only'

    return _render_payslip_response(
        request, emp, emp_type, selected_year, selected_month, month_name,
        days_in_month, absent_days_display, salary,
        basic_full, housing_full, transport_full, phone_full, other_allowance_full,
        incentives_commission, additions, leave_deduction, late_deduction,
        advance_ded, other_ded, carryover_ded, total_earnings, total_deductions,
        net_salary, salary_words,
    )


def _render_payslip_missing_structure(request, emp, selected_year, selected_month, month_name):
    """Clear error page instead of a guessed payslip when the employee has no
    approved SalaryStructure covering this pay period (Phase 5 salary revision
    workflow). Registers nothing in GeneratedDocument since no payslip was
    actually produced."""
    logger.warning(
        "Payslip blocked for %s %s/%s — no approved SalaryStructure on file",
        emp.name, selected_month, selected_year,
    )
    return render(request, 'payroll/payslip_missing_structure.html', {
        'emp': emp,
        'month_name': month_name,
        'selected_year': selected_year,
        'selected_month': selected_month,
    }, status=422)


def _render_payslip_response(
    request, emp, emp_type, selected_year, selected_month, month_name,
    days_in_month, absent_days_display, salary,
    basic_full, housing_full, transport_full, phone_full, other_allowance_full,
    incentives_commission, additions, leave_deduction, late_deduction,
    advance_ded, other_ded, carryover_ded, total_earnings, total_deductions,
    net_salary, salary_words,
):
    """Register the GeneratedDocument entry and render the payslip HTML.
    Shared by the live-computed path and the locked-snapshot (Mark as Paid) path."""
    if emp_type == 'inhouse':
        _gdoc, _ = GeneratedDocument.objects.get_or_create(
            doc_type='payslip', employee=emp, year=selected_year, month=selected_month,
            defaults={'remote_employee': None},
        )
    else:
        _gdoc, _ = GeneratedDocument.objects.get_or_create(
            doc_type='payslip', remote_employee=emp, year=selected_year, month=selected_month,
            defaults={'employee': None},
        )

    logger.info("Payslip viewed for %s %s/%s by %s", emp.name, selected_month, selected_year, request.user.username)
    return render(request, 'payroll/payslip.html', {
        'emp': emp,
        'month_name': month_name,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'days_in_month': days_in_month,
        'absent_days_display': absent_days_display,
        'salary': salary,
        'basic_full': basic_full,
        'housing_full': housing_full,
        'transport_full': transport_full,
        'phone_full': phone_full,
        'allowance_full': other_allowance_full,
        'incentives_commission': incentives_commission,
        'additions': additions,
        'leave_deduction': leave_deduction,
        'late_deduction': late_deduction,
        'advance_ded': advance_ded,
        'other_ded': other_ded,
        'carryover_ded': carryover_ded,
        'total_earnings': total_earnings,
        'total_deductions': total_deductions,
        'net_salary': net_salary,
        'salary_words': salary_words,
        'doc_ref': _gdoc.ref,
    })


# ============================================
# Payslip History (searchable archive of Mark-as-Paid records)
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def payslip_history(request):
    """Searchable archive of every payslip ever locked via Mark as Paid, for
    any employee — in-house or remote, active or since left — across all months.
    Each row links to the existing payslip view, which already renders from the
    locked PaidSalaryRecord snapshot when one exists for that employee/month."""
    search = request.GET.get('q', '').strip()
    year_filter = request.GET.get('year', '').strip()
    month_filter = request.GET.get('month', '').strip()
    type_filter = request.GET.get('type', '').strip()

    records = PaidSalaryRecord.objects.select_related('employee', 'remote_employee').filter(
        snapshot__isnull=False
    )

    if search:
        records = records.filter(
            Q(employee__name__icontains=search) | Q(remote_employee__name__icontains=search)
        )
    if year_filter.isdigit():
        records = records.filter(year=int(year_filter))
    if month_filter.isdigit():
        records = records.filter(month=int(month_filter))
    if type_filter == 'inhouse':
        records = records.filter(employee__isnull=False)
    elif type_filter == 'remote':
        records = records.filter(remote_employee__isnull=False)

    records = records.order_by('-year', '-month', 'employee__name', 'remote_employee__name')
    total_count = records.count()

    paginator = Paginator(records, 50)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    rows = []
    for rec in page_obj:
        emp = rec.employee or rec.remote_employee
        emp_type = 'inhouse' if rec.employee_id else 'remote'
        snap = rec.snapshot or {}
        rows.append({
            'record': rec,
            'employee': emp,
            'employee_type': emp_type,
            'department': snap.get('department') or getattr(emp, 'department', '') or '',
            'designation': snap.get('designation', ''),
            'month_name': MONTH_NAMES[rec.month] if 1 <= rec.month <= 12 else rec.month,
            'final_salary': rec.final_salary,
            'currency': rec.currency,
            'paid_at': rec.paid_at,
            'paid_by': rec.paid_by,
        })

    available_years = list(
        PaidSalaryRecord.objects.order_by().values_list('year', flat=True).distinct().order_by('-year')
    )

    return render(request, 'payroll/payslip_history.html', {
        'rows': rows,
        'page_obj': page_obj,
        'search': search,
        'year_filter': year_filter,
        'month_filter': month_filter,
        'type_filter': type_filter,
        'available_years': available_years,
        'months': MONTH_CHOICES,
        'total_count': total_count,
    })


# ============================================
# Advance Payment Voucher (print view)
# ============================================

@login_required
@user_passes_test(section_required('payroll'))
def advance_voucher_download(request):
    """Render a printable payment voucher page for an advance deduction."""
    entry_id = request.GET.get('entry_id')
    try:
        year = int(request.GET.get('year', 0))
        month = int(request.GET.get('month', 0))
    except (TypeError, ValueError):
        from django.http import Http404
        raise Http404("Invalid parameters.")

    entry = get_object_or_404(DeductionEntry, pk=entry_id, category='advance')
    emp = entry.employee or entry.remote_employee
    emp_type = 'inhouse' if entry.employee_id else 'remote'

    _month_full = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December',
    }

    e_start = entry.start_year * 12 + (entry.start_month - 1)
    target_idx = year * 12 + (month - 1)

    # Register document and get stable reference ID
    _gdoc, _ = GeneratedDocument.objects.get_or_create(
        doc_type='advance_voucher',
        deduction_entry=entry,
        year=year,
        month=month,
        defaults={
            'employee': entry.employee,
            'remote_employee': entry.remote_employee,
        },
    )

    advance_entries = [{
        'voucher_no': _gdoc.ref,
        'date': entry.created_at.date(),
        'towards': entry.note if entry.note else 'Cash Advance',
        'total_amount': float(entry.total_amount),
        'installment_amount': float(entry.installment_amount),
        'split_months': entry.split_months,
        'installment_num': target_idx - e_start + 1,
    }]

    return render(request, 'payroll/advance_voucher.html', {
        'emp': emp,
        'emp_type': emp_type,
        'advance_entries': advance_entries,
        'month': month,
        'year': year,
        'month_name': _month_full.get(month, str(month)),
    })


# ============================================
# Exchange Rate API
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
@require_http_methods(["POST"])
def save_exchange_rate(request):
    """Save or update the exchange rate for a foreign currency (INR, NPR, ...) for a given month."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    currency = (data.get('currency') or 'INR').strip().upper()
    year = data.get('year')
    month = data.get('month')
    rate = data.get('rate')

    if currency not in FOREIGN_CURRENCIES:
        return JsonResponse({'success': False, 'error': 'Currency must be one of: ' + ', '.join(FOREIGN_CURRENCIES)}, status=400)
    if not all([year, month, rate]):
        return JsonResponse({'success': False, 'error': 'year, month, and rate are required'}, status=400)

    try:
        rate_val = Decimal(str(rate))
        if rate_val <= 0:
            raise ValueError
    except (ValueError, Exception):
        return JsonResponse({'success': False, 'error': 'Rate must be a positive number'}, status=400)

    obj, created = ExchangeRate.objects.update_or_create(
        currency=currency, year=int(year), month=int(month),
        defaults={'rate': rate_val},
    )
    return JsonResponse({'success': True, 'currency': currency, 'rate': float(obj.rate), 'created': created})


# ============================================
# Test / Simplified Payroll Dashboard
# ============================================

@login_required
@user_passes_test(section_required('payroll'), login_url='/report/')
def payroll_test_dashboard(request):
    """Payroll dashboard with 4 tables by employee category."""
    selected_month, selected_year = get_selected_month_year(request)

    # Legacy frozen-month protection: if this month was previously frozen with the
    # old payroll system, skip carryover mutations so the historical DB state is
    # not overwritten.  PaidSalaryRecord overlay (populated by convert_frozen_to_paid)
    # will display the locked values instead.
    frozen_obj = FrozenPayrollMonth.objects.filter(
        year=selected_year, month=selected_month
    ).first()
    is_frozen = frozen_obj is not None

    # Default pay period (21st prev → 20th current) used for header display
    if selected_month == 1:
        prev_month, prev_year = 12, selected_year - 1
    else:
        prev_month, prev_year = selected_month - 1, selected_year
    default_period_start = datetime.date(prev_year, prev_month, 21)
    default_period_end = datetime.date(selected_year, selected_month, 20)
    default_days = (default_period_end - default_period_start).days + 1
    default_holidays = _count_holidays_in_period(default_period_start, default_period_end)

    # Per-employee pay period cache (keyed by cycle_start_day)
    _period_cache = {}

    def _emp_period(emp):
        day = emp.salary_cycle_start_day or 21
        if day not in _period_cache:
            _period_cache[day] = _get_employee_pay_period(day, selected_year, selected_month)
        return _period_cache[day]

    banks = list(Bank.objects.filter(is_active=True).order_by('name'))
    banks_json = json.dumps([
        {'id': b.id, 'name': b.name, 'rate': float(b.per_account_charge),
         'inr_rate': float(b.inr_per_account_charge) if b.inr_per_account_charge else None}
        for b in banks
    ])

    all_inhouse_tcr_ids = set(
        Employee.objects.filter(is_active=True)
        .exclude(tcr_id__isnull=True).exclude(tcr_id='')
        .values_list('tcr_id', flat=True)
    )

    # Table 1: Admin In-House
    admin_inhouse_emps = list(Employee.objects.filter(department='Admin', is_active=True).order_by('name'))
    admin_inhouse_rows = []
    for emp in admin_inhouse_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        admin_inhouse_rows.append(
            _get_inhouse_payroll_row(emp, selected_year, selected_month, p_start, p_end, p_hols, days_in_period=p_days)
        )

    # Table 2: Admin Remote
    admin_remote_qs = RemoteEmployee.objects.filter(department='Admin', is_active=True)
    if all_inhouse_tcr_ids:
        admin_remote_qs = admin_remote_qs.exclude(tcr_id__in=all_inhouse_tcr_ids)
    admin_remote_emps = list(admin_remote_qs.order_by('name'))
    admin_remote_rows = []
    for emp in admin_remote_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        _row = _get_sales_payroll_row(emp, selected_year, selected_month, 'remote', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
        _attach_gross_breakdown(_row, p_end)
        admin_remote_rows.append(_row)

    # Table 3: Sales - Fixed Salary (is_fixed_salary=True)
    sales_fixed_inhouse_emps = list(Employee.objects.filter(
        department='Sales', is_active=True, is_fixed_salary=True
    ).order_by('name'))
    sales_fixed_remote_qs = RemoteEmployee.objects.filter(
        is_active=True, is_fixed_salary=True
    ).exclude(department='Admin')
    if all_inhouse_tcr_ids:
        sales_fixed_remote_qs = sales_fixed_remote_qs.exclude(tcr_id__in=all_inhouse_tcr_ids)
    sales_fixed_remote_emps = list(sales_fixed_remote_qs.order_by('name'))
    sales_fixed_rows = []
    for emp in sales_fixed_inhouse_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        _row = _get_sales_payroll_row(emp, selected_year, selected_month, 'inhouse', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
        _attach_gross_breakdown(_row, p_end)
        sales_fixed_rows.append(_row)
    for emp in sales_fixed_remote_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        _row = _get_sales_payroll_row(emp, selected_year, selected_month, 'remote', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
        _attach_gross_breakdown(_row, p_end)
        sales_fixed_rows.append(_row)

    # Table 4: Sales - Performance Based (is_fixed_salary=False)
    sales_perf_inhouse_emps = list(Employee.objects.filter(
        department='Sales', is_active=True, is_fixed_salary=False
    ).order_by('name'))
    sales_perf_remote_qs = RemoteEmployee.objects.filter(
        is_active=True, is_fixed_salary=False
    ).exclude(department='Admin')
    if all_inhouse_tcr_ids:
        sales_perf_remote_qs = sales_perf_remote_qs.exclude(tcr_id__in=all_inhouse_tcr_ids)
    sales_perf_remote_emps = list(sales_perf_remote_qs.order_by('name'))
    sales_perf_rows = []
    for emp in sales_perf_inhouse_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        _row = _get_sales_payroll_row(emp, selected_year, selected_month, 'inhouse', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
        _attach_gross_breakdown(_row, p_end)
        sales_perf_rows.append(_row)
    for emp in sales_perf_remote_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        _row = _get_sales_payroll_row(emp, selected_year, selected_month, 'remote', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
        _attach_gross_breakdown(_row, p_end)
        sales_perf_rows.append(_row)

    # Table 4 (reference): Sales - Performance TEST — talktime-proportional daily
    # pay ("Method 2"), remote employees only. sales_perf_rows above already uses
    # this same formula (via _get_sales_payroll_row) for months >= July 2026, so
    # for those months this tab is redundant with the live one; it's kept so
    # pre-cutoff months can still be compared against what Method 2 would have paid.
    sales_perf_test_rows = []
    for emp in sales_perf_remote_emps:
        p_start, p_end, p_days, p_hols = _emp_period(emp)
        test_row = _get_sales_performance_test_row(emp, p_start, p_end, p_days, p_hols, year=selected_year, month=selected_month)
        _attach_gross_breakdown(test_row, p_end)
        live_row = _get_sales_payroll_row(
            emp, selected_year, selected_month, 'remote', banks, p_days, p_hols,
            period_start=p_start, period_end=p_end,
        )
        test_row['live_net_payroll'] = live_row['net_payroll']
        test_row['live_commission'] = live_row.get('commission', 0)
        test_row['diff'] = round(test_row['net_payroll_test'] - live_row['net_payroll'], 2)
        sales_perf_test_rows.append(test_row)
    # Totals/diff are finalized below (after the paid-snapshot overlay), since
    # already-paid employees need their "Live Net Payroll" replaced with the
    # locked snapshot value to stay consistent with the real Sales: Performance tab.

    # ---- Deductions & Additions ----
    _DED_COLS = ['advance', 'visa_status_change', 'clawback', 'leave_deduction', 'late_deduction', 'other_deduction']
    _ADD_COLS = ['last_month_balance', 'paid_leave', 'other_addition']
    _ALL_CATS = _DED_COLS + _ADD_COLS
    target_idx = selected_year * 12 + (selected_month - 1)
    _mnames = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

    active_by_emp = _defaultdict(list)
    all_deductions_list = []
    _recovery_map = _build_deduction_recovery_map()
    _today = timezone.localdate()
    _today_idx = _today.year * 12 + (_today.month - 1)
    for entry in DeductionEntry.objects.select_related('employee', 'remote_employee').order_by('-created_at'):
        emp_obj = entry.employee or entry.remote_employee
        emp_type_str = 'inhouse' if entry.employee else 'remote'
        start_idx = entry.start_year * 12 + (entry.start_month - 1)
        is_active_this_month = start_idx <= target_idx < start_idx + entry.split_months
        if is_active_this_month:
            active_by_emp[(emp_type_str, emp_obj.id)].append(entry)
        end_y, end_m = entry.end_month_year()
        end_idx = end_y * 12 + (end_m - 1)
        # "Recovered" only means something once the entry's own installment
        # schedule has actually finished (relative to *today*, not the
        # month being viewed) — an entry still mid-schedule or not yet
        # started is neither recovered nor "pending", it just hasn't run yet.
        is_upcoming = start_idx > _today_idx
        is_finished = end_idx < _today_idx
        all_deductions_list.append({
            'id': entry.id,
            'employee': emp_obj,
            'employee_type': emp_type_str,
            'currency': entry.currency,
            'category_display': entry.get_category_display(),
            'entry_type': entry.entry_type,
            'total_amount': float(entry.total_amount),
            'split_months': entry.split_months,
            'installment_amount': float(entry.installment_amount),
            'start_month_name': _mnames[entry.start_month],
            'start_year': entry.start_year,
            'end_month_name': _mnames[end_m],
            'end_year': end_y,
            'note': entry.note,
            'created_at': entry.created_at.strftime('%d %b %Y'),
            'is_recovered': is_finished and _entry_is_recovered(end_idx, (emp_type_str, emp_obj.id), _recovery_map, _today_idx),
            'is_active_this_month': is_active_this_month,
            'is_upcoming': is_upcoming,
            'is_finished': is_finished,
        })

    all_payroll_inhouse = list(Employee.objects.filter(
        is_active=True, department__in=['Admin', 'Sales']
    ).order_by('department', 'name'))
    all_payroll_remote_qs = RemoteEmployee.objects.filter(is_active=True)
    if all_inhouse_tcr_ids:
        all_payroll_remote_qs = all_payroll_remote_qs.exclude(tcr_id__in=all_inhouse_tcr_ids)
    all_payroll_remote = list(all_payroll_remote_qs.order_by('name'))

    incoming_carryovers = DeductionCarryover.objects.filter(
        to_year=selected_year, to_month=selected_month
    ).exclude(is_skipped=True)
    carryover_by_emp = {}
    for co in incoming_carryovers:
        key = ('inhouse', co.employee_id) if co.employee_id else ('remote', co.remote_employee_id)
        carryover_by_emp[key] = float(co.overflow_amount)

    # ---- Employee-wise deduction/addition history (Deductions tab) ----
    # Groups the same entries from all_deductions_list (already ordered newest-first
    # by created_at) under each employee, so the tab can show one row per employee
    # with a complete, drill-down history instead of one giant flat table.
    _ded_group_map = {}
    deduction_groups = []
    for entry_dict in all_deductions_list:
        key = (entry_dict['employee_type'], entry_dict['employee'].id)
        group = _ded_group_map.get(key)
        if group is None:
            group = {
                'employee': entry_dict['employee'],
                'employee_type': entry_dict['employee_type'],
                'currency': entry_dict['currency'],
                'entries': [],
                'lifetime_deduction': 0.0,
                'lifetime_addition': 0.0,
                'active_deduction': 0.0,
                'active_addition': 0.0,
                'active_count': 0,
                'pending_count': 0,
            }
            _ded_group_map[key] = group
            deduction_groups.append(group)
        group['entries'].append(entry_dict)
        if entry_dict['entry_type'] == 'deduction':
            group['lifetime_deduction'] = round(group['lifetime_deduction'] + entry_dict['total_amount'], 2)
        else:
            group['lifetime_addition'] = round(group['lifetime_addition'] + entry_dict['total_amount'], 2)
        if entry_dict['is_active_this_month']:
            group['active_count'] += 1
            if entry_dict['entry_type'] == 'deduction':
                group['active_deduction'] = round(group['active_deduction'] + entry_dict['installment_amount'], 2)
            else:
                group['active_addition'] = round(group['active_addition'] + entry_dict['installment_amount'], 2)
        if entry_dict['entry_type'] == 'deduction' and entry_dict['is_finished'] and not entry_dict['is_recovered']:
            group['pending_count'] += 1

    for group in deduction_groups:
        group['lifetime_net'] = round(group['lifetime_addition'] - group['lifetime_deduction'], 2)
        group['active_net'] = round(group['active_addition'] - group['active_deduction'], 2)
        group['entry_count'] = len(group['entries'])
        group['carryover_in'] = carryover_by_emp.get(
            (group['employee_type'], group['employee'].id), 0.0
        )
    deduction_groups.sort(key=lambda g: g['employee'].name.lower())

    # Build lookup from pre-computed payroll rows so leave/late deduction display
    # uses the employee's actual salary-cycle period instead of the calendar month summary.
    _payroll_row_lookup = {}
    for _pr in admin_inhouse_rows + sales_fixed_rows + sales_perf_rows:
        _payroll_row_lookup[('inhouse', _pr['employee'].id)] = _pr

    deduction_rows = []
    for emp in all_payroll_inhouse:
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get(('inhouse', emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        # Auto-fill leave/late for attendance-based employees using pre-computed payroll row
        if emp.salary and emp.payroll_type != 'performance':
            emp_p_start, emp_p_end, emp_p_days, _ = _emp_period(emp)
            daily = float(emp.salary) / emp_p_days
            pr = _payroll_row_lookup.get(('inhouse', emp.id))
            if pr:
                absent_d = pr.get('absent_days', 0) + pr.get('half_days', 0) * 0.5 + pr.get('bridge_sunday_count', 0)
                late_d = pr.get('late_days', 0)
            else:
                absent_d = 0
                late_d = 0
            cat['leave_deduction'] = round(daily * absent_d, 2)
            cat['late_deduction'] = round(daily * (late_d // 3) * 0.5, 2)
        carryover_in = carryover_by_emp.get(('inhouse', emp.id), 0.0)
        _has_salary = emp.salary and emp.payroll_type != 'performance'
        ded_for_total = [c for c in _DED_COLS if c not in ('leave_deduction', 'late_deduction')] if (emp.department == 'Admin' or _has_salary) else _DED_COLS
        total_ded = round(sum(cat[c] for c in ded_for_total) + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        deduction_rows.append({
            'employee': emp, 'employee_type': 'inhouse', 'currency': emp.currency,
            'categories': cat, 'carryover_in': carryover_in, 'ded_for_total': ded_for_total,
            'total_deductions': total_ded, 'total_additions': total_add,
            'net': round(total_add - total_ded, 2),
        })

    for emp in all_payroll_remote:
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get(('remote', emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        carryover_in = carryover_by_emp.get(('remote', emp.id), 0.0)
        total_ded = round(sum(cat[c] for c in _DED_COLS) + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        deduction_rows.append({
            'employee': emp, 'employee_type': 'remote', 'currency': emp.currency,
            'categories': cat, 'carryover_in': carryover_in, 'ded_for_total': _DED_COLS,
            'total_deductions': total_ded, 'total_additions': total_add,
            'net': round(total_add - total_ded, 2),
        })

    for row in sales_perf_rows:
        row['combined_deductions'] = round(row.get('deduction', 0) + row.get('reductions', 0), 2)

    # ---- Final Summary ----
    all_payroll_rows = admin_inhouse_rows + admin_remote_rows + sales_fixed_rows + sales_perf_rows
    payroll_by_emp = {}
    for row in all_payroll_rows:
        payroll_by_emp[(row['employee_type'], row['employee'].id)] = row
    ded_by_emp = {(r['employee_type'], r['employee'].id): r for r in deduction_rows}

    if selected_month == 12:
        _co_to_month, _co_to_year = 1, selected_year + 1
    else:
        _co_to_month, _co_to_year = selected_month + 1, selected_year

    # Employees already marked paid for this month have a locked PaidSalaryRecord
    # snapshot (including its own carryover_out at the time of payment). Their
    # carryover rows must not be recomputed/deleted on later page loads using
    # whatever the employee's *current* settings (salary, currency, cycle) are —
    # that silently destroys real debt when those settings change after payment.
    _paid_keys_this_month = set(
        PaidSalaryRecord.objects.filter(year=selected_year, month=selected_month).values_list(
            'employee_id', 'remote_employee_id'
        )
    )
    _paid_emp_keys = {('inhouse', e) for e, r in _paid_keys_this_month if e} | {
        ('remote', r) for e, r in _paid_keys_this_month if r
    }

    final_rows = []
    for emp in all_payroll_inhouse:
        key = ('inhouse', emp.id)
        p = payroll_by_emp.get(key)
        d = ded_by_emp.get(key)
        payroll_net = p['net_payroll'] if p else 0.0
        total_ded = d['total_deductions'] if d else 0.0
        total_add = d['total_additions'] if d else 0.0
        carryover_in = d['carryover_in'] if d else 0.0
        final_salary = round(payroll_net - total_ded + total_add, 2)
        if not is_frozen and key not in _paid_emp_keys:
            if final_salary < 0:
                DeductionCarryover.objects.update_or_create(
                    employee=emp, from_year=selected_year, from_month=selected_month,
                    defaults={'overflow_amount': Decimal(str(abs(final_salary))),
                              'to_year': _co_to_year, 'to_month': _co_to_month, 'remote_employee': None,
                              'currency': emp.currency},
                )
            else:
                DeductionCarryover.objects.filter(
                    employee=emp, from_year=selected_year, from_month=selected_month
                ).delete()
            if carryover_in > 0:
                inc_co = incoming_carryovers.filter(employee=emp).first()
                if inc_co:
                    inc_co.applied_amount = Decimal(str(min(carryover_in, float(inc_co.overflow_amount))))
                    inc_co.save(update_fields=['applied_amount'])
        if final_salary < 0:
            final_salary = 0.0
        carryover_out = 0.0
        if final_salary == 0.0 and round(payroll_net - total_ded + total_add, 2) < 0:
            carryover_out = round(abs(payroll_net - total_ded + total_add), 2)
        # Itemized deduction breakdown for the click-through modal on the
        # Deductions figure below — only categories actually counted toward
        # total_ded are included, plus carryover_in, so the modal always
        # reconciles exactly to the total shown in the table.
        _ded_for_total = d.get('ded_for_total', _DED_COLS) if d else _DED_COLS
        _categories = d['categories'] if d else {}
        ded_breakdown_json = json.dumps({
            'items': {c: _categories.get(c, 0.0) for c in _ded_for_total if _categories.get(c, 0.0)},
            'carryover_in': carryover_in,
        })
        # ---- Phase E6: itemized deduction columns --------------------------
        # Only categories inside _ded_for_total are shown as figures. Late and
        # Leave fall OUT of that set for Admin/fixed-salary in-house staff,
        # because their attendance is already priced into net_payroll — showing
        # a number there would read as a second deduction for the same absence
        # and the columns would no longer sum to the Deductions total.
        # `_shown` is therefore the single source of truth for both the value
        # and whether the cell is a figure or a greyed dash.
        def _ded_col(cat):
            return _categories.get(cat, 0.0) if cat in _ded_for_total else None
        _other_total = round(sum(
            _categories.get(c, 0.0) for c in OTHER_DEDUCTION_CATEGORIES if c in _ded_for_total
        ), 2)
        ded_cols = {
            'late': _ded_col('late_deduction'),
            'leave': _ded_col('leave_deduction'),
            'advance': _ded_col('advance'),
            'other': _other_total,
            'carryover': carryover_in,
        }
        final_rows.append({
            'employee': emp, 'employee_type': 'inhouse',
            'department': emp.department or 'In-House', 'currency': emp.currency,
            'payroll_net': payroll_net, 'total_deductions': total_ded,
            'total_additions': total_add, 'final_salary': final_salary,
            'carryover_in': carryover_in, 'carryover_out': carryover_out,
            'ded_breakdown_json': ded_breakdown_json,
            'ded_cols': ded_cols,
        })

    for emp in all_payroll_remote:
        key = ('remote', emp.id)
        p = payroll_by_emp.get(key)
        d = ded_by_emp.get(key)
        payroll_net = p['net_payroll'] if p else 0.0
        total_ded = d['total_deductions'] if d else 0.0
        total_add = d['total_additions'] if d else 0.0
        carryover_in = d['carryover_in'] if d else 0.0
        final_salary = round(payroll_net - total_ded + total_add, 2)
        if not is_frozen and key not in _paid_emp_keys:
            if final_salary < 0:
                DeductionCarryover.objects.update_or_create(
                    remote_employee=emp, from_year=selected_year, from_month=selected_month,
                    defaults={'overflow_amount': Decimal(str(abs(final_salary))),
                              'to_year': _co_to_year, 'to_month': _co_to_month, 'employee': None,
                              'currency': emp.currency},
                )
            else:
                DeductionCarryover.objects.filter(
                    remote_employee=emp, from_year=selected_year, from_month=selected_month
                ).delete()
            if carryover_in > 0:
                inc_co = incoming_carryovers.filter(remote_employee=emp).first()
                if inc_co:
                    inc_co.applied_amount = Decimal(str(min(carryover_in, float(inc_co.overflow_amount))))
                    inc_co.save(update_fields=['applied_amount'])
        if final_salary < 0:
            final_salary = 0.0
        carryover_out = 0.0
        if final_salary == 0.0 and round(payroll_net - total_ded + total_add, 2) < 0:
            carryover_out = round(abs(payroll_net - total_ded + total_add), 2)
        _ded_for_total = d.get('ded_for_total', _DED_COLS) if d else _DED_COLS
        _categories = d['categories'] if d else {}
        ded_breakdown_json = json.dumps({
            'items': {c: _categories.get(c, 0.0) for c in _ded_for_total if _categories.get(c, 0.0)},
            'carryover_in': carryover_in,
        })
        # ---- Phase E6: itemized deduction columns --------------------------
        # Only categories inside _ded_for_total are shown as figures. Late and
        # Leave fall OUT of that set for Admin/fixed-salary in-house staff,
        # because their attendance is already priced into net_payroll — showing
        # a number there would read as a second deduction for the same absence
        # and the columns would no longer sum to the Deductions total.
        # `_shown` is therefore the single source of truth for both the value
        # and whether the cell is a figure or a greyed dash.
        def _ded_col(cat):
            return _categories.get(cat, 0.0) if cat in _ded_for_total else None
        _other_total = round(sum(
            _categories.get(c, 0.0) for c in OTHER_DEDUCTION_CATEGORIES if c in _ded_for_total
        ), 2)
        ded_cols = {
            'late': _ded_col('late_deduction'),
            'leave': _ded_col('leave_deduction'),
            'advance': _ded_col('advance'),
            'other': _other_total,
            'carryover': carryover_in,
        }
        final_rows.append({
            'employee': emp, 'employee_type': 'remote',
            'department': getattr(emp, 'department', 'Remote') or 'Remote', 'currency': emp.currency,
            'payroll_net': payroll_net, 'total_deductions': total_ded,
            'total_additions': total_add, 'final_salary': final_salary,
            'carryover_in': carryover_in, 'carryover_out': carryover_out,
            'ded_breakdown_json': ded_breakdown_json,
            'ded_cols': ded_cols,
        })

    # Load full paid salary snapshots and overlay onto all rows
    paid_records = {}
    for rec in PaidSalaryRecord.objects.filter(year=selected_year, month=selected_month):
        if rec.employee_id:
            paid_records[('inhouse', rec.employee_id)] = rec
        elif rec.remote_employee_id:
            paid_records[('remote', rec.remote_employee_id)] = rec

    def _overlay_section_row(row):
        key = (row['employee_type'], row['employee'].id)
        rec = paid_records.get(key)
        if not rec or not rec.snapshot:
            row['is_paid'] = False
            row['is_partial'] = False
            row['pay_status'] = 'unpaid'
            return
        snap = rec.snapshot
        row['is_paid'] = True
        row['is_partial'] = rec.is_partial
        row['pay_status'] = 'partial' if rec.is_partial else 'paid'
        row['amount_paid'] = float(rec.effective_amount_paid)
        row['balance_due'] = float(rec.balance_due)
        row['paid_at'] = rec.paid_at
        row['paid_by'] = rec.paid_by
        row['paid_snapshot'] = snap
        row['paid_snapshot_json'] = json.dumps(snap)
        # Replace every computed value with the locked snapshot value
        row['salary'] = snap.get('salary', row.get('salary'))
        row['net_payroll'] = snap.get('net_payroll', row.get('net_payroll', 0))
        row['deduction'] = snap.get('deduction', row.get('deduction', 0))
        row['commission'] = snap.get('commission', row.get('commission', 0))
        row['incentives'] = snap.get('incentives', row.get('incentives', 0))
        row['reductions'] = snap.get('reductions', row.get('reductions', 0))
        if snap.get('full_days') is not None:
            row['full_days'] = snap['full_days']
        if snap.get('half_days') is not None:
            row['half_days'] = snap['half_days']
        if snap.get('absent_days') is not None:
            row['absent_days'] = snap['absent_days']
        if snap.get('late_days') is not None:
            row['late_days'] = snap['late_days']
        row['late_half_days'] = snap.get('late_half_days', row.get('late_half_days', 0))
        if snap.get('present_days') is not None:
            row['present_days'] = snap['present_days']
        if snap.get('bank_submissions'):
            row['bank_counts_list'] = [bs['count'] for bs in snap['bank_submissions']]
        row['combined_deductions'] = round(snap.get('deduction', 0) + snap.get('reductions', 0), 2)

    for row in admin_inhouse_rows + admin_remote_rows + sales_fixed_rows + sales_perf_rows:
        _overlay_section_row(row)

    for row in final_rows:
        key = (row['employee_type'], row['employee'].id)
        rec = paid_records.get(key)
        if not rec or not rec.snapshot:
            row['is_paid'] = False
            row['is_partial'] = False
            row['pay_status'] = 'unpaid'
            row['amount_paid'] = 0.0
            row['balance_due'] = row.get('final_salary', 0.0)
            row['payment_method'] = ''
            row['payment_method_label'] = ''
            row['payment_splits'] = None
            row['payment_splits_json'] = 'null'
            row['payment_date'] = None
            continue
        snap = rec.snapshot
        row['is_paid'] = True
        row['paid_at'] = rec.paid_at
        row['paid_by'] = rec.paid_by
        # ---- Phase E6: payment execution -------------------------------
        # is_paid stays True for a partial payment — the row IS locked and
        # must keep behaving as locked everywhere else in the page. The new
        # pay_status carries the three-way distinction for display only.
        row['is_partial'] = rec.is_partial
        row['pay_status'] = 'partial' if rec.is_partial else 'paid'
        row['amount_paid'] = float(rec.effective_amount_paid)
        row['balance_due'] = float(rec.balance_due)
        row['payment_method'] = rec.payment_method
        row['payment_splits'] = rec.payment_splits
        row['payment_splits_json'] = json.dumps(rec.payment_splits) if rec.payment_splits else 'null'
        if rec.payment_method == 'mixed' and rec.payment_splits:
            row['payment_method_label'] = ' + '.join(
                _payment_method_label(s['method'], row['employee']) for s in rec.payment_splits
            )
        else:
            row['payment_method_label'] = _payment_method_label(rec.payment_method, row['employee'])
        row['payment_date'] = rec.payment_date or (rec.paid_at.date() if rec.paid_at else None)
        row['paid_snapshot'] = snap
        row['paid_snapshot_json'] = json.dumps(snap)
        row['payroll_net'] = snap.get('net_payroll', row.get('payroll_net', 0))
        row['total_deductions'] = snap.get('total_deductions', row.get('total_deductions', 0))
        row['total_additions'] = snap.get('total_additions', row.get('total_additions', 0))
        row['carryover_in'] = snap.get('carryover_in', row.get('carryover_in', 0))
        row['carryover_out'] = snap.get('carryover_out', row.get('carryover_out', 0))
        row['final_salary'] = snap.get('final_salary', row.get('final_salary', 0))
        # Itemized deduction columns must read from the locked snapshot once a
        # month is paid, same as every other figure on this row. The None-mask
        # (which categories are excluded from the total) is a property of the
        # employee's department/salary type rather than of the payment, so the
        # live mask still applies — only the amounts come from the lock.
        _snap_cat = snap.get('deductions_breakdown')
        _live_cols = row.get('ded_cols')
        if _snap_cat and _live_cols:
            row['ded_cols'] = {
                'late': None if _live_cols['late'] is None else _snap_cat.get('late_deduction', 0.0),
                'leave': None if _live_cols['leave'] is None else _snap_cat.get('leave_deduction', 0.0),
                'advance': None if _live_cols['advance'] is None else _snap_cat.get('advance', 0.0),
                'other': round(sum(_snap_cat.get(c, 0.0) or 0.0 for c in OTHER_DEDUCTION_CATEGORIES), 2),
                'carryover': snap.get('carryover_in', 0) or 0,
            }

    # ---- Phase E6b: category label + Gross Pay breakdown on the summary table ----
    # Both are copied from the per-category rows rather than recomputed, so the
    # summary table and the five category tabs can never show different figures
    # for the same employee. An employee absent from every category list (should
    # not happen, but the summary is built from a wider query) simply falls back
    # to a plain label and no breakdown, rather than a fabricated one.
    _cat_label_by_emp = {}
    _gross_by_emp = {}
    # sales_perf_rows is split into its in-house/remote lists further down, so
    # the label is derived from employee_type here rather than referencing those
    # lists before they exist.
    for _rows, _label in (
        (admin_inhouse_rows, 'Admin: In-House'),
        (admin_remote_rows, 'Admin: Remote'),
        (sales_fixed_rows, 'Sales: Fixed'),
        (sales_perf_rows, None),
    ):
        for _r in _rows:
            _k = (_r['employee_type'], _r['employee'].id)
            if _label is None:
                _cat_label_by_emp[_k] = (
                    'Sales Perf: In-House' if _r['employee_type'] == 'inhouse'
                    else 'Sales Perf: Remote'
                )
            else:
                _cat_label_by_emp[_k] = _label
            if _r.get('has_salary_structure') is not None:
                _gross_by_emp[_k] = {
                    'basic': _r.get('basic_salary', 0.0),
                    'housing': _r.get('housing_allowance', 0.0),
                    'transport': _r.get('transport_allowance', 0.0),
                    'phone': _r.get('phone_allowance', 0.0),
                    'other': _r.get('other_allowance_amt', 0.0),
                    'total': _r.get('salary', 0.0),
                    'has_structure': _r.get('has_salary_structure', False),
                    'is_estimated': _r.get('is_estimated', False),
                }

    for row in final_rows:
        _k = (row['employee_type'], row['employee'].id)
        row['category_label'] = _cat_label_by_emp.get(_k) or row.get('department') or '—'
        row['category_slug'] = _cat_label_by_emp.get(_k, '').lower().replace(': ', '-').replace(' ', '-') or 'other'
        row['gross'] = _gross_by_emp.get(_k)
        # Two-letter initials for the avatar chip, from the employee's own name.
        _parts = [p for p in (row['employee'].name or '').split() if p]
        row['initials'] = ((_parts[0][0] if _parts else '?') + (_parts[1][0] if len(_parts) > 1 else '')).upper()

    final_total_aed = round(sum(r['final_salary'] for r in final_rows if r['currency'] == 'AED'), 2)
    final_total_inr = round(sum(r['final_salary'] for r in final_rows if r['currency'] == 'INR'), 2)
    final_total_npr = round(sum(r['final_salary'] for r in final_rows if r['currency'] == 'NPR'), 2)

    # Sales Performance Method 2 tab: "Live Net Payroll" shows the Payroll
    # Summary value (final_salary — net payroll plus/minus all deductions and
    # additions, and the locked paid figure if already marked as paid) for
    # each employee, so it's directly comparable to what admins actually see
    # in the Final Summary tab, not just the pre-deductions commission figure.
    final_salary_by_emp = {(row['employee_type'], row['employee'].id): row['final_salary'] for row in final_rows}
    for test_row in sales_perf_test_rows:
        key = (test_row['employee_type'], test_row['employee'].id)
        if key in final_salary_by_emp:
            test_row['live_net_payroll'] = final_salary_by_emp[key]
        test_row['diff'] = round(test_row['net_payroll_test'] - test_row['live_net_payroll'], 2)
    sales_perf_test_totals = {
        'test_net_aed': round(sum(r['net_payroll_test'] for r in sales_perf_test_rows if r.get('currency', 'AED') == 'AED'), 2),
        'test_net_inr': round(sum(r['net_payroll_test'] for r in sales_perf_test_rows if r.get('currency') == 'INR'), 2),
        'test_net_npr': round(sum(r['net_payroll_test'] for r in sales_perf_test_rows if r.get('currency') == 'NPR'), 2),
        'live_net_aed': round(sum(r['live_net_payroll'] for r in sales_perf_test_rows if r.get('currency', 'AED') == 'AED'), 2),
        'live_net_inr': round(sum(r['live_net_payroll'] for r in sales_perf_test_rows if r.get('currency') == 'INR'), 2),
        'live_net_npr': round(sum(r['live_net_payroll'] for r in sales_perf_test_rows if r.get('currency') == 'NPR'), 2),
    }

    # Visa provider breakdown
    _visa_order = ['Jumbo', 'OnTime', 'Taamul', 'own']
    _visa_labels = {'Jumbo': 'Jumbo', 'OnTime': 'OnTime', 'Taamul': 'Taamul', 'own': 'Own Visa'}
    _visa_groups = {}
    for _row in final_rows:
        _vp = _row['employee'].visa_provider or 'own'
        if _vp not in _visa_groups:
            _visa_groups[_vp] = {'key': _vp, 'label': _visa_labels.get(_vp, _vp), 'count': 0, 'total_aed': 0.0, 'total_inr': 0.0, 'total_npr': 0.0}
        _visa_groups[_vp]['count'] += 1
        _cur = _row['currency']
        if _cur == 'AED':
            _visa_groups[_vp]['total_aed'] = round(_visa_groups[_vp]['total_aed'] + _row['final_salary'], 2)
        elif _cur == 'NPR':
            _visa_groups[_vp]['total_npr'] = round(_visa_groups[_vp]['total_npr'] + _row['final_salary'], 2)
        else:
            _visa_groups[_vp]['total_inr'] = round(_visa_groups[_vp]['total_inr'] + _row['final_salary'], 2)
    visa_breakdown = [_visa_groups[k] for k in _visa_order if k in _visa_groups]
    visa_breakdown += [v for k, v in _visa_groups.items() if k not in _visa_order]

    inr_exchange_rate = None
    inr_rate_obj = ExchangeRate.objects.filter(currency='INR', year=selected_year, month=selected_month).first()
    if inr_rate_obj:
        inr_exchange_rate = float(inr_rate_obj.rate)

    npr_exchange_rate = None
    npr_rate_obj = ExchangeRate.objects.filter(currency='NPR', year=selected_year, month=selected_month).first()
    if npr_rate_obj:
        npr_exchange_rate = float(npr_rate_obj.rate)

    final_total_inr_aed = round(final_total_inr / inr_exchange_rate, 2) if (inr_exchange_rate and inr_exchange_rate > 0) else None
    final_total_npr_aed = round(final_total_npr / npr_exchange_rate, 2) if (npr_exchange_rate and npr_exchange_rate > 0) else None

    final_total_combined_aed = final_total_aed
    if final_total_inr_aed and final_total_inr > 0:
        final_total_combined_aed += final_total_inr_aed
    if final_total_npr_aed and final_total_npr > 0:
        final_total_combined_aed += final_total_npr_aed
    final_total_combined_aed = round(final_total_combined_aed, 2)

    all_employees_json = json.dumps(
        [{'id': e.id, 'name': e.name, 'type': 'inhouse', 'dept': e.department or '', 'currency': e.currency}
         for e in all_payroll_inhouse] +
        [{'id': e.id, 'name': e.name, 'type': 'remote', 'dept': 'Remote', 'currency': e.currency}
         for e in all_payroll_remote]
    )

    def _section_totals(rows):
        return {
            'aed': round(sum(r['net_payroll'] for r in rows if r.get('currency', 'AED') == 'AED'), 2),
            'inr': round(sum(r['net_payroll'] for r in rows if r.get('currency', 'AED') == 'INR'), 2),
            'npr': round(sum(r['net_payroll'] for r in rows if r.get('currency', 'AED') == 'NPR'), 2),
        }

    _pmonth_abbr = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    period_label = f"21 {_pmonth_abbr[prev_month]} – 20 {_pmonth_abbr[selected_month]} {selected_year}"

    # Cumulative unpaid salary report — all months up to and including selected
    _frozen_months_qs = FrozenPayrollMonth.objects.filter(
        Q(year__lt=selected_year) | Q(year=selected_year, month__lte=selected_month)
    ).order_by('year', 'month')

    _paid_keys_all = set()
    for _rec in PaidSalaryRecord.objects.filter(
        Q(year__lt=selected_year) | Q(year=selected_year, month__lte=selected_month)
    ):
        if _rec.employee_id:
            _paid_keys_all.add((_rec.year, _rec.month, 'inhouse', _rec.employee_id))
        elif _rec.remote_employee_id:
            _paid_keys_all.add((_rec.year, _rec.month, 'remote', _rec.remote_employee_id))

    _unpaid_entries = []
    _frozen_ym = set()

    for _frozen in _frozen_months_qs:
        _fy, _fm = _frozen.year, _frozen.month
        _frozen_ym.add((_fy, _fm))
        for _fr in _frozen.snapshot.get('final_rows', []):
            _eid = _fr.get('employee_id')
            _etype = _fr.get('employee_type')
            if not _eid or not _etype:
                continue
            if (_fy, _fm, _etype, _eid) not in _paid_keys_all:
                _unpaid_entries.append({
                    'year': _fy,
                    'month': _fm,
                    'month_label': f"{_pmonth_abbr[_fm]} {_fy}",
                    'employee_id': _eid,
                    'employee_type': _etype,
                    'employee_name': _fr.get('employee_name', ''),
                    'department': _fr.get('employee_department', '') or 'Unknown',
                    'location': _fr.get('employee_location', '') or 'Other',
                    'currency': _fr.get('currency', 'AED'),
                    'final_salary': float(_fr.get('final_salary', 0)),
                })

    # Current month if not yet frozen
    if (selected_year, selected_month) not in _frozen_ym:
        for _fr in final_rows:
            if not _fr.get('is_paid'):
                _emp = _fr['employee']
                _unpaid_entries.append({
                    'year': selected_year,
                    'month': selected_month,
                    'month_label': f"{_pmonth_abbr[selected_month]} {selected_year}",
                    'employee_id': _emp.id,
                    'employee_type': _fr['employee_type'],
                    'employee_name': _emp.name,
                    'department': getattr(_emp, 'department', '') or 'Unknown',
                    'location': getattr(_emp, 'location', '') or 'Other',
                    'currency': _fr['currency'],
                    'final_salary': float(_fr.get('final_salary', 0)),
                })

    # Enrich with visa_provider from current DB state (snapshot doesn't store it)
    _up_inhouse_ids = {e['employee_id'] for e in _unpaid_entries if e['employee_type'] == 'inhouse'}
    _up_remote_ids = {e['employee_id'] for e in _unpaid_entries if e['employee_type'] == 'remote'}
    _inhouse_visa_map = (
        {e.id: e.visa_provider or 'own' for e in Employee.objects.filter(id__in=_up_inhouse_ids).only('id', 'visa_provider')}
        if _up_inhouse_ids else {}
    )
    _remote_visa_map = (
        {e.id: e.visa_provider or 'own' for e in RemoteEmployee.objects.filter(id__in=_up_remote_ids).only('id', 'visa_provider')}
        if _up_remote_ids else {}
    )
    for _ue in _unpaid_entries:
        _vp = (_inhouse_visa_map if _ue['employee_type'] == 'inhouse' else _remote_visa_map).get(_ue['employee_id'], 'own')
        _ue['visa_provider'] = _vp
        _ue['visa_label'] = _visa_labels.get(_vp, 'Own Visa')

    # Aggregate per employee across all their unpaid months
    _emp_agg = {}
    for _ue in _unpaid_entries:
        _ekey = (_ue['employee_type'], _ue['employee_id'])
        if _ekey not in _emp_agg:
            _emp_agg[_ekey] = {
                'employee_type': _ue['employee_type'],
                'employee_id': _ue['employee_id'],
                'employee_name': _ue['employee_name'],
                'department': _ue['department'],
                'location': _ue['location'],
                'visa_provider': _ue['visa_provider'],
                'visa_label': _ue['visa_label'],
                'currency': _ue['currency'],
                'months': [],
                'total_aed': 0.0,
                'total_inr': 0.0,
                'total_npr': 0.0,
            }
        _emp_agg[_ekey]['months'].append(_ue['month_label'])
        if _ue['currency'] == 'AED':
            _emp_agg[_ekey]['total_aed'] = round(_emp_agg[_ekey]['total_aed'] + _ue['final_salary'], 2)
        elif _ue['currency'] == 'NPR':
            _emp_agg[_ekey]['total_npr'] = round(_emp_agg[_ekey]['total_npr'] + _ue['final_salary'], 2)
        else:
            _emp_agg[_ekey]['total_inr'] = round(_emp_agg[_ekey]['total_inr'] + _ue['final_salary'], 2)

    # Sort: department → name
    unpaid_employees = sorted(_emp_agg.values(), key=lambda r: (r['department'], r['employee_name']))

    def _grp_breakdown(entries, key_field, label_field=None):
        grps = {}
        for e in entries:
            k = e.get(key_field) or 'Unknown'
            lbl = e.get(label_field, k) if label_field else k
            if k not in grps:
                grps[k] = {'key': k, 'label': lbl, 'emp_count': 0, 'total_aed': 0.0, 'total_inr': 0.0, 'total_npr': 0.0, 'rows': []}
            grps[k]['emp_count'] += 1
            grps[k]['total_aed'] = round(grps[k]['total_aed'] + e['total_aed'], 2)
            grps[k]['total_inr'] = round(grps[k]['total_inr'] + e['total_inr'], 2)
            grps[k]['total_npr'] = round(grps[k]['total_npr'] + e.get('total_npr', 0), 2)
            grps[k]['rows'].append(e)
        return sorted(grps.values(), key=lambda g: g['label'])

    unpaid_by_dept = _grp_breakdown(unpaid_employees, 'department')
    unpaid_by_visa = _grp_breakdown(unpaid_employees, 'visa_provider', 'visa_label')
    _visa_sort_idx = {k: i for i, k in enumerate(['Jumbo', 'OnTime', 'Taamul', 'own'])}
    unpaid_by_visa.sort(key=lambda g: _visa_sort_idx.get(g['key'], 99))
    unpaid_by_location = _grp_breakdown(unpaid_employees, 'location')
    unpaid_total_aed = round(sum(e['total_aed'] for e in unpaid_employees), 2)
    unpaid_total_inr = round(sum(e['total_inr'] for e in unpaid_employees), 2)
    unpaid_total_npr = round(sum(e['total_npr'] for e in unpaid_employees), 2)
    unpaid_unique_employees = len(unpaid_employees)
    unpaid_entry_count = len(_unpaid_entries)

    # Carryover schedule — all records, ordered newest-first
    carryover_schedule = []
    for co in DeductionCarryover.objects.select_related('employee', 'remote_employee').order_by('-to_year', '-to_month', '-from_year', '-from_month'):
        emp = co.employee or co.remote_employee
        if not emp:
            continue
        emp_type = 'inhouse' if co.employee else 'remote'
        remaining = round(float(co.overflow_amount) - float(co.applied_amount), 2)
        if co.is_skipped:
            status = 'skipped'
        elif remaining <= 0:
            status = 'cleared'
        elif float(co.applied_amount) > 0:
            status = 'partial'
        else:
            status = 'pending'
        carryover_schedule.append({
            'id': co.id,
            'employee': emp,
            'employee_type': emp_type,
            'currency': co.currency,
            'from_month': co.from_month,
            'from_year': co.from_year,
            'from_label': f"{_pmonth_abbr[co.from_month]} {co.from_year}",
            'to_month': co.to_month,
            'to_year': co.to_year,
            'to_label': f"{_pmonth_abbr[co.to_month]} {co.to_year}",
            'overflow_amount': float(co.overflow_amount),
            'applied_amount': float(co.applied_amount),
            'remaining': remaining,
            'status': status,
            'is_skipped': co.is_skipped,
            'skipped_by': co.skipped_by,
            'skip_reason': co.skip_reason,
            'is_incoming': (co.to_year == selected_year and co.to_month == selected_month),
            'is_outgoing': (co.from_year == selected_year and co.from_month == selected_month),
        })

    carryover_pending_count = sum(1 for c in carryover_schedule if c['status'] not in ('cleared', 'skipped'))
    carryover_pending_aed = round(sum(c['remaining'] for c in carryover_schedule if c['currency'] == 'AED' and c['status'] not in ('cleared', 'skipped')), 2)
    carryover_pending_inr = round(sum(c['remaining'] for c in carryover_schedule if c['currency'] == 'INR' and c['status'] not in ('cleared', 'skipped')), 2)
    carryover_pending_npr = round(sum(c['remaining'] for c in carryover_schedule if c['currency'] == 'NPR' and c['status'] not in ('cleared', 'skipped')), 2)
    carryover_incoming_count = sum(1 for c in carryover_schedule if c['is_incoming'])
    carryover_outgoing_count = sum(1 for c in carryover_schedule if c['is_outgoing'])

    admin_inhouse_totals = _section_totals(admin_inhouse_rows)
    admin_remote_totals = _section_totals(admin_remote_rows)
    sales_fixed_totals = _section_totals(sales_fixed_rows)
    sales_perf_totals = _section_totals(sales_perf_rows)

    # Old-method Sales: Performance rows split by employee type: in-house gets its
    # own tab (no talktime data, so it was never part of Method 2), remote is kept
    # for reference against the live Method 2 tab.
    sales_perf_inhouse_rows = [r for r in sales_perf_rows if r['employee_type'] == 'inhouse']
    sales_perf_remote_rows = [r for r in sales_perf_rows if r['employee_type'] == 'remote']
    sales_perf_inhouse_totals = _section_totals(sales_perf_inhouse_rows)
    sales_perf_remote_totals = _section_totals(sales_perf_remote_rows)

    # ---- XLSX download ----
    if request.GET.get('format') == 'xlsx':
        tab = request.GET.get('tab', 'summary')
        month_label = MONTH_NAMES[selected_month]

        # Shared styles
        _TITLE_FONT = Font(bold=True, size=13)
        _HDR_FONT = Font(bold=True, color='FFFFFF')
        _HDR_FILL_PURPLE = PatternFill(start_color='5B21B6', end_color='5B21B6', fill_type='solid')
        _HDR_FILL_BLUE = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        _HDR_FILL_GREEN = PatternFill(start_color='065F46', end_color='065F46', fill_type='solid')
        _HDR_FILL_RED = PatternFill(start_color='991B1B', end_color='991B1B', fill_type='solid')
        _SECTION_FILL = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
        _TOTAL_FILL = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        _TOTAL_FONT = Font(bold=True)
        _PAID_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        _NEG_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        _THIN = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        _RIGHT = Alignment(horizontal='right')
        _CENTER = Alignment(horizontal='center')

        def _hdr(ws, row, cols, fill=None):
            fill = fill or _HDR_FILL_PURPLE
            for col, val in enumerate(cols, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _HDR_FONT
                c.fill = fill
                c.border = _THIN
                c.alignment = _CENTER

        def _title_row(ws, text, ncols):
            ws.cell(row=1, column=1, value=text).font = _TITLE_FONT
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

        def _num_cell(ws, row, col, val, negative=False, positive=False):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = _RIGHT
            c.border = _THIN
            c.number_format = '#,##0.00'
            if negative and val and val > 0:
                c.fill = _NEG_FILL
            return c

        def _str_cell(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _THIN
            return c

        wb = Workbook()

        # ── Summary ──
        if tab == 'summary':
            ws = wb.active
            ws.title = 'Summary'
            ncols = 9
            _title_row(ws, f'Payroll Summary — {period_label}', ncols)
            ws.row_dimensions[1].height = 22

            ws.append([])

            # Section breakdown mini-table
            ws.cell(row=3, column=1, value='Section Breakdown').font = Font(bold=True)
            _hdr(ws, 4, ['Section', 'Employees', 'Net AED', 'Net INR/NPR'], _HDR_FILL_BLUE)
            sections = [
                ('Admin In-House', len(admin_inhouse_rows), admin_inhouse_totals['aed'], admin_inhouse_totals['inr'] + admin_inhouse_totals['npr']),
                ('Admin Remote', len(admin_remote_rows), admin_remote_totals['aed'], admin_remote_totals['inr'] + admin_remote_totals['npr']),
                ('Sales Fixed', len(sales_fixed_rows), sales_fixed_totals['aed'], sales_fixed_totals['inr'] + sales_fixed_totals['npr']),
                ('Sales Performance', len(sales_perf_rows), sales_perf_totals['aed'], sales_perf_totals['inr'] + sales_perf_totals['npr']),
            ]
            r = 5
            for sec_name, cnt, aed, inr in sections:
                ws.cell(row=r, column=1, value=sec_name).border = _THIN
                ws.cell(row=r, column=2, value=cnt).border = _THIN
                _num_cell(ws, r, 3, aed)
                _num_cell(ws, r, 4, inr if inr else 0)
                r += 1
            # Totals row
            for col, val in enumerate([
                'Total', len(final_rows), final_total_aed, (final_total_inr + final_total_npr) or 0
            ], 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _THIN
                if col >= 3:
                    c.alignment = _RIGHT
                    c.number_format = '#,##0.00'
            r += 2

            # Per-employee final table
            ws.cell(row=r, column=1, value='Per-Employee Final Salary').font = Font(bold=True)
            r += 1
            headers = ['Employee', 'Section', 'Currency', 'Net Payroll', 'Deductions', 'Additions',
                       'Carryover In', 'Carryover Out', 'Final Salary', 'Status']
            _hdr(ws, r, headers, _HDR_FILL_PURPLE)
            r += 1
            for row in final_rows:
                paid_label = 'PAID' if row.get('is_paid') else 'Unpaid'
                _str_cell(ws, r, 1, row['employee'].name)
                _str_cell(ws, r, 2, row['department'])
                _str_cell(ws, r, 3, row['currency'])
                _num_cell(ws, r, 4, row['payroll_net'])
                _num_cell(ws, r, 5, row['total_deductions'], negative=True)
                _num_cell(ws, r, 6, row['total_additions'])
                _num_cell(ws, r, 7, row['carryover_in'], negative=True)
                _num_cell(ws, r, 8, row['carryover_out'])
                _num_cell(ws, r, 9, row['final_salary'])
                ws.cell(row=r, column=10, value=paid_label).border = _THIN
                if row.get('is_paid'):
                    for col in range(1, 11):
                        ws.cell(row=r, column=col).fill = _PAID_FILL
                r += 1
            # Totals
            total_vals = ['', 'Total', '', '', '', '', '', '']
            for col, val in enumerate(['', 'Total', '', '',
                sum(row['payroll_net'] for row in final_rows),
                sum(row['total_deductions'] for row in final_rows),
                sum(row['total_additions'] for row in final_rows),
                sum(row['carryover_in'] for row in final_rows),
                sum(row['carryover_out'] for row in final_rows),
                sum(row['final_salary'] for row in final_rows),
            ], 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _THIN
                if col >= 5:
                    c.alignment = _RIGHT
                    c.number_format = '#,##0.00'

            # Column widths
            for col, w in enumerate([28, 18, 10, 14, 14, 14, 14, 14, 14, 10], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        # ── Admin In-House ──
        elif tab == 'admin-inhouse':
            ws = wb.active
            ws.title = 'Admin In-House'
            headers = ['Name', 'Designation', 'Salary', 'Full Days', 'Half Days',
                       'Absent', 'Late Days', 'Deduction', 'Incentives', 'Reductions',
                       'Commission', 'Net (AED)', 'Net (INR/NPR)', 'Status']
            ncols = len(headers)
            _title_row(ws, f'Admin In-House — {month_label} {selected_year}', ncols)
            ws.append([])
            _hdr(ws, 3, headers, _HDR_FILL_BLUE)
            r = 4
            for row in admin_inhouse_rows:
                _str_cell(ws, r, 1, row['employee'].name)
                _str_cell(ws, r, 2, getattr(row['employee'], 'designation', '') or '')
                _num_cell(ws, r, 3, float(row.get('salary') or 0))
                ws.cell(row=r, column=4, value=row.get('full_days', 0)).border = _THIN
                ws.cell(row=r, column=5, value=row.get('half_days', 0)).border = _THIN
                ws.cell(row=r, column=6, value=row.get('absent_days', 0)).border = _THIN
                ws.cell(row=r, column=7, value=row.get('late_days', 0)).border = _THIN
                _num_cell(ws, r, 8, float(row.get('deduction') or 0), negative=True)
                _num_cell(ws, r, 9, float(row.get('incentives') or 0))
                _num_cell(ws, r, 10, float(row.get('reductions') or 0), negative=True)
                _num_cell(ws, r, 11, float(row.get('commission') or 0))
                net = float(row.get('net_payroll') or 0)
                if row.get('currency') == 'AED':
                    _num_cell(ws, r, 12, net)
                    _str_cell(ws, r, 13, '—')
                else:
                    _str_cell(ws, r, 12, '—')
                    _num_cell(ws, r, 13, net)
                ws.cell(row=r, column=14, value='PAID' if row.get('is_paid') else '').border = _THIN
                if row.get('is_paid'):
                    for col in range(1, ncols + 1):
                        ws.cell(row=r, column=col).fill = _PAID_FILL
                r += 1
            # Totals
            tots = [None, None, None, None, None, None, None,
                    sum(float(row.get('deduction') or 0) for row in admin_inhouse_rows),
                    sum(float(row.get('incentives') or 0) for row in admin_inhouse_rows),
                    sum(float(row.get('reductions') or 0) for row in admin_inhouse_rows),
                    sum(float(row.get('commission') or 0) for row in admin_inhouse_rows),
                    admin_inhouse_totals['aed'],
                    (admin_inhouse_totals['inr'] + admin_inhouse_totals['npr']) or 0,
                    None]
            for col, val in enumerate(tots, 1):
                c = ws.cell(row=r, column=col, value='Total' if col == 1 else val)
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _THIN
                if col >= 8 and val is not None:
                    c.alignment = _RIGHT
                    c.number_format = '#,##0.00'
            for col, w in enumerate([28, 18, 10, 10, 10, 10, 10, 12, 12, 12, 12, 14, 14, 8], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        # ── Admin Remote ──
        elif tab == 'admin-remote':
            ws = wb.active
            ws.title = 'Admin Remote'
            headers = ['Name', 'Currency', 'Salary', 'Present Days', 'Absent Days',
                       'Deduction', 'Incentives', 'Reductions', 'Net (AED)', 'Net (INR/NPR)', 'Status']
            ncols = len(headers)
            _title_row(ws, f'Admin Remote — {month_label} {selected_year}', ncols)
            ws.append([])
            _hdr(ws, 3, headers, _HDR_FILL_BLUE)
            r = 4
            for row in admin_remote_rows:
                _str_cell(ws, r, 1, row['employee'].name)
                _str_cell(ws, r, 2, row.get('currency', 'AED'))
                _num_cell(ws, r, 3, float(row.get('salary') or 0))
                ws.cell(row=r, column=4, value=row.get('present_days', '')).border = _THIN
                ws.cell(row=r, column=5, value=row.get('absent_days', '')).border = _THIN
                _num_cell(ws, r, 6, float(row.get('deduction') or 0), negative=True)
                _num_cell(ws, r, 7, float(row.get('incentives') or 0))
                _num_cell(ws, r, 8, float(row.get('reductions') or 0), negative=True)
                net = float(row.get('net_payroll') or 0)
                if row.get('currency') == 'AED':
                    _num_cell(ws, r, 9, net)
                    _str_cell(ws, r, 10, '—')
                else:
                    _str_cell(ws, r, 9, '—')
                    _num_cell(ws, r, 10, net)
                ws.cell(row=r, column=11, value='PAID' if row.get('is_paid') else '').border = _THIN
                if row.get('is_paid'):
                    for col in range(1, ncols + 1):
                        ws.cell(row=r, column=col).fill = _PAID_FILL
                r += 1
            for col, val in enumerate([
                'Total', None, None, None,
                sum(float(row.get('absent_days') or 0) for row in admin_remote_rows),
                sum(float(row.get('deduction') or 0) for row in admin_remote_rows),
                sum(float(row.get('incentives') or 0) for row in admin_remote_rows),
                sum(float(row.get('reductions') or 0) for row in admin_remote_rows),
                admin_remote_totals['aed'],
                (admin_remote_totals['inr'] + admin_remote_totals['npr']) or 0,
                None,
            ], 1):
                c = ws.cell(row=r, column=col, value='Total' if col == 1 else val)
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _THIN
                if col >= 5 and val is not None:
                    c.alignment = _RIGHT
                    c.number_format = '#,##0.00'
            for col, w in enumerate([28, 10, 12, 14, 14, 14, 14, 14, 14, 14, 8], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        # ── Sales Fixed ──
        elif tab == 'sales-fixed':
            ws = wb.active
            ws.title = 'Sales Fixed'
            headers = ['Name', 'Type', 'Currency', 'Salary', 'Present Days', 'Absent Days',
                       'Deduction', 'Commission', 'Incentives', 'Reductions',
                       'Net (AED)', 'Net (INR/NPR)', 'Status']
            ncols = len(headers)
            _title_row(ws, f'Sales — Fixed Salary — {month_label} {selected_year}', ncols)
            ws.append([])
            _hdr(ws, 3, headers, _HDR_FILL_GREEN)
            r = 4
            for row in sales_fixed_rows:
                _str_cell(ws, r, 1, row['employee'].name)
                _str_cell(ws, r, 2, row.get('employee_type', '').capitalize())
                _str_cell(ws, r, 3, row.get('currency', 'AED'))
                _num_cell(ws, r, 4, float(row.get('salary') or 0))
                ws.cell(row=r, column=5, value=row.get('present_days', '')).border = _THIN
                ws.cell(row=r, column=6, value=row.get('absent_days', '')).border = _THIN
                _num_cell(ws, r, 7, float(row.get('deduction') or 0), negative=True)
                _num_cell(ws, r, 8, float(row.get('commission') or 0))
                _num_cell(ws, r, 9, float(row.get('incentives') or 0))
                _num_cell(ws, r, 10, float(row.get('reductions') or 0), negative=True)
                net = float(row.get('net_payroll') or 0)
                if row.get('currency') == 'AED':
                    _num_cell(ws, r, 11, net)
                    _str_cell(ws, r, 12, '—')
                else:
                    _str_cell(ws, r, 11, '—')
                    _num_cell(ws, r, 12, net)
                ws.cell(row=r, column=13, value='PAID' if row.get('is_paid') else '').border = _THIN
                if row.get('is_paid'):
                    for col in range(1, ncols + 1):
                        ws.cell(row=r, column=col).fill = _PAID_FILL
                r += 1
            for col, val in enumerate([
                'Total', None, None, None, None,
                sum(float(row.get('absent_days') or 0) for row in sales_fixed_rows),
                sum(float(row.get('deduction') or 0) for row in sales_fixed_rows),
                sum(float(row.get('commission') or 0) for row in sales_fixed_rows),
                sum(float(row.get('incentives') or 0) for row in sales_fixed_rows),
                sum(float(row.get('reductions') or 0) for row in sales_fixed_rows),
                sales_fixed_totals['aed'],
                (sales_fixed_totals['inr'] + sales_fixed_totals['npr']) or 0,
                None,
            ], 1):
                c = ws.cell(row=r, column=col, value='Total' if col == 1 else val)
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _THIN
                if col >= 6 and val is not None:
                    c.alignment = _RIGHT
                    c.number_format = '#,##0.00'
            for col, w in enumerate([28, 10, 10, 12, 14, 14, 14, 14, 14, 14, 14, 14, 8], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        # ── Sales Performance (Old Method: all / in-house only / remote only) ──
        elif tab in ('sales-perf', 'sales-perf-inhouse', 'sales-perf-remote'):
            _sp_rows, _sp_totals, _sp_title = {
                'sales-perf': (sales_perf_rows, sales_perf_totals, 'Sales — Performance Based (Old Method)'),
                'sales-perf-inhouse': (sales_perf_inhouse_rows, sales_perf_inhouse_totals, 'Sales — Performance In-House (Old Method)'),
                'sales-perf-remote': (sales_perf_remote_rows, sales_perf_remote_totals, 'Sales — Performance Remote (Old Method)'),
            }[tab]
            ws = wb.active
            ws.title = 'Sales Performance'
            bank_headers = [b.name for b in banks]
            headers = ['Name', 'Type', 'Currency', 'Base Salary', 'Deductions'] + \
                      bank_headers + ['Commission', 'Incentives', 'Net (AED)', 'Net (INR/NPR)', 'Status']
            ncols = len(headers)
            _title_row(ws, f'{_sp_title} — {month_label} {selected_year}', ncols)
            ws.append([])
            _hdr(ws, 3, headers, _HDR_FILL_RED)
            r = 4
            for row in _sp_rows:
                _str_cell(ws, r, 1, row['employee'].name)
                _str_cell(ws, r, 2, row.get('employee_type', '').capitalize())
                _str_cell(ws, r, 3, row.get('currency', 'AED'))
                _num_cell(ws, r, 4, float(row.get('salary') or 0))
                _num_cell(ws, r, 5, float(row.get('combined_deductions') or 0), negative=True)
                bank_counts = row.get('bank_counts_list') or []
                for i, b in enumerate(banks):
                    count = bank_counts[i] if i < len(bank_counts) else 0
                    ws.cell(row=r, column=6 + i, value=count or 0).border = _THIN
                col_off = 6 + len(banks)
                _num_cell(ws, r, col_off, float(row.get('commission') or 0))
                _num_cell(ws, r, col_off + 1, float(row.get('incentives') or 0))
                net = float(row.get('net_payroll') or 0)
                if row.get('currency') == 'AED':
                    _num_cell(ws, r, col_off + 2, net)
                    _str_cell(ws, r, col_off + 3, '—')
                else:
                    _str_cell(ws, r, col_off + 2, '—')
                    _num_cell(ws, r, col_off + 3, net)
                ws.cell(row=r, column=col_off + 4, value='PAID' if row.get('is_paid') else '').border = _THIN
                if row.get('is_paid'):
                    for col in range(1, ncols + 1):
                        ws.cell(row=r, column=col).fill = _PAID_FILL
                r += 1
            # Totals
            perf_bank_totals = [
                sum((row.get('bank_counts_list') or [0] * len(banks))[i] if i < len(row.get('bank_counts_list') or []) else 0
                    for row in _sp_rows)
                for i in range(len(banks))
            ]
            tot_vals = ['Total', None, None, None,
                        sum(float(row.get('combined_deductions') or 0) for row in _sp_rows),
                        ] + perf_bank_totals + [
                        sum(float(row.get('commission') or 0) for row in _sp_rows),
                        sum(float(row.get('incentives') or 0) for row in _sp_rows),
                        _sp_totals['aed'],
                        (_sp_totals['inr'] + _sp_totals['npr']) or 0,
                        None,
                       ]
            for col, val in enumerate(tot_vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font = _TOTAL_FONT
                c.fill = _TOTAL_FILL
                c.border = _THIN
                if col >= 5 and val is not None and isinstance(val, (int, float)):
                    c.alignment = _RIGHT
                    c.number_format = '#,##0.00'
            for col, w in enumerate([28, 10, 10, 14, 14] + [12] * len(banks) + [14, 14, 14, 14, 8], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        # ── Deductions ──
        elif tab == 'deductions':
            ws = wb.active
            ws.title = 'Deductions & Additions'
            headers = ['Employee', 'Emp Type', 'Category', 'Direction', 'Note',
                       'Currency', 'Total Amount', 'Split Months', 'This Month (Installment)',
                       'Start Period', 'End Period', 'Recovery Status']
            ncols = len(headers)
            _title_row(ws, f'Deductions & Additions — {month_label} {selected_year}', ncols)
            ws.append([])
            _hdr(ws, 3, headers, _HDR_FILL_PURPLE)
            r = 4
            for entry in sorted(all_deductions_list, key=lambda e: e['employee'].name.lower()):
                end_period = f"{entry['end_month_name']} {entry['end_year']}" if entry['split_months'] > 1 else ''
                _str_cell(ws, r, 1, entry['employee'].name)
                _str_cell(ws, r, 2, entry['employee_type'].capitalize())
                _str_cell(ws, r, 3, entry['category_display'])
                direction = 'Deduction' if entry['entry_type'] == 'deduction' else 'Addition'
                _str_cell(ws, r, 4, direction)
                _str_cell(ws, r, 5, entry['note'] or '')
                _str_cell(ws, r, 6, entry['currency'])
                _num_cell(ws, r, 7, float(entry['total_amount']))
                ws.cell(row=r, column=8, value=entry['split_months']).border = _THIN
                inst = float(entry['installment_amount'])
                c = ws.cell(row=r, column=9, value=inst)
                c.alignment = _RIGHT
                c.border = _THIN
                c.number_format = '#,##0.00'
                if entry['entry_type'] == 'deduction':
                    c.fill = _NEG_FILL
                else:
                    c.fill = _PAID_FILL
                _str_cell(ws, r, 10, f"{entry['start_month_name']} {entry['start_year']}")
                _str_cell(ws, r, 11, end_period)
                if entry['is_active_this_month']:
                    _status = 'Active This Month'
                elif entry['is_upcoming']:
                    _status = 'Upcoming'
                elif not entry['is_finished']:
                    _status = 'In Progress'
                elif entry['entry_type'] == 'deduction':
                    _status = 'Recovered' if entry['is_recovered'] else 'Pending Recovery'
                else:
                    _status = 'Completed'
                _str_cell(ws, r, 12, _status)
                r += 1
            for col, w in enumerate([28, 12, 22, 12, 30, 10, 14, 14, 22, 16, 16, 14], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        # ── Carryovers ──
        elif tab == 'carryovers':
            ws = wb.active
            ws.title = 'Carryover Schedule'
            headers = ['Employee', 'Type', 'Currency',
                       'Salary went -ve in', 'Deducted from',
                       'Total to Deduct', 'Already Deducted', 'Still to Deduct',
                       'Status']
            ncols = len(headers)
            _title_row(ws, f'Deduction Carryover Schedule — {month_label} {selected_year}', ncols)
            ws.append([])
            _hdr(ws, 3, headers, _HDR_FILL_RED)
            r = 4
            for co in carryover_schedule:
                _str_cell(ws, r, 1, co['employee'].name)
                _str_cell(ws, r, 2, co['employee_type'].capitalize())
                _str_cell(ws, r, 3, co['currency'])
                _str_cell(ws, r, 4, co['from_label'])
                _str_cell(ws, r, 5, co['to_label'])
                _num_cell(ws, r, 6, co['overflow_amount'])
                _num_cell(ws, r, 7, co['applied_amount'])
                c8 = _num_cell(ws, r, 8, co['remaining'], negative=(co['remaining'] > 0))
                status_map = {'cleared': 'FULLY RECOVERED', 'partial': 'PARTIALLY DEDUCTED', 'pending': 'NOT YET DEDUCTED', 'skipped': 'SKIPPED / WAIVED'}
                _str_cell(ws, r, 9, status_map.get(co['status'], co['status']))
                if co['status'] in ('cleared', 'skipped'):
                    for col in range(1, ncols + 1):
                        ws.cell(row=r, column=col).fill = _PAID_FILL
                r += 1
            # Summary rows
            r += 1
            ws.cell(row=r, column=1, value=f'Pending: {carryover_pending_count} unresolved').font = Font(bold=True)
            ws.cell(row=r, column=6, value=carryover_pending_aed).number_format = '#,##0.00'
            ws.cell(row=r, column=7, value='AED balance').font = Font(italic=True)
            for col, w in enumerate([28, 12, 10, 18, 18, 18, 18, 18, 20], 1):
                ws.column_dimensions[_get_col_letter(col)].width = w

        else:
            ws = wb.active
            ws.title = 'Payroll'
            ws.cell(row=1, column=1, value=f'No data for tab: {tab}')

        # Freeze top rows and set filename
        try:
            ws.freeze_panes = ws.cell(row=4, column=1)
        except Exception:
            pass

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_tab = tab.replace('-', '_')
        filename = f"Payroll_{month_label}_{selected_year}_{safe_tab}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        logger.info("Payroll XLSX downloaded: tab=%s %s/%s by %s", tab, selected_month, selected_year, request.user.username)
        return response

    # ---- Phase E1: Workforce & ledger categories + Payment status summary ----
    # Read-only aggregation over data already computed above (final_rows,
    # admin_inhouse_rows, etc.) — no new queries, no new state. "Payment
    # status" is intentionally a two-bucket Paid/Unpaid split: today's
    # PaidSalaryRecord model only tracks a binary is_paid flag, there is no
    # partial-payment concept yet, so we don't fabricate one here.
    _wf_total_employees = len(final_rows)
    _wf_categories = [
        {'key': 'admin_inhouse', 'label': 'Admin: In-House', 'count': len(admin_inhouse_rows)},
        {'key': 'admin_remote', 'label': 'Admin: Remote', 'count': len(admin_remote_rows)},
        {'key': 'sales_fixed', 'label': 'Sales: Fixed', 'count': len(sales_fixed_rows)},
        {'key': 'sales_perf_remote', 'label': 'Sales Perf: Remote', 'count': len(sales_perf_remote_rows)},
        {'key': 'sales_perf_inhouse', 'label': 'Sales Perf: In-House', 'count': len(sales_perf_inhouse_rows)},
    ]
    for _wf_cat in _wf_categories:
        _wf_cat['pct'] = round((_wf_cat['count'] / _wf_total_employees * 100), 1) if _wf_total_employees else 0

    # ---- Phase E6: three-way payment status --------------------------------
    # "Paid" now means paid IN FULL. A partially-paid employee is its own
    # bucket, reported as amount disbursed against amount owed, and the balance
    # to pay is the sum of every outstanding amount — unpaid rows in full, plus
    # the still-owed remainder of partial ones. Before E6 the balance was just
    # the unpaid total, which would now understate what the business still owes.
    _full = [r for r in final_rows if r.get('pay_status') == 'paid']
    _part = [r for r in final_rows if r.get('pay_status') == 'partial']
    _none = [r for r in final_rows if r.get('pay_status', 'unpaid') == 'unpaid']

    def _sum(rows, field, currency):
        return round(sum(r.get(field, 0) or 0 for r in rows if r['currency'] == currency), 2)

    payment_status_summary = {
        'paid_count': len(_full),
        'partial_count': len(_part),
        'unpaid_count': len(_none),
        'paid_aed': _sum(_full, 'final_salary', 'AED'),
        'paid_inr': _sum(_full, 'final_salary', 'INR'),
        'paid_npr': _sum(_full, 'final_salary', 'NPR'),
        # For partials: what actually went out, and what it was owed against.
        'partial_paid_aed': _sum(_part, 'amount_paid', 'AED'),
        'partial_owed_aed': _sum(_part, 'final_salary', 'AED'),
        'partial_paid_inr': _sum(_part, 'amount_paid', 'INR'),
        'partial_owed_inr': _sum(_part, 'final_salary', 'INR'),
        'partial_paid_npr': _sum(_part, 'amount_paid', 'NPR'),
        'partial_owed_npr': _sum(_part, 'final_salary', 'NPR'),
        'unpaid_aed': _sum(_none, 'final_salary', 'AED'),
        'unpaid_inr': _sum(_none, 'final_salary', 'INR'),
        'unpaid_npr': _sum(_none, 'final_salary', 'NPR'),
        # Balance still to pay = unpaid in full + unrecovered remainder of partials.
        'balance_aed': round(_sum(_none, 'final_salary', 'AED') + _sum(_part, 'balance_due', 'AED'), 2),
        'balance_inr': round(_sum(_none, 'final_salary', 'INR') + _sum(_part, 'balance_due', 'INR'), 2),
        'balance_npr': round(_sum(_none, 'final_salary', 'NPR') + _sum(_part, 'balance_due', 'NPR'), 2),
    }

    # ---- Phase E4: Deductions & Carryovers ledger summary modal ----
    # Read-only aggregation over all_deductions_list (already built above) —
    # no new queries. Carryover figures reuse the carryover_pending_* totals
    # already computed and passed into the context below.
    _ledger_deduction_entries = 0
    _ledger_addition_entries = 0
    _ded_cat_counts = {}
    for _entry in all_deductions_list:
        if _entry['entry_type'] == 'deduction':
            _ledger_deduction_entries += 1
            _ded_cat_counts[_entry['category_display']] = _ded_cat_counts.get(_entry['category_display'], 0) + 1
        else:
            _ledger_addition_entries += 1
    ledger_deduction_categories = sorted(
        ({'label': label, 'count': count} for label, count in _ded_cat_counts.items()),
        key=lambda c: -c['count']
    )

    return render(request, 'payroll/test_dashboard.html', {
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_name': MONTH_NAMES[selected_month],
        'period_label': period_label,
        'period_start': default_period_start,
        'period_end': default_period_end,
        'months': MONTH_CHOICES,
        'years': YEAR_RANGE,
        'days_in_month': default_days,
        'total_holidays': default_holidays,
        'banks': banks,
        'banks_json': banks_json,
        'admin_inhouse_rows': admin_inhouse_rows,
        'admin_inhouse_totals': admin_inhouse_totals,
        'admin_remote_rows': admin_remote_rows,
        'admin_remote_totals': admin_remote_totals,
        'sales_fixed_rows': sales_fixed_rows,
        'sales_fixed_totals': sales_fixed_totals,
        'sales_perf_rows': sales_perf_rows,
        'sales_perf_totals': sales_perf_totals,
        'sales_perf_bank_totals': [
            sum(row['bank_counts_list'][i] if row.get('bank_counts_list') and i < len(row['bank_counts_list']) else 0
                for row in sales_perf_rows)
            for i in range(len(banks))
        ],
        'sales_perf_inhouse_rows': sales_perf_inhouse_rows,
        'sales_perf_inhouse_totals': sales_perf_inhouse_totals,
        'sales_perf_inhouse_bank_totals': [
            sum(row['bank_counts_list'][i] if row.get('bank_counts_list') and i < len(row['bank_counts_list']) else 0
                for row in sales_perf_inhouse_rows)
            for i in range(len(banks))
        ],
        'sales_perf_remote_rows': sales_perf_remote_rows,
        'sales_perf_remote_totals': sales_perf_remote_totals,
        'sales_perf_remote_bank_totals': [
            sum(row['bank_counts_list'][i] if row.get('bank_counts_list') and i < len(row['bank_counts_list']) else 0
                for row in sales_perf_remote_rows)
            for i in range(len(banks))
        ],
        'sales_perf_test_rows': sales_perf_test_rows,
        'sales_perf_test_totals': sales_perf_test_totals,
        'deduction_rows': deduction_rows,
        'all_deductions_list': all_deductions_list,
        'deduction_groups': deduction_groups,
        'final_rows': final_rows,
        'final_total_aed': final_total_aed,
        'final_total_inr': final_total_inr,
        'final_total_npr': final_total_npr,
        'final_total_inr_aed': final_total_inr_aed,
        'final_total_npr_aed': final_total_npr_aed,
        'final_total_combined_aed': final_total_combined_aed,
        'visa_breakdown': visa_breakdown,
        'unpaid_employees': unpaid_employees,
        'unpaid_by_dept': unpaid_by_dept,
        'unpaid_by_visa': unpaid_by_visa,
        'unpaid_by_location': unpaid_by_location,
        'unpaid_total_aed': unpaid_total_aed,
        'unpaid_total_inr': unpaid_total_inr,
        'unpaid_total_npr': unpaid_total_npr,
        'unpaid_unique_employees': unpaid_unique_employees,
        'unpaid_entry_count': unpaid_entry_count,
        'inr_exchange_rate': inr_exchange_rate,
        'npr_exchange_rate': npr_exchange_rate,
        'all_employees_json': all_employees_json,
        'DEDUCTION_CATEGORY_CHOICES': DEDUCTION_CATEGORY_CHOICES,
        'carryover_schedule': carryover_schedule,
        'carryover_pending_count': carryover_pending_count,
        'carryover_pending_aed': carryover_pending_aed,
        'carryover_pending_inr': carryover_pending_inr,
        'carryover_pending_npr': carryover_pending_npr,
        'carryover_incoming_count': carryover_incoming_count,
        'carryover_outgoing_count': carryover_outgoing_count,
        'is_frozen': is_frozen,
        'frozen_at': frozen_obj.frozen_at if frozen_obj else None,
        'frozen_by': frozen_obj.frozen_by if frozen_obj else None,
        # Phase E1 — Workforce & ledger categories / Payment status summary
        'wf_categories': _wf_categories,
        'wf_total_employees': _wf_total_employees,
        'ledger_deductions_count': len(all_deductions_list),
        'ledger_carryovers_count': len(carryover_schedule),
        'payment_status_summary': payment_status_summary,
        # Phase E10 — the category colour key. Defined once here so the legend
        # include, the avatar chips and the row stripes can never list a
        # different set of categories from one another.
        'category_legend_items': [
            ('admin-in-house', 'Admin: In-House'),
            ('admin-remote', 'Admin: Remote'),
            ('sales-fixed', 'Sales: Fixed'),
            ('sales-perf-remote', 'Sales Perf: Remote'),
            # NB: 'in-house' with the hyphen — this must match the slug produced
            # by category_slug above ("Sales Perf: In-House" -> lowercased, ": "
            # and " " both to "-"), or the swatch silently loses its colour.
            ('sales-perf-in-house', 'Sales Perf: In-House'),
        ],
        # Phase E6: distinct sections present in final_rows, for the category
        # filter dropdown. Derived from the rows themselves rather than a fixed
        # list, so a new department never silently drops out of the filter.
        'final_sections': sorted({r['department'] for r in final_rows if r.get('department')}),
        # Phase E4 — Deductions & Carryovers ledger summary modal
        'ledger_deduction_categories': ledger_deduction_categories,
        'ledger_deduction_entries': _ledger_deduction_entries,
        'ledger_addition_entries': _ledger_addition_entries,
    })


def _payment_method_label(method, emp):
    """Human label for a payment method (Phase E6).

    Bank Transfer is qualified with the employee's manpower visa provider —
    'Bank Transfer (Jumbo)' — because in practice that is what distinguishes
    one transfer run from another. Own-visa employees have no provider set, so
    they stay plain 'Bank Transfer' rather than showing an empty bracket.
    """
    if not method:
        return ''
    if method != 'bank_transfer':
        return dict(PAYMENT_METHOD_CHOICES).get(method, method)
    provider = getattr(emp, 'visa_provider', None)
    return f'Bank Transfer ({provider})' if provider else 'Bank Transfer'


def _sync_wps_shortfall_advance(emp, emp_type, year, month, final_salary,
                                amount_paid, payment_method, actor):
    """Phase E6 — WPS shortfall becomes a recoverable advance.

    When an employee's WPS-registered salary is lower than their computed net
    salary, the bank can only disburse the registered figure. The difference is
    not a missing payment — the business still owes it — so it is booked as an
    'advance' DeductionEntry starting the following month, where the existing
    deduction/carryover machinery recovers it automatically. If that month also
    cannot absorb it, the standing carryover logic rolls the balance forward on
    its own; nothing extra is needed here.

    Only WPS does this. A short Cash or Bank Transfer payment is a deliberate
    part-payment decision, not a registration ceiling, so it simply leaves a
    pending balance (see PaidSalaryRecord.balance_due).

    Idempotent by design: mark_paid_salary uses update_or_create and can be
    re-run on the same employee/month (e.g. correcting an amount). The entry is
    therefore keyed by a deterministic marker in `note` and updated in place —
    re-running never stacks duplicate advances, and correcting a payment upward
    to the full amount removes the advance entirely.
    """
    marker = f'[wps-shortfall:{year}-{month:02d}]'
    emp_filter = (
        {'employee': emp, 'remote_employee': None} if emp_type == 'inhouse'
        else {'remote_employee': emp, 'employee': None}
    )
    existing = DeductionEntry.objects.filter(
        category='advance', note__startswith=marker, **emp_filter
    ).first()

    shortfall = final_salary - amount_paid
    if payment_method != 'wps' or shortfall <= 0:
        # No shortfall (or no longer a WPS payment): retract any advance this
        # same month previously raised, so a corrected payment doesn't leave a
        # phantom deduction hanging over next month.
        if existing:
            existing.delete()
            logger.info(
                'WPS shortfall advance retracted for %s %s/%s (now settled in full)',
                emp.name, month, year,
            )
        return

    # Recover starting the month after the one being paid.
    _next_idx = year * 12 + (month - 1) + 1
    _rec_year, _rec_month = divmod(_next_idx, 12)
    _rec_month += 1

    note_text = (
        f'{marker} Auto-raised: WPS disbursed {amount_paid} of {final_salary} '
        f'{emp.currency} for {month}/{year}. Shortfall recorded as an advance '
        f'for recovery from {_rec_month}/{_rec_year} payroll.'
    )

    if existing:
        existing.total_amount = shortfall
        existing.start_year = _rec_year
        existing.start_month = _rec_month
        existing.note = note_text
        existing.currency = emp.currency
        existing.save(update_fields=[
            'total_amount', 'start_year', 'start_month', 'note', 'currency',
        ])
    else:
        DeductionEntry.objects.create(
            category='advance',
            total_amount=shortfall,
            currency=emp.currency,
            split_months=1,
            start_year=_rec_year,
            start_month=_rec_month,
            note=note_text,
            **emp_filter,
        )

    # Surface it on the employee's Notes & Timeline so the shortfall is visible
    # where payroll staff actually look, not only in the deductions ledger.
    PayrollNote.objects.create(
        text=(
            f'WPS shortfall {shortfall} {emp.currency} recorded as an advance — '
            f'auto-recovery scheduled for {_rec_month}/{_rec_year}.'
        ),
        created_by=actor or 'system',
        **emp_filter,
    )


@login_required
@user_passes_test(section_required('payroll'))
def mark_paid_salary(request):
    """
    Re-compute the full payroll for the given employees and lock every value
    into an immutable snapshot. Future changes to attendance, employee settings,
    bank rates, or deductions will not affect the stored figures.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        year = int(data['year'])
        month = int(data['month'])
        emp_list = data['employees']  # [{id, type, amount?}]
    except (KeyError, ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'error': str(e)}, status=400)

    # ---- Phase E6: payment execution details --------------------------------
    # Method and value date apply to the whole batch (one "Mark as Paid" action
    # settles a selection the same way); the disbursed amount is per employee,
    # since a partial payment is an individual decision. An omitted amount means
    # "pay in full", which keeps every pre-E6 caller working unchanged.
    payment_method = (data.get('payment_method') or '').strip()
    # 'mixed' is a derived value (set automatically when an employee's payment
    # is split across methods below) — never a method someone can pick directly.
    _valid_methods = {m[0] for m in PAYMENT_METHOD_CHOICES if m[0] != 'mixed'}
    if payment_method and payment_method not in _valid_methods:
        return JsonResponse({'error': f'Unknown payment method: {payment_method}'}, status=400)

    payment_date = None
    if data.get('payment_date'):
        try:
            payment_date = datetime.datetime.strptime(data['payment_date'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'error': 'payment_date must be YYYY-MM-DD'}, status=400)

    # Per-employee disbursed amount, keyed the same way as the employee list.
    # A blank/absent value is full payment; a negative one is rejected outright
    # rather than silently clamped, since it would corrupt the balance figures.
    #
    # An employee may instead carry `splits: [{method, amount}, ...]` — the
    # same payment settled across more than one method (e.g. part WPS, part
    # cash). Its total becomes that employee's disbursed amount and overrides
    # `amount`/the batch-level payment_method for that employee only.
    _split_date = str(payment_date or datetime.date.today())

    amount_by_key = {}
    splits_by_key = {}
    for e in emp_list:
        key = (e['type'], int(e['id']))
        raw_splits = e.get('splits')
        if raw_splits:
            parsed = []
            total = Decimal('0')
            for s in raw_splits:
                method = (s.get('method') or '').strip()
                if method not in _valid_methods:
                    return JsonResponse({'error': f'Unknown payment method: {method}'}, status=400)
                try:
                    amt = Decimal(str(s.get('amount')))
                except (ArithmeticError, ValueError, TypeError):
                    return JsonResponse({'error': f"Invalid split amount for employee {e.get('id')}"}, status=400)
                if amt < 0:
                    return JsonResponse({'error': 'Paid amount cannot be negative'}, status=400)
                if amt == 0:
                    continue
                parsed.append({'method': method, 'amount': str(amt), 'date': _split_date})
                total += amt
            if not parsed:
                return JsonResponse({'error': f"At least one payment method with an amount is required for employee {e.get('id')}"}, status=400)
            splits_by_key[key] = parsed
            amount_by_key[key] = total
            continue
        if e.get('amount') in (None, ''):
            continue
        try:
            amt = Decimal(str(e['amount']))
        except (ArithmeticError, ValueError, TypeError):
            return JsonResponse({'error': f"Invalid amount for employee {e.get('id')}"}, status=400)
        if amt < 0:
            return JsonResponse({'error': 'Paid amount cannot be negative'}, status=400)
        amount_by_key[key] = amt

    inhouse_ids = {int(e['id']) for e in emp_list if e['type'] == 'inhouse'}
    remote_ids = {int(e['id']) for e in emp_list if e['type'] == 'remote'}

    banks = list(Bank.objects.filter(is_active=True).order_by('name'))

    _DED_COLS = ['advance', 'visa_status_change', 'clawback', 'leave_deduction', 'late_deduction', 'other_deduction']
    _ADD_COLS = ['last_month_balance', 'paid_leave', 'other_addition']
    _ALL_CATS = _DED_COLS + _ADD_COLS
    target_idx = year * 12 + (month - 1)

    _period_cache = {}
    def _emp_period_mp(emp):
        day = emp.salary_cycle_start_day or 21
        if day not in _period_cache:
            _period_cache[day] = _get_employee_pay_period(day, year, month)
        return _period_cache[day]

    # Active deduction entries for this month
    active_by_emp = _defaultdict(list)
    for entry in DeductionEntry.objects.select_related('employee', 'remote_employee'):
        emp_obj = entry.employee or entry.remote_employee
        if not emp_obj:
            continue
        et = 'inhouse' if entry.employee_id else 'remote'
        start_idx = entry.start_year * 12 + (entry.start_month - 1)
        if start_idx <= target_idx < start_idx + entry.split_months:
            active_by_emp[(et, emp_obj.id)].append(entry)

    # Incoming carryovers
    incoming_carryovers = DeductionCarryover.objects.filter(to_year=year, to_month=month).exclude(is_skipped=True)
    carryover_by_emp = {}
    for co in incoming_carryovers:
        key = ('inhouse', co.employee_id) if co.employee_id else ('remote', co.remote_employee_id)
        carryover_by_emp[key] = float(co.overflow_amount)

    inhouse_summaries = {
        s.employee_id: s for s in MonthlySummary.objects.filter(year=year, month=month)
    }

    now_utc = datetime.datetime.now(datetime.timezone.utc)
    count = 0

    def _build_bank_submissions(row):
        result = []
        counts = row.get('bank_counts_list') or []
        for i, b in enumerate(banks):
            result.append({
                'bank_id': b.id, 'bank_name': b.name,
                'count': counts[i] if i < len(counts) else 0,
                'rate_aed': float(b.per_account_charge),
                'rate_inr': float(b.inr_per_account_charge) if b.inr_per_account_charge else None,
            })
        return result

    def _compute_deductions(emp, emp_type, p_start, p_end, p_days):
        cat = {c: 0.0 for c in _ALL_CATS}
        for ded_entry in active_by_emp.get((emp_type, emp.id), []):
            cat[ded_entry.category] = round(cat[ded_entry.category] + float(ded_entry.installment_amount), 2)
        carryover_in = carryover_by_emp.get((emp_type, emp.id), 0.0)
        if emp_type == 'inhouse':
            summary = inhouse_summaries.get(emp.id)
            if emp.salary and emp.payroll_type != 'performance':
                daily = float(emp.salary) / p_days
                bridge_count = len(get_bridge_sunday_days(emp, p_start, p_end))
                cat['leave_deduction'] = round(daily * ((summary.leave_days if summary else 0) + bridge_count), 2)
                cat['late_deduction'] = round(daily * ((summary.late_days if summary else 0) // 3) * 0.5, 2)
            _has_salary = emp.salary and emp.payroll_type != 'performance'
            ded_for_total = [c for c in _DED_COLS if c not in ('leave_deduction', 'late_deduction')] if (emp.department == 'Admin' or _has_salary) else _DED_COLS
        else:
            ded_for_total = _DED_COLS
        total_ded = round(sum(cat[c] for c in ded_for_total) + carryover_in, 2)
        total_add = round(sum(cat[c] for c in _ADD_COLS), 2)
        return cat, carryover_in, total_ded, total_add

    def _save_snapshot(emp, emp_type, row, cat, carryover_in, total_ded, total_add, p_start, p_end, p_days, p_hols):
        payroll_net = row['net_payroll']
        final_salary = round(payroll_net - total_ded + total_add, 2)
        carryover_out = 0.0
        if final_salary < 0:
            carryover_out = round(abs(final_salary), 2)
            final_salary = 0.0

        snapshot = {
            # Identity & settings at lock time
            'employee_name': emp.name,
            'employee_type': emp_type,
            'department': emp.department or '',
            'currency': emp.currency,
            'designation': getattr(emp, 'designation', '') or '',
            'salary': float(emp.salary or 0),
            'payroll_type': emp.payroll_type or 'attendance',
            'is_fixed_salary': emp.is_fixed_salary,
            # Pay period
            'pay_period_start': str(p_start),
            'pay_period_end': str(p_end),
            'days_in_period': p_days,
            'total_holidays': p_hols,
            # Attendance
            'full_days': row.get('full_days'),
            'half_days': row.get('half_days'),
            'absent_days': row.get('absent_days'),
            'late_days': row.get('late_days'),
            'late_half_days': row.get('late_half_days', 0),
            'present_days': row.get('present_days'),
            'is_fixed_salary_attendance': row.get('is_fixed_salary', False),
            'is_attendance_based': row.get('is_attendance_based', False),
            # Payroll line items
            'net_payroll': payroll_net,
            'deduction': row.get('deduction', 0),
            'commission': row.get('commission', 0),
            'incentives': row.get('incentives', 0),
            'reductions': row.get('reductions', 0),
            'bank_submissions': _build_bank_submissions(row),
            # Deductions & additions breakdown
            'deductions_breakdown': cat,
            'carryover_in': carryover_in,
            'total_deductions': total_ded,
            'total_additions': total_add,
            # Final
            'carryover_out': carryover_out,
            'final_salary': final_salary,
        }

        # ---- Phase E6: what was actually disbursed --------------------------
        # Default to the full computed salary so an unspecified amount behaves
        # exactly as "Mark as Paid" always has. An amount above the computed
        # salary is capped rather than rejected: over-disbursement is a real
        # thing that happens, but recording it here would make balance_due
        # negative and corrupt the "balance to pay" totals.
        _final_dec = Decimal(str(final_salary))
        _amount_paid = amount_by_key.get((emp_type, emp.id), _final_dec)
        if _amount_paid > _final_dec:
            _amount_paid = _final_dec

        # A per-employee split overrides the batch-level method for that
        # employee only; 'mixed' is set automatically once more than one
        # distinct method is actually used (a split with every leg on the
        # same method — e.g. two WPS instalments — stays that single method).
        _splits = splits_by_key.get((emp_type, emp.id))
        if _splits:
            _distinct_methods = {s['method'] for s in _splits}
            _pm = _distinct_methods.pop() if len(_distinct_methods) == 1 else 'mixed'
        else:
            _pm = payment_method

        _defaults = {
            'final_salary': _final_dec,
            'currency': emp.currency,
            'paid_by': request.user.username,
            'paid_at': now_utc,
            'snapshot': snapshot,
            'amount_paid': _amount_paid,
            'payment_method': _pm,
            'payment_date': payment_date or now_utc.date(),
            'payment_splits': _splits,
        }

        if emp_type == 'inhouse':
            PaidSalaryRecord.objects.update_or_create(
                employee_id=emp.id, remote_employee=None, year=year, month=month,
                defaults=_defaults,
            )
        else:
            PaidSalaryRecord.objects.update_or_create(
                remote_employee_id=emp.id, employee=None, year=year, month=month,
                defaults=_defaults,
            )

        _sync_wps_shortfall_advance(
            emp, emp_type, year, month, _final_dec, _amount_paid,
            _pm, request.user.username,
        )

    # Process inhouse employees
    if inhouse_ids:
        all_inhouse_tcr = set(
            Employee.objects.filter(is_active=True)
            .exclude(tcr_id__isnull=True).exclude(tcr_id='')
            .values_list('tcr_id', flat=True)
        )
        for emp in Employee.objects.filter(id__in=inhouse_ids, is_active=True):
            p_start, p_end, p_days, p_hols = _emp_period_mp(emp)
            if emp.department == 'Admin':
                row = _get_inhouse_payroll_row(emp, year, month, p_start, p_end, p_hols, days_in_period=p_days)
            else:
                row = _get_sales_payroll_row(emp, year, month, 'inhouse', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
            cat, carryover_in, total_ded, total_add = _compute_deductions(emp, 'inhouse', p_start, p_end, p_days)
            _save_snapshot(emp, 'inhouse', row, cat, carryover_in, total_ded, total_add, p_start, p_end, p_days, p_hols)
            count += 1

    # Process remote employees
    if remote_ids:
        all_inhouse_tcr = set(
            Employee.objects.filter(is_active=True)
            .exclude(tcr_id__isnull=True).exclude(tcr_id='')
            .values_list('tcr_id', flat=True)
        )
        for emp in RemoteEmployee.objects.filter(id__in=remote_ids, is_active=True):
            p_start, p_end, p_days, p_hols = _emp_period_mp(emp)
            row = _get_sales_payroll_row(emp, year, month, 'remote', banks, p_days, p_hols, period_start=p_start, period_end=p_end)
            cat, carryover_in, total_ded, total_add = _compute_deductions(emp, 'remote', p_start, p_end, p_days)
            _save_snapshot(emp, 'remote', row, cat, carryover_in, total_ded, total_add, p_start, p_end, p_days, p_hols)
            count += 1

    return JsonResponse({'success': True, 'count': count})


@login_required
@user_passes_test(section_required('payroll'))
def unmark_paid_salary(request):
    """Remove the paid lock for selected employees for a given month."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        year = int(data['year'])
        month = int(data['month'])
        employees = data['employees']
    except (KeyError, ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'error': str(e)}, status=400)

    for emp in employees:
        emp_id = int(emp['id'])
        emp_type = emp['type']
        if emp_type == 'inhouse':
            PaidSalaryRecord.objects.filter(employee_id=emp_id, year=year, month=month).delete()
        else:
            PaidSalaryRecord.objects.filter(remote_employee_id=emp_id, year=year, month=month).delete()

    return JsonResponse({'success': True})


@login_required
@user_passes_test(section_required('payroll'))
def add_partial_payment(request):
    """Record an additional installment against an employee already marked
    partially paid for a month — without touching the locked payroll snapshot.

    Unlike mark_paid_salary (which recomputes and overwrites the whole row),
    this only adds to amount_paid and appends to payment_splits, so an earlier
    partial payment's method/date is preserved alongside the new one. Only
    valid against a record that is still partial; a fully-paid month has
    nothing left to add.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        year = int(data['year'])
        month = int(data['month'])
        emp_type = data['employee_type']
        emp_id = int(data['employee_id'])
        raw_splits = data['splits']
    except (KeyError, ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'error': str(e)}, status=400)

    if emp_type not in ('inhouse', 'remote'):
        return JsonResponse({'error': f'Unknown employee type: {emp_type}'}, status=400)

    payment_date = datetime.date.today()
    if data.get('payment_date'):
        try:
            payment_date = datetime.datetime.strptime(data['payment_date'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'error': 'payment_date must be YYYY-MM-DD'}, status=400)

    _valid_methods = {m[0] for m in PAYMENT_METHOD_CHOICES if m[0] != 'mixed'}
    new_splits = []
    new_total = Decimal('0')
    for s in raw_splits:
        method = (s.get('method') or '').strip()
        if method not in _valid_methods:
            return JsonResponse({'error': f'Unknown payment method: {method}'}, status=400)
        try:
            amt = Decimal(str(s.get('amount')))
        except (ArithmeticError, ValueError, TypeError):
            return JsonResponse({'error': 'Invalid payment amount'}, status=400)
        if amt < 0:
            return JsonResponse({'error': 'Paid amount cannot be negative'}, status=400)
        if amt == 0:
            continue
        new_splits.append({'method': method, 'amount': str(amt), 'date': str(payment_date)})
        new_total += amt
    if not new_splits:
        return JsonResponse({'error': 'At least one payment method with an amount is required'}, status=400)

    emp_filter = (
        {'employee_id': emp_id, 'remote_employee': None} if emp_type == 'inhouse'
        else {'remote_employee_id': emp_id, 'employee': None}
    )
    try:
        rec = PaidSalaryRecord.objects.get(year=year, month=month, **emp_filter)
    except PaidSalaryRecord.DoesNotExist:
        return JsonResponse({'error': 'No payment record found for this employee/month — use Mark as Paid first.'}, status=404)

    if not rec.is_partial:
        return JsonResponse({'error': 'This employee is already paid in full for this month.'}, status=400)

    # A record with no split history yet was paid via a single method — fold
    # that original payment in as the first entry so the ledger is complete.
    prior_splits = rec.payment_splits or [{
        'method': rec.payment_method or 'cash',
        'amount': str(rec.effective_amount_paid),
        'date': str(rec.payment_date or (rec.paid_at.date() if rec.paid_at else payment_date)),
    }]
    all_splits = prior_splits + new_splits

    combined_total = rec.effective_amount_paid + new_total
    if combined_total > rec.final_salary:
        combined_total = rec.final_salary

    distinct_methods = {s['method'] for s in all_splits}
    new_pm = distinct_methods.pop() if len(distinct_methods) == 1 else 'mixed'

    rec.amount_paid = combined_total
    rec.payment_method = new_pm
    rec.payment_splits = all_splits
    rec.payment_date = payment_date
    rec.save(update_fields=['amount_paid', 'payment_method', 'payment_splits', 'payment_date'])

    if emp_type == 'inhouse':
        emp = Employee.objects.filter(id=emp_id).first()
    else:
        emp = RemoteEmployee.objects.filter(id=emp_id).first()
    if emp:
        _sync_wps_shortfall_advance(
            emp, emp_type, year, month, rec.final_salary, combined_total,
            new_pm, request.user.username,
        )

    return JsonResponse({
        'success': True,
        'amount_paid': str(combined_total),
        'balance_due': str(rec.balance_due),
        'is_partial': rec.is_partial,
    })
