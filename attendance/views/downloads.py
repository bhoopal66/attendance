"""
Download views for generating XLSX reports.
Handles monthly report downloads for both in-house and remote employees.
"""

import datetime
import calendar
import logging
from datetime import time
from io import BytesIO

from django.db.models import Prefetch
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import (
    AttendanceRecord, Employee, Holiday, LeaveRequest, MonthlySummary,
    RemoteCallRecord, RemoteEmployee,
)
from .utils import (
    MONTH_NAMES, get_active_special_periods_for_month,
    get_employee_shift_for_date, get_holiday_data,
    get_remote_thresholds_from_period, get_saturday_shift,
    get_selected_month_year,
    remote_employee_uses_performance_v2, compute_sales_performance_v2_days,
)
from .reports import _compute_remote_calendar

logger = logging.getLogger('attendance')

# Shared Excel styles
TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
RED_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
HOLIDAY_FILL = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="C7D2FE", end_color="C7D2FE", fill_type="solid")


def _safe_filename(name):
    """Sanitize a string for use in a filename."""
    safe = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
    return safe.replace(' ', '_')


def _build_xlsx_response(wb, filename):
    """Write a workbook to an HttpResponse for download."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()

    response = HttpResponse(
        data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Content-Length'] = len(data)
    return response


def _style_header_row(ws, row_num, num_cols, fill=None):
    """Apply header styles to a row."""
    fill = fill or HEADER_FILL
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _style_data_row(ws, row_num, num_cols, fill=None, center_from=2):
    """Apply border and optional fill to a data row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = THIN_BORDER
        if col >= center_from:
            cell.alignment = Alignment(horizontal='center')
        if fill:
            cell.fill = fill


def _get_approved_leave_dates(employee, month_start, month_end):
    """Get approved leave dates, excluding Sundays and holidays."""
    approved_leaves = LeaveRequest.objects.filter(
        employee=employee,
        status='approved',
        start_date__lte=month_end,
        end_date__gte=month_start
    )

    holidays_set = set(
        Holiday.objects.filter(date__range=(month_start, month_end)).values_list('date', flat=True)
    )

    leave_dates = set()
    for leave in approved_leaves:
        start = max(leave.start_date, month_start)
        end = min(leave.end_date, month_end)
        curr = start
        while curr <= end:
            if curr.weekday() != 6 and curr not in holidays_set:
                leave_dates.add(curr)
            curr += datetime.timedelta(days=1)

    return leave_dates


@login_required
def download_report(request):
    """Generate and download XLSX report for the selected month."""
    selected_month, selected_year = get_selected_month_year(request)
    month_name = MONTH_NAMES[selected_month]

    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    month_start = datetime.date(selected_year, selected_month, 1)
    month_end = datetime.date(selected_year, selected_month, days_in_month)

    show_inactive = request.GET.get('show_inactive', '') == '1'

    summaries = MonthlySummary.objects.filter(
        year=selected_year, month=selected_month
    ).select_related('employee')

    if not show_inactive:
        summaries = summaries.filter(employee__is_active=True)
    summaries = summaries.order_by('employee__name')

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {selected_year}"

    ws.merge_cells('A1:G1')
    ws['A1'] = f"Attendance Report - {month_name} {selected_year}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    headers = [
        'Employee Name', 'Leave Days', 'Half Days', 'Late Arrivals',
        'Late Penalty', 'Paid Leave', 'Total Deductions'
    ]
    ws.append(headers)
    _style_header_row(ws, 3, len(headers))

    for summary in summaries:
        employee = summary.employee
        paid_leave_dates = _get_approved_leave_dates(employee, month_start, month_end)
        paid_leave_count = len(paid_leave_dates)

        half_days = summary.half_days or 0
        late_days = summary.late_days or 0
        late_half_days = late_days // 3
        late_penalty = late_half_days * 0.5
        leave_days = summary.leave_days or 0
        total_deductions = leave_days + (half_days * 0.5) + late_penalty

        ws.append([
            employee.name, leave_days, half_days, late_days,
            late_penalty, paid_leave_count, total_deductions
        ])
        _style_data_row(ws, ws.max_row, 7)

    for col, width in [('A', 30), ('B', 12), ('C', 12), ('D', 14),
                       ('E', 14), ('F', 12), ('G', 18)]:
        ws.column_dimensions[col].width = width

    filename = f"Attendance_Report_{selected_year}_{selected_month:02d}.xlsx"
    return _build_xlsx_response(wb, filename)


@login_required
def download_employee_report(request, employee_id):
    """Generate and download XLSX report for a single employee for the selected month."""
    employee = get_object_or_404(Employee, id=employee_id)
    selected_month, selected_year = get_selected_month_year(request)
    month_name = MONTH_NAMES[selected_month]

    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    month_start = datetime.date(selected_year, selected_month, 1)
    month_end = datetime.date(selected_year, selected_month, days_in_month)

    holiday_dates, _, _ = get_holiday_data(month_start, month_end)

    records = AttendanceRecord.objects.filter(
        employee=employee,
        date__year=selected_year,
        date__month=selected_month
    ).order_by('date')
    records_dict = {r.date: r for r in records}

    emp_shift_start, emp_shift_end = get_employee_shift_for_date(employee, month_start)
    sat_shift_start, sat_shift_end = get_saturday_shift()

    approved_leave_dates = set()
    for leave in LeaveRequest.objects.filter(
        employee=employee, status='approved',
        start_date__lte=month_end, end_date__gte=month_start
    ):
        start = max(leave.start_date, month_start)
        end = min(leave.end_date, month_end)
        curr = start
        while curr <= end:
            approved_leave_dates.add(curr)
            curr += datetime.timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = employee.name[:20]

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Attendance Report - {month_name} {selected_year}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Employee: {employee.name} (ID: {employee.person_id})"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])

    headers = ['Date', 'Day', 'First In', 'Last Out', 'Duration', 'Status']
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    full_days = half_days = leave_days = paid_leave_days = late_arrivals = holidays_count = 0
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    for day in range(1, days_in_month + 1):
        date = datetime.date(selected_year, selected_month, day)
        weekday = date.weekday()
        day_name = day_names[weekday]
        is_sunday = weekday == 6
        is_holiday = date in holiday_dates
        is_saturday = weekday == 5

        record = records_dict.get(date)

        if is_sunday or is_holiday:
            holidays_count += 1
            status = "Holiday"
            first_in = last_out = duration = "-"
            fill = HOLIDAY_FILL
        elif record:
            first_in = record.first_in.strftime("%H:%M") if record.first_in else "-"
            last_out = record.last_out.strftime("%H:%M") if record.last_out else "-"
            duration = str(record.work_duration) if record.work_duration else "-"

            total_secs = record.work_duration.total_seconds() if record.work_duration else 0
            arrived_after_noon = record.first_in and record.first_in.hour >= 12 and not is_saturday

            # Compare using hour:minute only (ignore seconds)
            fi_hm = (record.first_in.hour, record.first_in.minute) if record.first_in else None
            if is_saturday:
                is_late = fi_hm is not None and fi_hm > (sat_shift_start.hour, sat_shift_start.minute)
                left_early = record.last_out and (
                    record.last_out.hour < sat_shift_end.hour or
                    (record.last_out.hour == sat_shift_end.hour and
                     record.last_out.minute < sat_shift_end.minute)
                )
            else:
                is_late = fi_hm is not None and fi_hm > (emp_shift_start.hour, emp_shift_start.minute)
                left_early = record.last_out and (
                    record.last_out.hour < emp_shift_end.hour or
                    (record.last_out.hour == emp_shift_end.hour and
                     record.last_out.minute < emp_shift_end.minute)
                )

            is_half_day = arrived_after_noon

            if total_secs == 0:
                if date in approved_leave_dates and not is_sunday and not is_holiday:
                    status = "Paid Leave"
                    fill = BLUE_FILL
                    paid_leave_days += 1
                else:
                    status = "Leave"
                    fill = RED_FILL
                    leave_days += 1
            elif is_half_day:
                status = "Half Day"
                fill = YELLOW_FILL
                half_days += 1
                if is_late:
                    late_arrivals += 1
            elif is_late:
                status = "Late"
                fill = YELLOW_FILL
                full_days += 1
                late_arrivals += 1
            else:
                status = "Present"
                fill = GREEN_FILL
                full_days += 1
        else:
            first_in = last_out = duration = "-"
            if date <= datetime.date.today():
                if date in approved_leave_dates and not is_sunday and not is_holiday:
                    status = "Paid Leave"
                    fill = BLUE_FILL
                    paid_leave_days += 1
                else:
                    status = "Leave"
                    fill = RED_FILL
                    leave_days += 1
            else:
                status = "-"
                fill = None

        ws.append([date.strftime("%Y-%m-%d"), day_name, first_in, last_out, duration, status])
        _style_data_row(ws, ws.max_row, 6, fill=fill, center_from=1)

    ws.append([])
    ws.append([])
    summary_row = ws.max_row + 1
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Monthly Summary"
    ws.cell(row=summary_row, column=1).font = TITLE_FONT
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')

    late_half_days = late_arrivals // 3
    late_penalty = late_half_days * 0.5
    total_deductions = leave_days + (half_days * 0.5) + late_penalty

    for label, value in [
        ("Leave Days", leave_days),
        ("Half Days", half_days),
        ("Late Arrivals", f"{late_arrivals} ({late_half_days} x 0.5 penalty)" if late_half_days > 0 else late_arrivals),
        ("Paid Leave Days", paid_leave_days),
        ("Holidays/Sundays", holidays_count),
        ("Full Days", full_days),
        ("Total Deductions", total_deductions),
    ]:
        ws.append([label, value])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    for col, width in [('A', 14), ('B', 8), ('C', 10), ('D', 10), ('E', 12), ('F', 12)]:
        ws.column_dimensions[col].width = width

    safe_name = _safe_filename(employee.name)
    filename = f"{safe_name}_Attendance_{selected_year}_{selected_month:02d}.xlsx"
    return _build_xlsx_response(wb, filename)


@login_required
def download_remote_report(request):
    """Generate and download XLSX report for remote team.

    Computed live via _compute_remote_calendar (the same function backing the
    on-screen report), rather than the RemoteMonthlySummary table, which is
    only refreshed by the manually-run recalculate_summaries command and
    never applies special-shift-period thresholds or the Sales:Performance
    "Method 2" model — so it can silently disagree with both the calendar and
    payroll. Computing live keeps this download exactly in sync with both.
    """
    selected_month, selected_year = get_selected_month_year(request)
    month_name = MONTH_NAMES[selected_month]

    show_inactive = request.GET.get('show_inactive', '') == '1'

    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    month_start = datetime.date(selected_year, selected_month, 1)
    month_end = datetime.date(selected_year, selected_month, days_in_month)
    holiday_dates, _, _ = get_holiday_data(month_start, month_end)
    special_periods = get_active_special_periods_for_month(month_start, month_end)

    today = datetime.date.today()
    is_current_month = selected_year == today.year and selected_month == today.month
    current_day = today.day if is_current_month else 32

    records_qs = RemoteCallRecord.objects.filter(
        date__year=selected_year, date__month=selected_month
    ).order_by('date')
    remote_qs = RemoteEmployee.objects.prefetch_related(
        Prefetch('remotecallrecord_set', queryset=records_qs, to_attr='filtered_records')
    )
    if not show_inactive:
        remote_qs = remote_qs.filter(is_active=True)
    employees = list(remote_qs.order_by('name'))

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {selected_year}"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Remote Team Attendance Report - {month_name} {selected_year}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    headers = ['Employee Name', 'Present Days', 'Partial Days', 'Half Days', 'Absent Days', 'Total Deductions']
    ws.append(headers)
    _style_header_row(ws, 3, len(headers), fill=HEADER_FILL_PURPLE)

    for emp in employees:
        _, stats = _compute_remote_calendar(
            emp, days_in_month, selected_year, selected_month,
            holiday_dates, current_day, special_periods=special_periods,
        )
        present_days = stats['present_days']
        proportional_days = stats.get('proportional_days', 0)
        half_days = stats['half_days']
        absent_days = stats['absent_days']
        total_deductions = absent_days + (half_days * 0.5) + (proportional_days * 0.5)

        ws.append([emp.name, present_days, proportional_days, half_days, absent_days, total_deductions])
        _style_data_row(ws, ws.max_row, 6)

    for col, width in [('A', 30), ('B', 14), ('C', 13), ('D', 12), ('E', 14), ('F', 18)]:
        ws.column_dimensions[col].width = width

    filename = f"Remote_Attendance_Report_{selected_year}_{selected_month:02d}.xlsx"
    return _build_xlsx_response(wb, filename)


@login_required
def download_remote_employee_report(request, employee_id):
    """Generate and download XLSX report for a single remote employee."""
    employee = get_object_or_404(RemoteEmployee, id=employee_id)
    selected_month, selected_year = get_selected_month_year(request)
    month_name = MONTH_NAMES[selected_month]

    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    month_start = datetime.date(selected_year, selected_month, 1)
    month_end = datetime.date(selected_year, selected_month, days_in_month)

    holiday_dates, _, _ = get_holiday_data(month_start, month_end)
    special_periods = get_active_special_periods_for_month(month_start, month_end)

    records = RemoteCallRecord.objects.filter(
        employee=employee,
        date__year=selected_year,
        date__month=selected_month
    ).order_by('date')
    records_dict = {r.date: r for r in records}

    wb = Workbook()
    ws = wb.active
    ws.title = employee.name[:20]

    ws.merge_cells('A1:E1')
    ws['A1'] = f"Remote Call Statistics - {month_name} {selected_year}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Employee: {employee.name} (Extension: {employee.extension_id})"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])

    headers = ['Date', 'Day', 'Answered Calls', 'Talk Duration', 'Status']
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    present_days = half_days = proportional_days = absent_days = holidays_count = total_calls = total_talk_minutes = 0
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    # Sales:Performance remote employees (attendance-based, not fixed-salary,
    # salaried) whose pay is governed by the "Method 2" talktime-proportional
    # model use the same shared classifier as the calendar/payroll, so this
    # download never disagrees with what the employee was actually paid.
    v2_days = None
    if remote_employee_uses_performance_v2(employee, selected_year, selected_month):
        v2_days = compute_sales_performance_v2_days(employee, month_start, month_end, holiday_dates)['days']

    for day in range(1, days_in_month + 1):
        date = datetime.date(selected_year, selected_month, day)
        weekday = date.weekday()
        day_name = day_names[weekday]
        is_sunday = weekday == 6
        is_holiday = date in holiday_dates

        record = records_dict.get(date)

        if v2_days is not None:
            day_info = v2_days.get(date)
            answered_calls = record.answered_calls or 0 if record else "-"
            talk_min = int(record.total_talk_duration.total_seconds() / 60) if (record and record.total_talk_duration) else 0
            talk_duration = f"{talk_min} min" if record else "-"
            if record and record.total_talk_duration:
                total_calls += answered_calls
                total_talk_minutes += talk_min

            if day_info and day_info['classification'] == 'non_working':
                holidays_count += 1
                status = "Holiday"
                fill = HOLIDAY_FILL
            elif day_info and (record or date <= datetime.date.today()):
                classification = day_info['classification']
                if classification == 'full':
                    status = "Present"
                    fill = GREEN_FILL
                    present_days += 1
                elif classification == 'proportional':
                    status = "Partial"
                    fill = PARTIAL_FILL
                    proportional_days += 1
                elif classification == 'half':
                    status = "Half Day"
                    fill = YELLOW_FILL
                    half_days += 1
                else:
                    status = "Absent" if record else "No Data"
                    fill = RED_FILL
                    absent_days += 1
            else:
                status = "-"
                fill = None

            ws.append([date.strftime("%Y-%m-%d"), day_name, answered_calls, talk_duration, status])
            _style_data_row(ws, ws.max_row, 5, fill=fill, center_from=1)
            continue

        if is_sunday or is_holiday:
            holidays_count += 1
            status = "Holiday"
            answered_calls = "-"
            talk_duration = "-"
            fill = HOLIDAY_FILL
        elif record:
            answered_calls = record.answered_calls or 0
            talk_min = int(record.total_talk_duration.total_seconds() / 60) if record.total_talk_duration else 0
            talk_duration = f"{talk_min} min"
            total_calls += answered_calls
            total_talk_minutes += talk_min

            # Check for special shift period with remote thresholds
            active_period = None
            if special_periods:
                active_period = next(
                    (p for p in special_periods if p.start_date <= date <= p.end_date),
                    None
                )
            remote_thresholds = get_remote_thresholds_from_period(active_period) if active_period else None
            att_status = record.calculate_attendance_status(thresholds=remote_thresholds) if remote_thresholds else record.attendance_status

            if att_status == 'present':
                status = "Present"
                fill = GREEN_FILL
                present_days += 1
            elif att_status == 'half_day':
                status = "Half Day"
                fill = YELLOW_FILL
                half_days += 1
            else:
                status = "Absent"
                fill = RED_FILL
                absent_days += 1
        else:
            answered_calls = "-"
            talk_duration = "-"
            if date <= datetime.date.today():
                status = "No Data"
                fill = RED_FILL
                absent_days += 1
            else:
                status = "-"
                fill = None

        ws.append([date.strftime("%Y-%m-%d"), day_name, answered_calls, talk_duration, status])
        _style_data_row(ws, ws.max_row, 5, fill=fill, center_from=1)

    ws.append([])
    ws.append([])
    summary_row = ws.max_row + 1
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Monthly Summary"
    ws.cell(row=summary_row, column=1).font = TITLE_FONT
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')

    total_deductions = absent_days + (half_days * 0.5) + (proportional_days * 0.5)

    summary_rows = [("Absent Days", absent_days), ("Half Days", half_days)]
    if v2_days is not None:
        summary_rows.append(("Partial Days", proportional_days))
    summary_rows += [
        ("Present Days", present_days),
        ("Holidays/Sundays", holidays_count),
        ("Total Calls Answered", total_calls),
        ("Total Talk Time", f"{total_talk_minutes} min"),
        ("Total Deductions", total_deductions),
    ]

    for label, value in summary_rows:
        ws.append([label, value])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    for col, width in [('A', 14), ('B', 8), ('C', 15), ('D', 15), ('E', 12)]:
        ws.column_dimensions[col].width = width

    safe_name = _safe_filename(employee.name)
    filename = f"{safe_name}_Remote_Stats_{selected_year}_{selected_month:02d}.xlsx"
    return _build_xlsx_response(wb, filename)
